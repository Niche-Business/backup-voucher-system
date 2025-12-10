from flask import Flask, request, jsonify, session
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import text
from flask_cors import CORS
from flask_mail import Mail, Message
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_compress import Compress
from flask_socketio import SocketIO
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import os
import secrets
from email_service import email_service
from charity_verification import verify_charity_number
from sms_service import sms_service
import stripe_payment
from wallet_blueprint import wallet_bp, init_wallet_blueprint
from admin_enhancements import init_admin_enhancements
from vcse_verification import init_vcse_verification

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'vcse-charity-platform-secret-key-2024')

# Database configuration - PostgreSQL for production, SQLite for local development
DATABASE_URL = os.environ.get('DATABASE_URL')
if DATABASE_URL:
    # Render provides DATABASE_URL, but we need to replace postgres:// with postgresql://
    if DATABASE_URL.startswith('postgres://'):
        DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
else:
    # Fallback to SQLite for local development
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///vcse_charity.db'

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_recycle': 300,
}

# Email configuration (using environment variables with fallback to demo mode)
app.config['MAIL_SERVER'] = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.environ.get('MAIL_PORT', 587))
app.config['MAIL_USE_TLS'] = os.environ.get('MAIL_USE_TLS', 'True') == 'True'
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME', 'noreply@bakup.org')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD', '')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_DEFAULT_SENDER', 'BAK UP E-Voucher <noreply@bakup.org>')
app.config['MAIL_SUPPRESS_SEND'] = os.environ.get('MAIL_SUPPRESS_SEND', 'True') == 'True'  # Set to True to disable emails in dev (default: True)

db = SQLAlchemy(app)
CORS(app, supports_credentials=True, origins=['*'])
mail = Mail(app)
socketio = SocketIO(app, cors_allowed_origins='*', manage_session=False)

# Initialize Flask-Compress for Gzip compression
Compress(app)

# Initialize Flask-Limiter for rate limiting
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://"
)

# Enhanced Database Models
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    first_name = db.Column(db.String(50), nullable=False)
    last_name = db.Column(db.String(50), nullable=False)
    phone = db.Column(db.String(20))
    user_type = db.Column(db.String(20), nullable=False)  # recipient, vendor, vcse, admin
    organization_name = db.Column(db.String(100))
    shop_name = db.Column(db.String(100))
    address = db.Column(db.Text)
    postcode = db.Column(db.String(10))
    city = db.Column(db.String(50))
    # date_of_birth = db.Column(db.Date)  # Recipient date of birth - COMMENTED OUT temporarily
    charity_commission_number = db.Column(db.String(50))
    shop_category = db.Column(db.String(50))  # African, Caribbean, Mixed African & Caribbean, Indian/South Asian, Eastern European, Middle Eastern
    is_verified = db.Column(db.Boolean, default=False)
    verification_token = db.Column(db.String(100))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    last_login = db.Column(db.DateTime)
    login_count = db.Column(db.Integer, default=0)
    is_active = db.Column(db.Boolean, default=True)
    account_status = db.Column(db.String(30), default='ACTIVE')  # ACTIVE, PENDING_VERIFICATION, REJECTED
    rejection_reason = db.Column(db.Text)  # Reason for rejection (if rejected)
    verified_at = db.Column(db.DateTime)  # When account was verified by admin
    verified_by_admin_id = db.Column(db.Integer, db.ForeignKey('user.id'))  # Admin who verified
    balance = db.Column(db.Float, default=0.0)  # For VCFSE organizations to load money
    allocated_balance = db.Column(db.Float, default=0.0)  # Funds allocated by admin to VCFSE
    
    # Food To Go preferred shop for recipients
    preferred_shop_id = db.Column(db.Integer, db.ForeignKey('vendor_shop.id'))  # Recipient's preferred shop

class VendorShop(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    vendor_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    shop_name = db.Column(db.String(100), nullable=False)
    address = db.Column(db.Text, nullable=False)
    postcode = db.Column(db.String(10))
    city = db.Column(db.String(50))
    town = db.Column(db.String(50))  # Specific town: Wellingborough, Kettering, Corby, Northampton, Daventry, Brackley, Towcester
    phone = db.Column(db.String(20))
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    vendor = db.relationship('User', foreign_keys=[vendor_id], backref='shops')

class Voucher(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), unique=True, nullable=False)
    value = db.Column(db.Float, nullable=False)
    recipient_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    issued_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)  # Admin or VCFSE
    vendor_restrictions = db.Column(db.Text)  # JSON list of allowed vendor IDs
    expiry_date = db.Column(db.Date, nullable=False)
    status = db.Column(db.String(20), default='active')  # active, redeemed, expired, reassigned
    redeemed_at = db.Column(db.DateTime)
    redeemed_by_vendor = db.Column(db.Integer, db.ForeignKey('user.id'))
    redeemed_at_shop_id = db.Column(db.Integer, db.ForeignKey('vendor_shop.id'))  # Shop where voucher was redeemed
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    reassignment_count = db.Column(db.Integer, default=0)  # Track number of reassignments
    reassignment_history = db.Column(db.Text)  # JSON array of reassignment records
    original_recipient_id = db.Column(db.Integer, db.ForeignKey('user.id'))  # Original recipient
    
    # Food To Go shop assignment fields
    assign_shop_method = db.Column(db.String(50), default='specific_shop')  # 'specific_shop' or 'recipient_to_choose'
    recipient_selected_shop_id = db.Column(db.Integer, db.ForeignKey('vendor_shop.id'))  # Shop selected by recipient
    
    # Wallet integration fields
    issued_by_user_id = db.Column(db.Integer, db.ForeignKey('user.id'))  # School/VCFSE who issued the voucher
    deducted_from_wallet = db.Column(db.Boolean, default=False)  # Whether voucher value was deducted from issuer's wallet
    wallet_transaction_id = db.Column(db.Integer, db.ForeignKey('wallet_transaction.id'))  # Link to wallet transaction
    
    recipient = db.relationship('User', foreign_keys=[recipient_id], backref='received_vouchers')
    issuer = db.relationship('User', foreign_keys=[issued_by], backref='issued_vouchers')
    redeemer = db.relationship('User', foreign_keys=[redeemed_by_vendor], backref='redeemed_vouchers')
    original_recipient = db.relationship('User', foreign_keys=[original_recipient_id])
    issued_by_user = db.relationship('User', foreign_keys=[issued_by_user_id], backref='wallet_issued_vouchers')
    wallet_transaction = db.relationship('WalletTransaction', foreign_keys=[wallet_transaction_id])

class Category(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False)
    type = db.Column(db.String(20), nullable=False)  # edible, non-edible
    icon = db.Column(db.String(10), default='📦')
    description = db.Column(db.Text)

class Item(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(100), nullable=False)
    description = db.Column(db.Text)
    category_id = db.Column(db.Integer, db.ForeignKey('category.id'), nullable=False)
    vendor_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    quantity = db.Column(db.Float, nullable=False)
    unit = db.Column(db.String(20), nullable=False)
    expiry_date = db.Column(db.Date)
    pickup_location = db.Column(db.String(200))
    pickup_instructions = db.Column(db.Text)
    collection_time_limit = db.Column(db.DateTime)  # Time limit for collection
    status = db.Column(db.String(20), default='available')  # available, claimed, collected, expired
    claimed_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    claimed_at = db.Column(db.DateTime)
    collected_at = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    category = db.relationship('Category', backref='items')
    vendor = db.relationship('User', foreign_keys=[vendor_id], backref='listed_items')
    claimer = db.relationship('User', foreign_keys=[claimed_by], backref='claimed_items')

class UserNotification(db.Model):
    __tablename__ = 'user_notifications'
    
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    title = db.Column(db.String(100), nullable=False)
    message = db.Column(db.Text, nullable=False)
    type = db.Column(db.String(20), default='info')  # info, success, warning, error
    is_read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    user = db.relationship('User', backref='user_notifications')

class LoginSession(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    login_time = db.Column(db.DateTime, default=datetime.utcnow)
    logout_time = db.Column(db.DateTime)
    session_duration = db.Column(db.Integer)  # in minutes
    ip_address = db.Column(db.String(45))
    user_agent = db.Column(db.Text)
    
    user = db.relationship('User', backref='login_sessions')

class PasswordResetToken(db.Model):
    __tablename__ = 'password_reset_token'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    token = db.Column(db.String(255), unique=True, nullable=False)
    expires_at = db.Column(db.DateTime, nullable=False)
    used = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    user = db.relationship('User', backref='password_reset_tokens')

class Report(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    generated_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    report_type = db.Column(db.String(50), nullable=False)  # admin_full, vcse_specific
    date_from = db.Column(db.Date, nullable=False)
    date_to = db.Column(db.Date, nullable=False)
    data = db.Column(db.Text)  # JSON data
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    generator = db.relationship('User', backref='generated_reports')

class SurplusItem(db.Model):
    __tablename__ = 'surplus_item'
    id = db.Column(db.Integer, primary_key=True)
    vendor_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    shop_id = db.Column(db.Integer, db.ForeignKey('vendor_shop.id'), nullable=False)
    item_name = db.Column(db.String(200), nullable=False)
    quantity = db.Column(db.String(100), nullable=False)
    unit = db.Column(db.String(50))  # kg, liters, pieces, etc.
    category = db.Column(db.String(50), nullable=False)  # edible, non-edible
    item_type = db.Column(db.String(20), default='free')  # 'discount' or 'free'
    price = db.Column(db.Numeric(10, 2))  # Price for discount items, NULL for free items
    original_price = db.Column(db.Numeric(10, 2))  # Original price before discount
    description = db.Column(db.Text)
    expiry_date = db.Column(db.Date)  # Product expiry/best before date
    status = db.Column(db.String(20), default='available')  # available, claimed, collected
    posted_at = db.Column(db.DateTime, default=datetime.utcnow)
    claimed_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    claimed_at = db.Column(db.DateTime)
    collected_at = db.Column(db.DateTime)
    
    vendor = db.relationship('User', foreign_keys=[vendor_id], backref='surplus_posted_items')
    shop = db.relationship('VendorShop', backref='shop_surplus_items')
    claimer = db.relationship('User', foreign_keys=[claimed_by], backref='surplus_claimed_items')

class ShoppingCart(db.Model):
    __tablename__ = 'shopping_cart'
    id = db.Column(db.Integer, primary_key=True)
    recipient_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    surplus_item_id = db.Column(db.Integer, db.ForeignKey('surplus_item.id'), nullable=False)
    quantity = db.Column(db.Integer, default=1, nullable=False)
    added_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    recipient = db.relationship('User', backref='cart_items')
    surplus_item = db.relationship('SurplusItem', backref='in_carts')

class Order(db.Model):
    __tablename__ = 'order'
    id = db.Column(db.Integer, primary_key=True)
    vcse_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)  # VCFSE organization placing order
    surplus_item_id = db.Column(db.Integer, db.ForeignKey('surplus_item.id'), nullable=False)
    client_name = db.Column(db.String(200), nullable=False)  # Client receiving the order
    client_mobile = db.Column(db.String(20), nullable=False)
    client_email = db.Column(db.String(120), nullable=False)
    quantity = db.Column(db.Integer, default=1, nullable=False)
    status = db.Column(db.String(20), default='pending')  # pending, confirmed, collected, cancelled
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    vcse = db.relationship('User', backref='placed_orders')
    surplus_item = db.relationship('SurplusItem', backref='orders')

class PayoutRequest(db.Model):
    __tablename__ = 'payout_request'
    id = db.Column(db.Integer, primary_key=True)
    vendor_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)  # Shop requesting payout
    shop_id = db.Column(db.Integer, db.ForeignKey('vendor_shop.id'), nullable=False)  # Specific shop
    amount = db.Column(db.Float, nullable=False)  # Amount requested
    status = db.Column(db.String(20), default='pending')  # pending, approved, rejected, paid
    bank_name = db.Column(db.String(100))  # Bank details
    account_number = db.Column(db.String(50))
    sort_code = db.Column(db.String(20))
    account_holder_name = db.Column(db.String(100))
    notes = db.Column(db.Text)  # Additional notes from vendor
    admin_notes = db.Column(db.Text)  # Notes from admin
    requested_at = db.Column(db.DateTime, default=datetime.utcnow)
    reviewed_at = db.Column(db.DateTime)  # When admin reviewed
    reviewed_by = db.Column(db.Integer, db.ForeignKey('user.id'))  # Admin who reviewed
    paid_at = db.Column(db.DateTime)  # When payment was made
    
    vendor = db.relationship('User', foreign_keys=[vendor_id], backref='payout_requests')
    shop = db.relationship('VendorShop', backref='payouts')
    reviewer = db.relationship('User', foreign_keys=[reviewed_by], backref='reviewed_payouts')

class WalletTransaction(db.Model):
    """Wallet transactions for schools and VCFSEs"""
    __tablename__ = 'wallet_transaction'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    transaction_type = db.Column(db.String(20), nullable=False)  # credit, debit, allocation
    amount = db.Column(db.Float, nullable=False)
    balance_before = db.Column(db.Float, nullable=False)
    balance_after = db.Column(db.Float, nullable=False)
    description = db.Column(db.Text)
    reference = db.Column(db.String(100))  # Payment reference, voucher code, etc.
    payment_method = db.Column(db.String(50))  # stripe, bank_transfer, admin_allocation, manual
    payment_reference = db.Column(db.String(100))  # Stripe payment ID, etc.
    status = db.Column(db.String(20), default='completed')  # pending, completed, failed
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))  # Admin who made allocation, or self
    
    # Relationships
    user = db.relationship('User', foreign_keys=[user_id], backref='wallet_transactions')
    creator = db.relationship('User', foreign_keys=[created_by])

class PaymentTransaction(db.Model):
    """Stripe payment transactions for VCFSE fund loading"""
    __tablename__ = 'payment_transaction'
    id = db.Column(db.Integer, primary_key=True)
    
    # User Information
    vcse_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    
    # Payment Details
    amount = db.Column(db.Float, nullable=False)  # Amount in GBP
    currency = db.Column(db.String(3), default='GBP')
    
    # Stripe Information
    stripe_payment_intent_id = db.Column(db.String(255), unique=True)
    stripe_customer_id = db.Column(db.String(255))  # For saved payment methods
    payment_method_id = db.Column(db.String(255))  # Card used
    
    # Status Tracking
    status = db.Column(db.String(20), default='pending')
    # Status values: pending, processing, succeeded, failed, cancelled, refunded
    
    # Transaction Metadata
    description = db.Column(db.Text)
    failure_reason = db.Column(db.Text)  # Error message if failed
    
    # Timestamps
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    completed_at = db.Column(db.DateTime)  # When payment succeeded
    
    # Relationship
    vcse = db.relationship('User', backref='payment_transactions')

# Notification Models
class Notification(db.Model):
    __tablename__ = 'notifications'
    
    id = db.Column(db.Integer, primary_key=True)
    type = db.Column(db.String(50), nullable=False)  # 'discounted_item' or 'free_item'
    shop_id = db.Column(db.Integer, db.ForeignKey('vendor_shop.id'), nullable=False)
    item_id = db.Column(db.Integer)  # Reference to SurplusItem
    target_group = db.Column(db.String(50), nullable=False)  # 'recipient', 'vcse', 'school', 'all'
    message = db.Column(db.Text, nullable=False)
    item_name = db.Column(db.String(200))
    shop_name = db.Column(db.String(200))
    quantity = db.Column(db.String(50))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_read = db.Column(db.Boolean, default=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'))  # For user-specific notifications
    
    def to_dict(self):
        return {
            'id': self.id,
            'type': self.type,
            'shop_id': self.shop_id,
            'item_id': self.item_id,
            'target_group': self.target_group,
            'message': self.message,
            'item_name': self.item_name,
            'shop_name': self.shop_name,
            'quantity': self.quantity,
            'created_at': self.created_at.isoformat() if self.created_at else None,
            'is_read': self.is_read
        }

class NotificationPreference(db.Model):
    __tablename__ = 'notification_preferences'
    
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), unique=True, nullable=False)
    sound_enabled = db.Column(db.Boolean, default=True)
    email_enabled = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    def to_dict(self):
        return {
            'user_id': self.user_id,
            'sound_enabled': self.sound_enabled,
            'email_enabled': self.email_enabled
        }

# Initialize and register wallet blueprint
init_wallet_blueprint(db, User, Voucher, WalletTransaction)
app.register_blueprint(wallet_bp)

# Initialize admin enhancement endpoints
# Note: Transaction model doesn't exist yet, so we pass None for now
init_admin_enhancements(app, db, User, VendorShop, Voucher, None, email_service)

# Initialize VCFSE Verification System
init_vcse_verification(app, db, User, email_service)

# Initialize Notifications System
from notifications_system import notifications_bp, init_socketio, init_notifications_system
init_notifications_system(db, Notification, NotificationPreference, User, socketio)
app.register_blueprint(notifications_bp)
init_socketio(socketio)

# Initialize Analytics Dashboard
from analytics_dashboard import analytics_bp
app.register_blueprint(analytics_bp)

# Initialize Expiration Reminders System
from expiration_reminders import expiration_bp, init_expiration_reminders
init_expiration_reminders(db, Voucher, User, email_service)
app.register_blueprint(expiration_bp)

# Initialize Export System
from export_reports import export_bp, init_export_system
init_export_system(db, User, Voucher, SurplusItem, WalletTransaction)
app.register_blueprint(export_bp)

# Initialize Audit Log System
from audit_log import audit_bp, init_audit_system, log_activity
init_audit_system(app, db, User)
app.register_blueprint(audit_bp)

# Initialize Security Enhancements
from security_enhancements import init_security_enhancements, rate_limit, csrf_protect
init_security_enhancements(app)

# Initialize Bulk Import System
from bulk_import import bulk_import_bp, init_bulk_import
init_bulk_import(db, User, Voucher, email_service)
app.register_blueprint(bulk_import_bp)

# Initialize Vendor Metrics System
from vendor_metrics import vendor_metrics_bp, init_vendor_metrics
init_vendor_metrics(db, User, Voucher, SurplusItem, WalletTransaction, VendorShop)
app.register_blueprint(vendor_metrics_bp)

# Initialize Bulk Upload System for VCFSE/Schools
from bulk_upload import bulk_upload_bp
app.register_blueprint(bulk_upload_bp)

# Initialize notifications migration endpoint
from migrate_notifications import create_notifications_migration_endpoint
create_notifications_migration_endpoint(app, db)

# Helper Functions
def send_verification_email(user):
    try:
        msg = Message(
            'Verify Your BAK UP Account',
            sender=app.config['MAIL_DEFAULT_SENDER'],
            recipients=[user.email]
        )
        msg.body = f'''
Hello {user.first_name},

Welcome to BAK UP E-Voucher System! Please verify your account by clicking the link below:

http://localhost:5000/api/verify/{user.verification_token}

Thank you for joining our mission to reduce waste and support communities!

Best regards,
BAK UP Team
        '''
        if not app.config['MAIL_SUPPRESS_SEND']:
            mail.send(msg)
        return True
    except Exception as e:
        print(f"Email sending failed: {e}")
        return False

def send_welcome_email(user):
    """Send welcome email to new users"""
    try:
        user_type_names = {
            'recipient': 'Voucher Recipient',
            'vendor': 'Shop Vendor',
            'vcse': 'VCFSE Organization',
            'admin': 'System Administrator'
        }
        
        msg = Message(
            'Welcome to BAK UP E-Voucher System!',
            sender=app.config['MAIL_DEFAULT_SENDER'],
            recipients=[user.email]
        )
        
        msg.html = f'''
<!DOCTYPE html>
<html>
<head>
    <style>
        body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
        .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
        .header {{ background: linear-gradient(135deg, #4CAF50 0%, #2d5a27 100%); color: white; padding: 30px; text-align: center; border-radius: 10px 10px 0 0; }}
        .content {{ background: #f8f9fa; padding: 30px; border-radius: 0 0 10px 10px; }}
        .button {{ background: #4CAF50; color: white; padding: 12px 30px; text-decoration: none; border-radius: 5px; display: inline-block; margin: 20px 0; }}
        .footer {{ text-align: center; margin-top: 30px; color: #666; font-size: 12px; }}
        h1 {{ margin: 0; font-size: 28px; }}
        h2 {{ color: #2d5a27; }}
        .feature {{ background: white; padding: 15px; margin: 10px 0; border-left: 4px solid #4CAF50; border-radius: 5px; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🎫 Welcome to BAK UP!</h1>
            <p>Reducing waste, supporting communities, strengthening local economies</p>
        </div>
        <div class="content">
            <h2>Hello {user.first_name}!</h2>
            <p>Thank you for joining the BAK UP E-Voucher System as a <strong>{user_type_names.get(user.user_type, user.user_type)}</strong>.</p>
            
            <p>Your account has been successfully created and is ready to use!</p>
            
            <div class="feature">
                <strong>📧 Your Email:</strong> {user.email}
            </div>
            
            <h2>What's Next?</h2>
            <p>You can now log in to your account and start using the platform:</p>
            
            <a href="https://8080-ifslfm1mlvuoydk0g6mnh-090f122e.manusvm.computer" class="button">Access Your Dashboard</a>
            
            <h2>Need Help?</h2>
            <p>If you have any questions or need assistance, please don't hesitate to contact our support team.</p>
            
            <p>Together, we can make a difference in reducing food waste and supporting our community!</p>
            
            <p>Best regards,<br>
            <strong>The BAK UP Team</strong></p>
        </div>
        <div class="footer">
            <p>BAK UP CIC - Northamptonshire Community E-Voucher Scheme</p>
            <p>This is an automated message. Please do not reply to this email.</p>
        </div>
    </div>
</body>
</html>
        '''
        
        msg.body = f'''
Hello {user.first_name}!

Thank you for joining the BAK UP E-Voucher System as a {user_type_names.get(user.user_type, user.user_type)}.

Your account has been successfully created and is ready to use!

Email: {user.email}

You can now log in to your account and start using the platform.

If you have any questions or need assistance, please don't hesitate to contact our support team.

Together, we can make a difference in reducing food waste and supporting our community!

Best regards,
The BAK UP Team

BAK UP CIC - Northamptonshire Community E-Voucher Scheme
        '''
        
        if not app.config['MAIL_SUPPRESS_SEND']:
            mail.send(msg)
            print(f"Welcome email sent to {user.email}")
        else:
            print(f"Email suppressed (dev mode): Would send welcome email to {user.email}")
        return True
    except Exception as e:
        print(f"Welcome email sending failed: {e}")
        return False

def create_notification(user_id, title, message, notification_type='info'):
    notification = UserNotification(
        user_id=user_id,
        title=title,
        message=message,
        type=notification_type
    )
    db.session.add(notification)
    db.session.commit()

def generate_voucher_code():
    return f"BAK{secrets.token_hex(4).upper()}"

def track_login(user_id, ip_address=None, user_agent=None):
    user = User.query.get(user_id)
    if user:
        user.last_login = datetime.utcnow()
        user.login_count += 1
        
        login_session = LoginSession(
            user_id=user_id,
            ip_address=ip_address,
            user_agent=user_agent
        )
        db.session.add(login_session)
        db.session.commit()

# API Routes
@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({
        'status': 'healthy',
        'message': 'BAK UP E-Voucher System API is running',
        'features': [
            'User Registration with Password Security',
            'Multi-Shop Vendor Support',
            'VCFSE Money Loading & Voucher Issuing',
            'Login Frequency Tracking',
            'Voucher Reassignment',
            'Comprehensive Reporting System',
            'Real-time Surplus Food Notifications',
            'Mobile PWA Support'
        ],
        'timestamp': datetime.utcnow().isoformat()
    })

@app.route('/api/register', methods=['POST'])
def register():
    try:
        data = request.get_json()
        print(f"\n{'='*80}")
        print(f"REGISTRATION REQUEST RECEIVED:")
        print(f"Email: {data.get('email')}")
        print(f"User Type: {data.get('user_type')} (type: {type(data.get('user_type'))})")
        print(f"Full data keys: {list(data.keys())}")
        print(f"{'='*80}\n")
        
        # Validate required fields
        required_fields = ['email', 'password', 'first_name', 'last_name', 'user_type']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'{field} is required'}), 400
        
        # VCFSE-specific validation: Charity Commission number is mandatory
        if data['user_type'] == 'vcse':
            if not data.get('charity_commission_number'):
                return jsonify({'error': 'Charity Commission Registration Number is required for VCFSE organizations'}), 400
            if not data.get('organization_name'):
                return jsonify({'error': 'Organization name is required for VCFSE organizations'}), 400
            
            # Verify charity number AND organization name with UK Charity Commission
            verification_result = verify_charity_number(
                data['charity_commission_number'],
                data['organization_name']  # Pass organization name for matching
            )
            
            if not verification_result['valid']:
                error_message = verification_result['message']
                
                # If name mismatch, provide the registered name
                if verification_result.get('name_match') == False:
                    error_message += f"\n\nThe registered charity name is: '{verification_result.get('charity_name')}'\nPlease use this exact name when registering."
                
                return jsonify({
                    'error': error_message,
                    'charity_number': data['charity_commission_number'],
                    'registered_name': verification_result.get('charity_name')
                }), 400
        
        # Check if user already exists
        if User.query.filter_by(email=data['email']).first():
            return jsonify({'error': 'Email already registered'}), 400
        
        # Validate user type
        if data['user_type'] not in ['recipient', 'vendor', 'vcse', 'school', 'admin']:
            return jsonify({'error': 'Invalid user type'}), 400
        
        # Validate password length
        if len(data['password']) < 6:
            return jsonify({'error': 'Password must be at least 6 characters long'}), 400
        
        # Create verification token
        verification_token = secrets.token_urlsafe(32)
        
        # Determine account status: VCFSE organizations are auto-approved if charity verified
        # Since we already verified the charity number above (line 615), VCFSE is approved
        account_status = 'ACTIVE'
        
        # Create new user
        user = User(
            email=data['email'],
            password_hash=generate_password_hash(data['password']),
            first_name=data['first_name'],
            last_name=data['last_name'],
            phone=data.get('phone', ''),
            user_type=data['user_type'],
            organization_name=data.get('organization_name', ''),
            shop_name=data.get('shop_name', ''),
            address=data.get('address', ''),
            postcode=data.get('postcode', ''),
            city=data.get('city', ''),
            charity_commission_number=data.get('charity_commission_number', ''),
            shop_category=data.get('shop_category', ''),
            verification_token=verification_token,
            account_status=account_status
        )
        
        db.session.add(user)
        db.session.commit()
        
        # Auto-create shop for vendors
        if user.user_type == 'vendor' and user.shop_name:
            vendor_shop = VendorShop(
                vendor_id=user.id,
                shop_name=user.shop_name,
                address=user.address or '',
                postcode=user.postcode or '',
                city=user.city or '',
                phone=user.phone or '',
                is_active=True
            )
            db.session.add(vendor_shop)
            db.session.commit()
        
        # Send welcome email (all users are now auto-approved)
        try:
            email_service.send_welcome_email(
                user_email=user.email,
                user_name=f"{user.first_name} {user.last_name}",
                user_type=user.user_type
            )
            
            # Notify admin when VCFSE registers (informational only)
            if user.user_type == 'vcse':
                admin_users = User.query.filter_by(user_type='admin').all()
                for admin in admin_users:
                    try:
                        email_service.send_email(
                            to_email=admin.email,
                            subject=f"New VCFSE Registration: {user.organization_name}",
                            html_content=f"""
                            <h3>New VCFSE Organization Registered</h3>
                            <p><strong>Organization:</strong> {user.organization_name}</p>
                            <p><strong>Charity Number:</strong> {user.charity_commission_number}</p>
                            <p><strong>Contact:</strong> {user.first_name} {user.last_name}</p>
                            <p><strong>Email:</strong> {user.email}</p>
                            <p><strong>Status:</strong> Auto-approved (charity verified)</p>
                            <p>The organization has been automatically approved and can now access the platform.</p>
                            """
                        )
                    except:
                        pass  # Don't fail registration if admin notification fails
                        
        except Exception as email_error:
            print(f"Warning: Could not send email: {email_error}")
        
        # For demo purposes, auto-verify
        user.is_verified = True
        db.session.commit()
        
        # Create welcome notification
        user_type_names = {
            'recipient': 'Voucher Recipient',
            'vendor': 'Food Vendor',
            'vcse': 'VCFSE Organization',
            'admin': 'System Administrator'
        }
        user_type_name = user_type_names.get(user.user_type, 'User')
        
        welcome_message = f"Welcome to BAK UP E-Voucher Platform, {user.first_name}! Your account as a {user_type_name} has been successfully created. You can now access all features available to your account type. If you need any assistance, please contact our support team."
        
        create_notification(
            user.id,
            '🎉 Welcome to BAK UP!',
            welcome_message,
            'success'
        )
        
        # Return appropriate message based on account status
        if user.account_status == 'PENDING_VERIFICATION':
            return jsonify({
                'message': 'Registration submitted! Your VCFSE organization is pending verification. You will receive an email once your account is approved.',
                'user_id': user.id,
                'account_status': 'PENDING_VERIFICATION'
            }), 201
        else:
            return jsonify({
                'message': 'Registration successful! Welcome email sent to ' + user.email,
                'user_id': user.id,
                'account_status': 'ACTIVE'
            }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Registration failed: {str(e)}'}), 500

@app.route('/api/login', methods=['POST'])
@limiter.limit("5 per 15 minutes")
def login():
    try:
        data = request.get_json()
        
        if not data.get('email') or not data.get('password'):
            return jsonify({'error': 'Email and password are required'}), 400
        
        user = User.query.filter_by(email=data['email']).first()
        
        if not user or not check_password_hash(user.password_hash, data['password']):
            return jsonify({'error': 'Invalid email or password'}), 401
        
        # Check account status for VCFSE organizations
        if user.account_status == 'PENDING_VERIFICATION':
            return jsonify({
                'error': 'Your account is pending verification. You will receive an email once your VCFSE organization has been approved by our administrators.',
                'account_status': 'PENDING_VERIFICATION'
            }), 403
        
        if user.account_status == 'REJECTED':
            return jsonify({
                'error': f'Your account application was rejected. Reason: {user.rejection_reason or "Please contact admin@bakupcic.co.uk for more information."}',
                'account_status': 'REJECTED'
            }), 403
        
        if not user.is_verified:
            return jsonify({'error': 'Please verify your email before logging in'}), 401
        
        if not user.is_active:
            return jsonify({'error': 'Account is deactivated'}), 401
        
        # Track login
        track_login(user.id, request.remote_addr, request.headers.get('User-Agent'))
        
        # Create session
        session['user_id'] = user.id
        session['user_type'] = user.user_type
        
        return jsonify({
            'message': 'Login successful',
            'user': {
                'id': user.id,
                'email': user.email,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'user_type': user.user_type,
                'organization_name': user.organization_name,
                'shop_name': user.shop_name,
                'balance': user.balance
            }
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Login failed: {str(e)}'}), 500

@app.route('/api/logout', methods=['POST'])
def logout():
    """Logout user and clear session"""
    try:
        session.clear()
        return jsonify({'success': True, 'message': 'Logged out successfully'}), 200
    except Exception as e:
        return jsonify({'error': f'Logout failed: {str(e)}'}), 500

@app.route('/api/check-auth', methods=['GET'])
def check_auth():
    """Check if user is authenticated and return user data"""
    try:
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'authenticated': False}), 200
        
        user = User.query.get(user_id)
        
        if not user:
            session.clear()
            return jsonify({'authenticated': False}), 200
        
        return jsonify({
            'authenticated': True,
            'user': {
                'id': user.id,
                'email': user.email,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'user_type': user.user_type,
                'organization_name': user.organization_name,
                'shop_name': user.shop_name,
                'balance': user.balance
            }
        }), 200
        
    except Exception as e:
        return jsonify({'authenticated': False, 'error': str(e)}), 200

@app.route('/api/forgot-password', methods=['POST'])
def forgot_password():
    """Request password reset link"""
    try:
        data = request.get_json()
        email = data.get('email')
        
        if not email:
            return jsonify({'error': 'Email is required'}), 400
        
        user = User.query.filter_by(email=email).first()
        
        if not user:
            # Don't reveal if email exists for security
            return jsonify({'message': 'If the email exists, a password reset link has been sent'}), 200
        
        # Generate reset token
        import secrets
        from datetime import datetime, timedelta
        token = secrets.token_urlsafe(32)
        expires_at = datetime.utcnow() + timedelta(hours=1)
        
        # Store token in database using ORM
        reset_token = PasswordResetToken(
            user_id=user.id,
            token=token,
            expires_at=expires_at
        )
        db.session.add(reset_token)
        db.session.commit()
        
        # Send password reset email using SendGrid
        print(f"[FORGOT-PASSWORD] Attempting to send email to {user.email}")
        print(f"[FORGOT-PASSWORD] Email service enabled: {email_service.enabled}")
        try:
            result = email_service.send_password_reset_email(
                user_email=user.email,
                user_name=f"{user.first_name} {user.last_name}",
                reset_token=token
            )
            print(f"[FORGOT-PASSWORD] Email send result: {result}")
        except Exception as email_error:
            print(f"[FORGOT-PASSWORD] ERROR: {str(email_error)}")
            import traceback
            traceback.print_exc()
        
        return jsonify({
            'message': 'If the email exists, a password reset link has been sent'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to process request: {str(e)}'}), 500

@app.route('/api/reset-password', methods=['POST'])
def reset_password():
    """Reset password using token"""
    try:
        data = request.get_json()
        token = data.get('token')
        new_password = data.get('password')
        
        if not token or not new_password:
            return jsonify({'error': 'Token and new password are required'}), 400
        
        # Find valid token using ORM
        from datetime import datetime
        reset_token = PasswordResetToken.query.filter_by(token=token).first()
        
        if not reset_token:
            return jsonify({'error': 'Invalid or expired reset link'}), 400
        
        if reset_token.used:
            return jsonify({'error': 'This reset link has already been used'}), 400
        
        if reset_token.expires_at < datetime.utcnow():
            return jsonify({'error': 'This reset link has expired'}), 400
        
        # Update password
        user = User.query.get(reset_token.user_id)
        user.password_hash = generate_password_hash(new_password)
        
        # Mark token as used
        reset_token.used = True
        
        db.session.commit()
        
        return jsonify({'message': 'Password reset successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to reset password: {str(e)}'}), 500

# Initialize database and categories
def init_db():
    with app.app_context():
        db.create_all()
        
        # Create default categories if they don't exist
        if Category.query.count() == 0:
            categories = [
                Category(name='Fruits', type='edible', icon='🍎', description='Fresh and dried fruits'),
                Category(name='Vegetables', type='edible', icon='🥕', description='Fresh vegetables and salads'),
                Category(name='Canned Foods', type='edible', icon='🥫', description='Canned and preserved foods'),
                Category(name='Dairy', type='edible', icon='🥛', description='Milk, cheese, yogurt and dairy products'),
                Category(name='Bakery', type='edible', icon='🍞', description='Bread, pastries and baked goods'),
                Category(name='Clothing', type='non-edible', icon='👕', description='Clothing and accessories'),
                Category(name='Household Items', type='non-edible', icon='🏠', description='Home and kitchen items'),
                Category(name='Books', type='non-edible', icon='📚', description='Books and educational materials'),
                Category(name='Toys', type='non-edible', icon='🧸', description='Toys and games'),
                Category(name='Other Goods', type='non-edible', icon='📦', description='Other miscellaneous items')
            ]
            
            for category in categories:
                db.session.add(category)
            
            db.session.commit()
            print("Database initialized with default categories")

# ============================================
# VCFSE Money Loading and Voucher Issuing Routes
# ============================================

@app.route('/api/vcse/load-money', methods=['POST'])
def vcse_load_money():
    """VCFSE organizations can load money onto their account"""
    try:
        data = request.get_json()
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vcse':
            return jsonify({'error': 'Only VCFSE organizations can load money'}), 403
        
        amount = data.get('amount')
        if not amount or amount <= 0:
            return jsonify({'error': 'Invalid amount'}), 400
        
        # Update balance
        user.balance += amount
        db.session.commit()
        
        # Create notification
        create_notification(
            user_id,
            'Money Loaded Successfully',
            f'£{amount:.2f} has been added to your account. New balance: £{user.balance:.2f}',
            'success'
        )
        
        return jsonify({
            'message': 'Money loaded successfully',
            'new_balance': user.balance
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to load money: {str(e)}'}), 500

@app.route('/api/vcse/analytics', methods=['GET'])
def vcse_analytics():
    """Get analytics data for VCFSE dashboard"""
    try:
        from sqlalchemy import func
        from datetime import datetime, timedelta
        
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vcse':
            return jsonify({'error': 'Only VCFSE organizations can view analytics'}), 403
        
        # Get all vouchers for this VCFSE
        vouchers = Voucher.query.filter_by(issued_by=user_id).all()
        
        # Total metrics
        total_vouchers = len(vouchers)
        total_value = sum(float(v.value) for v in vouchers)
        active_vouchers = len([v for v in vouchers if v.status == 'active'])
        redeemed_vouchers = len([v for v in vouchers if v.status == 'redeemed'])
        expired_vouchers = len([v for v in vouchers if v.status == 'expired'])
        
        # Status breakdown
        status_breakdown = {
            'active': active_vouchers,
            'redeemed': redeemed_vouchers,
            'expired': expired_vouchers
        }
        
        # Issuance trend (last 30 days)
        thirty_days_ago = datetime.utcnow() - timedelta(days=30)
        recent_vouchers = [v for v in vouchers if v.created_at and v.created_at >= thirty_days_ago]
        
        # Group by date
        issuance_by_date = {}
        for v in recent_vouchers:
            date_key = v.created_at.strftime('%Y-%m-%d')
            issuance_by_date[date_key] = issuance_by_date.get(date_key, 0) + 1
        
        # Fill in missing dates with 0
        trend_data = []
        for i in range(30):
            date = (datetime.utcnow() - timedelta(days=29-i)).strftime('%Y-%m-%d')
            trend_data.append({
                'date': date,
                'count': issuance_by_date.get(date, 0)
            })
        
        # Value distributed by status
        value_by_status = {
            'active': sum(float(v.value) for v in vouchers if v.status == 'active'),
            'redeemed': sum(float(v.value) for v in vouchers if v.status == 'redeemed'),
            'expired': sum(float(v.value) for v in vouchers if v.status == 'expired')
        }
        
        return jsonify({
            'total_vouchers': total_vouchers,
            'total_value': total_value,
            'active_vouchers': active_vouchers,
            'redeemed_vouchers': redeemed_vouchers,
            'expired_vouchers': expired_vouchers,
            'status_breakdown': status_breakdown,
            'issuance_trend': trend_data,
            'value_by_status': value_by_status
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get analytics: {str(e)}'}), 500

@app.route('/api/vcse/vouchers', methods=['GET'])
def vcse_get_vouchers():
    """Get all vouchers issued by this VCFSE organization"""
    try:
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vcse':
            return jsonify({'error': 'Only VCFSE organizations can access this'}), 403
        
        # Get query parameters for filtering
        status_filter = request.args.get('status', 'all')  # all, active, redeemed, expired
        search_query = request.args.get('search', '').lower()
        
        # Query vouchers issued by this VCFSE
        query = Voucher.query.filter_by(issued_by=user_id)
        
        # Apply status filter
        if status_filter != 'all':
            query = query.filter_by(status=status_filter)
        
        vouchers = query.order_by(Voucher.created_at.desc()).all()
        
        # Build response with recipient details
        vouchers_data = []
        for voucher in vouchers:
            recipient = User.query.get(voucher.recipient_id)
            
            voucher_info = {
                'id': voucher.id,
                'code': voucher.code,
                'value': float(voucher.value),
                'status': voucher.status,
                'created_at': voucher.created_at.isoformat() if voucher.created_at else None,
                'expiry_date': voucher.expiry_date.isoformat() if voucher.expiry_date else None,
                'redeemed_date': voucher.redeemed_at.isoformat() if voucher.redeemed_at else None,
                'reassignment_count': voucher.reassignment_count or 0,
                'recipient': {
                    'name': f"{recipient.first_name} {recipient.last_name}" if recipient else 'Unknown',
                    'email': recipient.email if recipient else '',
                    'phone': recipient.phone if recipient else ''
                }
            }
            
            # Apply search filter
            if search_query:
                searchable_text = f"{voucher.code} {voucher_info['recipient']['name']} {voucher_info['recipient']['email']}".lower()
                if search_query not in searchable_text:
                    continue
            
            vouchers_data.append(voucher_info)
        
        return jsonify({
            'vouchers': vouchers_data,
            'total_count': len(vouchers_data)
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get vouchers: {str(e)}'}), 500

@app.route('/api/vcse/recipients', methods=['GET'])
def vcse_get_recipients():
    """Get all recipients that this VCFSE has issued vouchers to"""
    try:
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vcse':
            return jsonify({'error': 'Only VCFSE organizations can access this'}), 403
        
        # Get all unique recipients from vouchers issued by this VCFSE
        vouchers = Voucher.query.filter_by(issued_by=user_id).all()
        recipient_ids = set(v.recipient_id for v in vouchers if v.recipient_id)
        
        recipients_data = []
        for recipient_id in recipient_ids:
            recipient = User.query.get(recipient_id)
            if recipient:
                recipients_data.append({
                    'id': recipient.id,
                    'name': f"{recipient.first_name} {recipient.last_name}",
                    'email': recipient.email,
                    'phone': recipient.phone or '',
                    'address': recipient.address or ''
                })
        
        # Sort by name
        recipients_data.sort(key=lambda x: x['name'])
        
        return jsonify({
            'recipients': recipients_data,
            'total_count': len(recipients_data)
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get recipients: {str(e)}'}), 500

@app.route('/api/vcse/voucher-pdf/<int:voucher_id>', methods=['GET'])
def vcse_voucher_pdf(voucher_id):
    """Generate PDF for a specific voucher with QR code"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.units import inch
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from io import BytesIO
        import qrcode
        from PIL import Image
        
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vcse':
            return jsonify({'error': 'Only VCFSE organizations can download vouchers'}), 403
        
        # Get voucher
        voucher = Voucher.query.get(voucher_id)
        if not voucher or voucher.issued_by != user_id:
            return jsonify({'error': 'Voucher not found or access denied'}), 404
        
        recipient = User.query.get(voucher.recipient_id)
        
        # Create PDF
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        
        # Header - BAK UP branding
        c.setFillColor(colors.HexColor('#4CAF50'))
        c.rect(0, height - 2*inch, width, 2*inch, fill=True, stroke=False)
        c.setFillColor(colors.white)
        c.setFont('Helvetica-Bold', 32)
        c.drawCentredString(width/2, height - 1.2*inch, 'BAK UP E-Voucher')
        c.setFont('Helvetica', 14)
        c.drawCentredString(width/2, height - 1.6*inch, 'Supporting Families Through Education & Care')
        
        # Voucher Code - Large and prominent
        c.setFillColor(colors.black)
        c.setFont('Helvetica-Bold', 48)
        c.drawCentredString(width/2, height - 3*inch, voucher.code)
        c.setFont('Helvetica', 12)
        c.setFillColor(colors.grey)
        c.drawCentredString(width/2, height - 3.3*inch, 'Voucher Code')
        
        # Generate QR Code
        qr = qrcode.QRCode(version=1, box_size=10, border=2)
        qr.add_data(voucher.code)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")
        
        # Save QR code to BytesIO
        qr_buffer = BytesIO()
        qr_img.save(qr_buffer, format='PNG')
        qr_buffer.seek(0)
        
        # Draw QR code on PDF
        c.drawImage(qr_buffer, width/2 - 1.5*inch, height - 6*inch, width=3*inch, height=3*inch)
        
        # Voucher Details Box
        y_position = height - 7*inch
        c.setFillColor(colors.HexColor('#f5f5f5'))
        c.rect(1*inch, y_position - 2.5*inch, width - 2*inch, 2.5*inch, fill=True, stroke=True)
        
        c.setFillColor(colors.black)
        c.setFont('Helvetica-Bold', 14)
        c.drawString(1.2*inch, y_position - 0.4*inch, 'Voucher Details')
        
        c.setFont('Helvetica', 11)
        c.drawString(1.2*inch, y_position - 0.8*inch, f'Recipient: {recipient.first_name} {recipient.last_name}' if recipient else 'Recipient: Unknown')
        c.drawString(1.2*inch, y_position - 1.1*inch, f'Email: {recipient.email}' if recipient else '')
        c.drawString(1.2*inch, y_position - 1.4*inch, f'Phone: {recipient.phone}' if recipient else '')
        
        c.setFont('Helvetica-Bold', 14)
        c.setFillColor(colors.HexColor('#4CAF50'))
        c.drawString(1.2*inch, y_position - 1.9*inch, f'Value: £{float(voucher.value):.2f}')
        
        c.setFillColor(colors.black)
        c.setFont('Helvetica', 11)
        c.drawString(1.2*inch, y_position - 2.2*inch, f'Issue Date: {voucher.created_at.strftime("%d %B %Y") if voucher.created_at else ""}')
        c.drawString(1.2*inch, y_position - 2.5*inch, f'Expiry Date: {voucher.expiry_date.strftime("%d %B %Y") if voucher.expiry_date else ""}')
        
        # Terms and Conditions
        y_position = height - 10*inch
        c.setFont('Helvetica-Bold', 10)
        c.drawString(1*inch, y_position, 'Terms & Conditions:')
        c.setFont('Helvetica', 8)
        terms = [
            '1. This voucher can be redeemed at participating local shops for food and essential items.',
            '2. The voucher must be presented at the time of purchase.',
            '3. The voucher cannot be exchanged for cash.',
            '4. The voucher is valid until the expiry date shown above.',
            '5. Any unused balance will be forfeited after expiry.',
            '6. For assistance, contact your VCFSE organization or visit backup-voucher-system.onrender.com'
        ]
        y_pos = y_position - 0.2*inch
        for term in terms:
            c.drawString(1*inch, y_pos, term)
            y_pos -= 0.15*inch
        
        # Footer
        c.setFont('Helvetica-Oblique', 8)
        c.setFillColor(colors.grey)
        c.drawCentredString(width/2, 0.5*inch, f'Generated on {datetime.utcnow().strftime("%d %B %Y at %H:%M UTC")} | BAK UP E-Voucher System')
        
        c.save()
        buffer.seek(0)
        
        # Send PDF
        from flask import send_file
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'voucher_{voucher.code}.pdf'
        )
        
    except Exception as e:
        return jsonify({'error': f'Failed to generate PDF: {str(e)}'}), 500

@app.route('/api/vcse/export-vouchers', methods=['GET'])
def vcse_export_vouchers():
    """Export all vouchers issued by VCFSE to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vcse':
            return jsonify({'error': 'Only VCFSE organizations can export vouchers'}), 403
        
        # Get all vouchers issued by this VCFSE
        vouchers = Voucher.query.filter_by(issued_by=user_id).order_by(Voucher.created_at.desc()).all()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Voucher Orders"
        
        # Header styling
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Headers
        headers = ['Voucher Code', 'Recipient Name', 'Email', 'Phone', 'Value (£)', 'Status', 'Issue Date', 'Expiry Date', 'Redeemed Date']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for row_idx, voucher in enumerate(vouchers, 2):
            recipient = User.query.get(voucher.recipient_id)
            
            ws.cell(row=row_idx, column=1, value=voucher.code)
            ws.cell(row=row_idx, column=2, value=f"{recipient.first_name} {recipient.last_name}" if recipient else 'Unknown')
            ws.cell(row=row_idx, column=3, value=recipient.email if recipient else '')
            ws.cell(row=row_idx, column=4, value=recipient.phone if recipient else '')
            ws.cell(row=row_idx, column=5, value=float(voucher.value))
            ws.cell(row=row_idx, column=6, value=voucher.status.upper())
            ws.cell(row=row_idx, column=7, value=voucher.created_at.strftime('%Y-%m-%d %H:%M') if voucher.created_at else '')
            ws.cell(row=row_idx, column=8, value=voucher.expiry_date.strftime('%Y-%m-%d') if voucher.expiry_date else '')
            ws.cell(row=row_idx, column=9, value=voucher.redeemed_date.strftime('%Y-%m-%d %H:%M') if voucher.redeemed_date else '-')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Send file
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'voucher_orders_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        return jsonify({'error': f'Failed to export vouchers: {str(e)}'}), 500

@app.route('/api/vcse/reports/pdf', methods=['GET'])
def vcse_generate_pdf_report():
    """Generate comprehensive PDF report for VCFSE organization"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.units import inch
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.platypus import Table, TableStyle
        from io import BytesIO
        from flask import send_file
        
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vcse':
            return jsonify({'error': 'Only VCFSE organizations can generate reports'}), 403
        
        # Get date range from query parameters
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        
        # Parse dates
        from datetime import datetime, timedelta
        if date_from:
            start_date = datetime.strptime(date_from, '%Y-%m-%d')
        else:
            start_date = datetime.now() - timedelta(days=30)
        
        if date_to:
            end_date = datetime.strptime(date_to, '%Y-%m-%d')
        else:
            end_date = datetime.now()
        
        # Get vouchers issued by this VCFSE in date range
        vouchers = Voucher.query.filter(
            Voucher.issued_by == user_id,
            Voucher.created_at >= start_date,
            Voucher.created_at <= end_date
        ).all()
        
        # Calculate statistics
        total_vouchers = len(vouchers)
        total_value = sum(float(v.value) for v in vouchers)
        active_vouchers = len([v for v in vouchers if v.status == 'active'])
        redeemed_vouchers = len([v for v in vouchers if v.status == 'redeemed'])
        expired_vouchers = len([v for v in vouchers if v.status == 'expired'])
        
        # Create PDF
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        
        # Header
        c.setFillColor(colors.HexColor('#4CAF50'))
        c.rect(0, height - 2*inch, width, 2*inch, fill=True, stroke=False)
        c.setFillColor(colors.white)
        c.setFont('Helvetica-Bold', 28)
        c.drawCentredString(width/2, height - 1*inch, 'VCFSE Activity Report')
        c.setFont('Helvetica', 12)
        c.drawCentredString(width/2, height - 1.4*inch, f'{user.organization_name or user.email}')
        c.drawCentredString(width/2, height - 1.7*inch, f'Period: {start_date.strftime("%d %B %Y")} to {end_date.strftime("%d %B %Y")}')
        
        # Summary Statistics
        y_pos = height - 2.5*inch
        c.setFillColor(colors.black)
        c.setFont('Helvetica-Bold', 16)
        c.drawString(1*inch, y_pos, 'Summary Statistics')
        
        y_pos -= 0.4*inch
        c.setFont('Helvetica', 12)
        c.drawString(1*inch, y_pos, f'Total Vouchers Issued: {total_vouchers}')
        y_pos -= 0.25*inch
        c.drawString(1*inch, y_pos, f'Total Value Issued: £{total_value:.2f}')
        y_pos -= 0.25*inch
        c.setFillColor(colors.HexColor('#4CAF50'))
        c.drawString(1*inch, y_pos, f'Active Vouchers: {active_vouchers}')
        y_pos -= 0.25*inch
        c.setFillColor(colors.HexColor('#2196F3'))
        c.drawString(1*inch, y_pos, f'Redeemed Vouchers: {redeemed_vouchers}')
        y_pos -= 0.25*inch
        c.setFillColor(colors.HexColor('#f44336'))
        c.drawString(1*inch, y_pos, f'Expired Vouchers: {expired_vouchers}')
        
        # Voucher List
        y_pos -= 0.6*inch
        c.setFillColor(colors.black)
        c.setFont('Helvetica-Bold', 14)
        c.drawString(1*inch, y_pos, 'Voucher Details')
        
        y_pos -= 0.3*inch
        c.setFont('Helvetica', 9)
        
        # Table headers
        c.setFont('Helvetica-Bold', 9)
        c.drawString(1*inch, y_pos, 'Code')
        c.drawString(2.2*inch, y_pos, 'Recipient')
        c.drawString(3.8*inch, y_pos, 'Value')
        c.drawString(4.5*inch, y_pos, 'Status')
        c.drawString(5.3*inch, y_pos, 'Issue Date')
        c.drawString(6.5*inch, y_pos, 'Expiry')
        
        y_pos -= 0.05*inch
        c.line(1*inch, y_pos, width - 1*inch, y_pos)
        y_pos -= 0.15*inch
        
        c.setFont('Helvetica', 8)
        for voucher in vouchers[:30]:  # Limit to first 30 vouchers
            if y_pos < 1.5*inch:
                # Start new page
                c.showPage()
                y_pos = height - 1*inch
                c.setFont('Helvetica', 8)
            
            recipient = User.query.get(voucher.recipient_id)
            recipient_name = f"{recipient.first_name} {recipient.last_name}" if recipient else "Unknown"
            
            c.drawString(1*inch, y_pos, voucher.code[:12])
            c.drawString(2.2*inch, y_pos, recipient_name[:20])
            c.drawString(3.8*inch, y_pos, f'£{float(voucher.value):.2f}')
            
            # Color-coded status
            if voucher.status == 'active':
                c.setFillColor(colors.HexColor('#4CAF50'))
            elif voucher.status == 'redeemed':
                c.setFillColor(colors.HexColor('#2196F3'))
            else:
                c.setFillColor(colors.HexColor('#f44336'))
            c.drawString(4.5*inch, y_pos, voucher.status.upper())
            c.setFillColor(colors.black)
            
            c.drawString(5.3*inch, y_pos, voucher.created_at.strftime('%d/%m/%Y') if voucher.created_at else '')
            c.drawString(6.5*inch, y_pos, voucher.expiry_date.strftime('%d/%m/%Y') if voucher.expiry_date else '')
            
            y_pos -= 0.2*inch
        
        if total_vouchers > 30:
            y_pos -= 0.2*inch
            c.setFont('Helvetica-Oblique', 9)
            c.drawString(1*inch, y_pos, f'... and {total_vouchers - 30} more vouchers (download Excel for complete list)')
        
        # Footer
        c.setFont('Helvetica-Oblique', 8)
        c.setFillColor(colors.grey)
        c.drawCentredString(width/2, 0.5*inch, f'Generated on {datetime.now().strftime("%d %B %Y at %H:%M")} | BAK UP E-Voucher System')
        
        c.save()
        buffer.seek(0)
        
        # Send PDF
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'vcse_report_{start_date.strftime("%Y%m%d")}_to_{end_date.strftime("%Y%m%d")}.pdf'
        )
        
    except Exception as e:
        return jsonify({'error': f'Failed to generate PDF report: {str(e)}'}), 500

@app.route('/api/school/place-order', methods=['POST'])
def school_place_order():
    """School/Care organizations can place orders for To Go items on behalf of families"""
    try:
        data = request.get_json()
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'school':
            return jsonify({'error': 'Only schools can place orders'}), 403
        
        # Create order
        order = Order(
            vcse_id=user_id,  # Using same field for schools
            surplus_item_id=data['surplus_item_id'],
            client_name=data['client_name'],
            client_mobile=data['client_mobile'],
            client_email=data['client_email'],
            quantity=data.get('quantity', 1),
            status='pending'
        )
        
        db.session.add(order)
        db.session.commit()
        
        return jsonify({'message': 'Order placed successfully', 'order_id': order.id}), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to place order: {str(e)}'}), 500

@app.route('/api/vcse/place-order', methods=['POST'])
def vcse_place_order():
    """VCFSE organizations can place orders for To Go items on behalf of clients"""
    try:
        data = request.get_json()
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vcse':
            return jsonify({'error': 'Only VCFSE organizations can place orders'}), 403
        
        # Validate required fields
        surplus_item_id = data.get('surplus_item_id')
        client_name = data.get('client_name')
        client_mobile = data.get('client_mobile')
        client_email = data.get('client_email')
        quantity = data.get('quantity', 1)
        
        if not all([surplus_item_id, client_name, client_mobile, client_email]):
            return jsonify({'error': 'All fields are required'}), 400
        
        # Check if item exists and is available
        item = SurplusItem.query.get(surplus_item_id)
        if not item:
            return jsonify({'error': 'Item not found'}), 404
        
        if item.status != 'available':
            return jsonify({'error': 'Item is no longer available'}), 400
        
        # Check if sufficient quantity available
        if item.quantity < quantity:
            return jsonify({'error': f'Insufficient quantity. Only {item.quantity} available'}), 400
        
        # Reduce item quantity
        item.quantity -= quantity
        
        # If quantity reaches 0, mark as unavailable
        if item.quantity == 0:
            item.status = 'unavailable'
        
        # Create order
        order = Order(
            vcse_id=user_id,
            surplus_item_id=surplus_item_id,
            client_name=client_name,
            client_mobile=client_mobile,
            client_email=client_email,
            quantity=quantity,
            status='pending'
        )
        
        db.session.add(order)
        db.session.commit()
        
        # Notify vendor
        create_notification(
            item.vendor_id,
            'New Order Received',
            f'VCFSE organization has ordered {quantity}x {item.item_name} for client {client_name}',
            'info'
        )
        
        return jsonify({
            'message': 'Order placed successfully',
            'order_id': order.id
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to place order: {str(e)}'}), 500

@app.route('/api/vcse/issue-voucher', methods=['POST'])
def vcse_issue_voucher():
    """VCFSE organizations can issue vouchers to recipients"""
    try:
        data = request.get_json()
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vcse':
            return jsonify({'error': 'Only VCFSE organizations can issue vouchers'}), 403
        
        # Validate required fields
        recipient_first_name = data.get('recipient_first_name')
        recipient_last_name = data.get('recipient_last_name')
        recipient_email = data.get('recipient_email')
        recipient_date_of_birth = data.get('recipient_date_of_birth')
        recipient_phone = data.get('recipient_phone')
        recipient_address = data.get('recipient_address')
        recipient_city = data.get('recipient_city', '')  # Town/City field
        recipient_postcode = data.get('recipient_postcode', '')
        value = data.get('value')
        expiry_days = data.get('expiry_days', 30)
        selected_shops = data.get('selected_shops')  # List of shop IDs or 'all'
        assign_shop_method = data.get('assign_shop_method', 'specific_shop')  # 'specific_shop' or 'recipient_to_choose'
        
        if not all([recipient_first_name, recipient_last_name, recipient_email, recipient_phone, value]):
            return jsonify({'error': 'All recipient details and value are required'}), 400
        
        if value <= 0:
            return jsonify({'error': 'Voucher value must be positive'}), 400
        
        # Check if VCFSE has sufficient allocated balance from admin
        if user.allocated_balance < value:
            return jsonify({'error': f'Insufficient allocated funds. Current allocated balance: £{user.allocated_balance:.2f}'}), 400
        
        # Find or create recipient
        recipient = User.query.filter_by(email=recipient_email).first()
        if not recipient:
            # Auto-create recipient account with provided details
            import secrets
            temp_password = secrets.token_urlsafe(12)
            # Parse date of birth if provided
            dob = None
            if recipient_date_of_birth:
                try:
                    from datetime import datetime
                    dob = datetime.strptime(recipient_date_of_birth, '%Y-%m-%d').date()
                except:
                    pass
            
            recipient = User(
                email=recipient_email,
                password_hash=generate_password_hash(temp_password),
                first_name=recipient_first_name,
                last_name=recipient_last_name,
                phone=recipient_phone,
                address=recipient_address,
                city=recipient_city,
                postcode=recipient_postcode,
                # date_of_birth=dob,  # Commented out temporarily
                user_type='recipient',
                is_verified=True,
                is_active=True
            )
            db.session.add(recipient)
            db.session.flush()  # Get recipient ID without committing
        
        if recipient.user_type != 'recipient':
            return jsonify({'error': 'Can only issue vouchers to recipients'}), 400
        
        # Generate voucher code
        voucher_code = generate_voucher_code()
        
        # Calculate expiry date
        expiry_date = datetime.utcnow().date() + timedelta(days=expiry_days)
        
        # Create voucher with vendor restrictions
        import json
        vendor_restrictions = None
        if selected_shops and selected_shops != 'all':
            vendor_restrictions = json.dumps(selected_shops)  # Store as JSON array
        
        voucher = Voucher(
            code=voucher_code,
            value=value,
            recipient_id=recipient.id,
            issued_by=user_id,
            expiry_date=expiry_date,
            status='active',
            vendor_restrictions=vendor_restrictions,
            original_recipient_id=recipient.id,
            reassignment_count=0,
            assign_shop_method=assign_shop_method
        )
        
        # Deduct from VCFSE allocated balance
        user.allocated_balance -= value
        
        db.session.add(voucher)
        db.session.commit()
        
        # Create notifications
        create_notification(
            recipient.id,
            'New Voucher Received',
            f'You have received a £{value:.2f} voucher from {user.organization_name}. Code: {voucher_code}',
            'success'
        )
        
        create_notification(
            user_id,
            'Voucher Issued',
            f'Voucher {voucher_code} for £{value:.2f} issued to {recipient.first_name} {recipient.last_name}',
            'success'
        )
        
        # Notify all admins about voucher issuance
        admins = User.query.filter_by(user_type='admin', is_active=True).all()
        for admin in admins:
            create_notification(
                admin.id,
                'Voucher Issued by VCFSE',
                f'{user.organization_name} issued voucher {voucher_code} for £{value:.2f} to {recipient.first_name} {recipient.last_name} ({recipient.email})',
                'info'
            )
        
        # Send SMS notification to recipient with voucher code
        if recipient.phone:
            sms_result = sms_service.send_voucher_code(
                recipient.phone,
                voucher_code,
                f"{recipient.first_name} {recipient.last_name}",
                value
            )
            if not sms_result.get('success'):
                print(f"Failed to send SMS: {sms_result.get('error')}")
        
        # Send email notification to recipient with voucher details
        if recipient.email:
            email_result = email_service.send_voucher_issued_email(
                recipient.email,
                f"{recipient.first_name} {recipient.last_name}",
                voucher_code,
                value,
                user.organization_name
            )
            if not email_result:
                print(f"Failed to send email to {recipient.email}")
        
        return jsonify({
            'message': 'Voucher issued successfully',
            'voucher_code': voucher_code,
            'new_balance': user.balance
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to issue voucher: {str(e)}'}), 500

@app.route('/api/vcse/issue-vouchers-bulk', methods=['POST'])
def vcse_issue_vouchers_bulk():
    """VCFSE organizations can issue vouchers in bulk via CSV upload"""
    try:
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vcse':
            return jsonify({'error': 'Only VCFSE organizations can issue vouchers'}), 403
        
        # Check if file was uploaded
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.endswith('.csv'):
            return jsonify({'error': 'Only CSV files are supported'}), 400
        
        # Parse CSV file
        from bulk_voucher_handler import parse_csv_recipients, create_bulk_vouchers
        parse_result = parse_csv_recipients(file)
        
        if not parse_result['success']:
            return jsonify({'error': parse_result['error']}), 400
        
        if parse_result['total_count'] == 0:
            return jsonify({'error': 'No valid recipients found in CSV'}), 400
        
        # Get optional parameters
        expiry_days = int(request.form.get('expiry_days', 30))
        selected_shops = request.form.get('selected_shops', 'all')
        if selected_shops != 'all':
            import json
            try:
                selected_shops = json.loads(selected_shops)
            except:
                selected_shops = 'all'
        
        # Create vouchers
        result = create_bulk_vouchers(
            db,
            User,
            Voucher,
            parse_result['recipients'],
            user_id,
            expiry_days,
            selected_shops
        )
        
        if 'error' in result:
            return jsonify({'error': result['error']}), 400
        
        # Send notifications to successful recipients
        for success_item in result['successful']:
            recipient = User.query.filter_by(email=success_item['email']).first()
            if recipient:
                # Create in-app notification
                create_notification(
                    recipient.id,
                    'New Voucher Received',
                    f'You have received a £{success_item["value"]:.2f} voucher from {user.organization_name}. Code: {success_item["voucher_code"]}',
                    'success'
                )
                
                # Send SMS if phone available
                if recipient.phone:
                    sms_service.send_voucher_code(
                        recipient.phone,
                        success_item['voucher_code'],
                        success_item['name'],
                        success_item['value']
                    )
                
                # Send email if email available
                if recipient.email:
                    email_service.send_voucher_issued_email(
                        recipient.email,
                        success_item['name'],
                        success_item['voucher_code'],
                        success_item['value'],
                        user.organization_name
                    )
        
        return jsonify({
            'message': 'Bulk voucher issuance completed',
            'summary': {
                'total_processed': result['success_count'] + result['failure_count'],
                'successful': result['success_count'],
                'failed': result['failure_count'],
                'total_value': result['total_value'],
                'remaining_balance': float(user.allocated_balance)
            },
            'successful_vouchers': result['successful'],
            'failed_vouchers': result['failed'],
            'csv_errors': parse_result['errors']
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to process bulk vouchers: {str(e)}'}), 500

@app.route('/api/admin/check-expirations', methods=['POST'])
def admin_check_expirations():
    """Admin endpoint to manually trigger expiration check"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        from expiration_manager import check_and_expire_vouchers
        result = check_and_expire_vouchers(db, Voucher)
        
        if not result['success']:
            return jsonify({'error': result['error']}), 500
        
        return jsonify({
            'message': 'Expiration check completed',
            'expired_count': result['expired_count']
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to check expirations: {str(e)}'}), 500

@app.route('/api/admin/expiring-soon', methods=['GET'])
def admin_get_expiring_soon():
    """Get vouchers expiring within specified days"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        days_ahead = int(request.args.get('days', 7))
        
        from expiration_manager import get_expiring_soon_vouchers
        result = get_expiring_soon_vouchers(Voucher, User, days_ahead)
        
        if not result['success']:
            return jsonify({'error': result['error']}), 500
        
        return jsonify({
            'vouchers': result['vouchers'],
            'count': result['count'],
            'days_ahead': days_ahead
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get expiring vouchers: {str(e)}'}), 500

@app.route('/api/admin/send-expiration-alerts', methods=['POST'])
def admin_send_expiration_alerts():
    """Send alerts for vouchers expiring soon"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        days_ahead = int(request.get_json().get('days', 7))
        
        from expiration_manager import get_expiring_soon_vouchers, send_expiration_alerts
        
        # Get expiring vouchers
        vouchers_result = get_expiring_soon_vouchers(Voucher, User, days_ahead)
        if not vouchers_result['success']:
            return jsonify({'error': vouchers_result['error']}), 500
        
        # Send alerts
        alert_result = send_expiration_alerts(
            vouchers_result['vouchers'],
            sms_service,
            email_service
        )
        
        return jsonify({
            'message': 'Expiration alerts sent',
            'total_vouchers': vouchers_result['count'],
            'alerts_sent': alert_result['sent_count'],
            'alerts_failed': alert_result['failed_count']
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to send alerts: {str(e)}'}), 500

@app.route('/api/admin/expired-report', methods=['GET'])
def admin_get_expired_report():
    """Get report of expired vouchers"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Get date range from query params
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        
        start_date = None
        end_date = None
        
        if start_date_str:
            from datetime import datetime
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        if end_date_str:
            from datetime import datetime
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        from expiration_manager import get_expired_vouchers_report
        result = get_expired_vouchers_report(Voucher, User, start_date, end_date)
        
        if not result['success']:
            return jsonify({'error': result['error']}), 500
        
        return jsonify({
            'vouchers': result['vouchers'],
            'count': result['count'],
            'total_value': result['total_value']
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to generate report: {str(e)}'}), 500

@app.route('/api/admin/export/vouchers', methods=['GET'])
def admin_export_vouchers():
    """Export vouchers to CSV"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Get filter parameters
        status = request.args.get('status')  # active, redeemed, expired
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        
        query = Voucher.query
        
        if status:
            query = query.filter(Voucher.status == status)
        
        if start_date_str and hasattr(Voucher, 'created_at'):
            from datetime import datetime
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            query = query.filter(Voucher.created_at >= start_date)
        
        if end_date_str and hasattr(Voucher, 'created_at'):
            from datetime import datetime
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            query = query.filter(Voucher.created_at <= end_date)
        
        vouchers = query.all()
        
        from transaction_export import export_vouchers_csv
        csv_data = export_vouchers_csv(vouchers, User)
        
        from flask import make_response
        response = make_response(csv_data)
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename=vouchers_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        return response
        
    except Exception as e:
        return jsonify({'error': f'Failed to export vouchers: {str(e)}'}), 500

@app.route('/api/admin/export/surplus-items', methods=['GET'])
def admin_export_surplus_items():
    """Export surplus items to CSV"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        surplus_items = SurplusItem.query.all()
        
        from transaction_export import export_surplus_items_csv
        csv_data = export_surplus_items_csv(surplus_items, User)
        
        from flask import make_response
        from datetime import datetime
        response = make_response(csv_data)
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename=surplus_items_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        return response
        
    except Exception as e:
        return jsonify({'error': f'Failed to export surplus items: {str(e)}'}), 500

@app.route('/api/admin/export/users', methods=['GET'])
def admin_export_users():
    """Export users to CSV"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Get filter parameters
        user_type = request.args.get('user_type')  # vcse, vendor, school, recipient
        
        query = User.query
        if user_type:
            query = query.filter(User.user_type == user_type)
        
        users = query.all()
        
        from transaction_export import export_users_csv
        csv_data = export_users_csv(users)
        
        from flask import make_response
        from datetime import datetime
        response = make_response(csv_data)
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename=users_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        return response
        
    except Exception as e:
        return jsonify({'error': f'Failed to export users: {str(e)}'}), 500

@app.route('/api/admin/export/financial-report', methods=['GET'])
def admin_export_financial_report():
    """Export comprehensive financial report to CSV"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Get date range
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        
        start_date = None
        end_date = None
        
        if start_date_str:
            from datetime import datetime
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        if end_date_str:
            from datetime import datetime
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        
        from transaction_export import generate_financial_report_csv
        csv_data = generate_financial_report_csv(Voucher, User, start_date, end_date)
        
        from flask import make_response
        from datetime import datetime
        response = make_response(csv_data)
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename=financial_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        return response
        
    except Exception as e:
        return jsonify({'error': f'Failed to export financial report: {str(e)}'}), 500

@app.route('/api/admin/export/impact-report', methods=['GET'])
def admin_export_impact_report():
    """Export impact report to CSV"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        from transaction_export import export_impact_report_csv
        csv_data = export_impact_report_csv(Voucher, SurplusItem, User)
        
        from flask import make_response
        from datetime import datetime
        response = make_response(csv_data)
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename=impact_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        return response
        
    except Exception as e:
        return jsonify({'error': f'Failed to export impact report: {str(e)}'}), 500

@app.route('/api/admin/check-low-balances', methods=['GET'])
def admin_check_low_balances():
    """Check for VCFSE organizations with low balances"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        threshold = float(request.args.get('threshold', 50.0))
        
        from balance_alerts import check_low_balances
        result = check_low_balances(User, threshold)
        
        if not result['success']:
            return jsonify({'error': result['error']}), 500
        
        return jsonify({
            'organizations': result['organizations'],
            'count': result['count'],
            'threshold': threshold
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to check balances: {str(e)}'}), 500

@app.route('/api/admin/send-balance-alerts', methods=['POST'])
def admin_send_balance_alerts():
    """Send alerts to organizations with low balances"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        data = request.get_json()
        threshold = float(data.get('threshold', 50.0))
        
        from balance_alerts import check_low_balances, send_low_balance_alerts
        
        # Get organizations with low balance
        balance_result = check_low_balances(User, threshold)
        if not balance_result['success']:
            return jsonify({'error': balance_result['error']}), 500
        
        # Send alerts
        alert_result = send_low_balance_alerts(
            balance_result['organizations'],
            sms_service,
            email_service,
            user.email  # Admin email
        )
        
        return jsonify({
            'message': 'Balance alerts sent',
            'organizations_alerted': balance_result['count'],
            'alerts_sent': alert_result['sent_count'],
            'alerts_failed': alert_result['failed_count']
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to send alerts: {str(e)}'}), 500

@app.route('/api/admin/balance-summary', methods=['GET'])
def admin_get_balance_summary():
    """Get balance summary for all VCFSE organizations"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        from balance_alerts import get_balance_summary
        result = get_balance_summary(User)
        
        if not result['success']:
            return jsonify({'error': result['error']}), 500
        
        return jsonify(result['summary']), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get balance summary: {str(e)}'}), 500

@app.route('/api/vcse/balance', methods=['GET'])
def vcse_get_balance():
    """Get current VCFSE balance"""
    try:
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vcse':
            return jsonify({'error': 'Only VCFSE organizations can view balance'}), 403
        
        return jsonify({
            'balance': user.balance,
            'allocated_balance': user.allocated_balance,
            'organization_name': user.organization_name
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get balance: {str(e)}'}), 500

# ============================================
# Stripe Payment Integration Routes
# ============================================

@app.route('/api/config/stripe', methods=['GET'])
def get_stripe_config():
    """Get Stripe publishable key for frontend initialization"""
    try:
        publishable_key = stripe_payment.get_publishable_key()
        return jsonify({'publishableKey': publishable_key}), 200
    except Exception as e:
        return jsonify({'error': f'Failed to get Stripe config: {str(e)}'}), 500

@app.route('/api/admin/stripe-diagnostic', methods=['GET'])
def stripe_diagnostic():
    """Admin-only endpoint to diagnose Stripe configuration issues"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        import stripe
        raw_secret = os.getenv('STRIPE_SECRET_KEY', '')
        raw_publishable = os.getenv('STRIPE_PUBLISHABLE_KEY', '')
        
        diagnostic = {
            'secret_key_configured': bool(raw_secret),
            'secret_key_length': len(raw_secret) if raw_secret else 0,
            'secret_key_starts_with': raw_secret[:10] if raw_secret else 'N/A',
            'secret_key_format_valid': raw_secret.startswith('sk_test_') or raw_secret.startswith('sk_live_') if raw_secret else False,
            'secret_key_has_bearer_prefix': raw_secret.startswith('Bearer ') or raw_secret.startswith('bearer ') if raw_secret else False,
            'publishable_key_configured': bool(raw_publishable),
            'publishable_key_length': len(raw_publishable) if raw_publishable else 0,
            'publishable_key_starts_with': raw_publishable[:10] if raw_publishable else 'N/A',
            'current_stripe_api_key': stripe.api_key[:15] + '...' if stripe.api_key else 'Not set',
            'stripe_api_key_length': len(stripe.api_key) if stripe.api_key else 0
        }
        
        return jsonify(diagnostic), 200
        
    except Exception as e:
        return jsonify({'error': f'Diagnostic failed: {str(e)}'}), 500

@app.route('/api/admin/email-diagnostic', methods=['GET'])
def email_diagnostic():
    """Admin-only endpoint to diagnose email configuration issues"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        gmail_user_raw = os.getenv('GMAIL_USER', '')
        gmail_password_raw = os.getenv('GMAIL_APP_PASSWORD', '')
        from_email_raw = os.getenv('FROM_EMAIL', gmail_user_raw)
        
        # Show both raw and sanitized values
        gmail_user_sanitized = gmail_user_raw.strip() if gmail_user_raw else ''
        gmail_password_sanitized = gmail_password_raw.strip() if gmail_password_raw else ''
        
        diagnostic = {
            'email_service_enabled': email_service.enabled,
            'gmail_user_raw': repr(gmail_user_raw) if gmail_user_raw else 'Not set',
            'gmail_user_sanitized': gmail_user_sanitized if gmail_user_sanitized else 'Not set',
            'gmail_user_has_whitespace': gmail_user_raw != gmail_user_sanitized if gmail_user_raw else False,
            'gmail_app_password_configured': bool(gmail_password_raw),
            'gmail_app_password_length_raw': len(gmail_password_raw) if gmail_password_raw else 0,
            'gmail_app_password_length_sanitized': len(gmail_password_sanitized) if gmail_password_sanitized else 0,
            'gmail_app_password_has_whitespace': gmail_password_raw != gmail_password_sanitized if gmail_password_raw else False,
            'from_email': from_email_raw.strip() if from_email_raw else 'Not set',
            'smtp_server': email_service.smtp_server,
            'smtp_port': email_service.smtp_port,
            'app_url': email_service.app_url,
            'smtp_user_in_service': email_service.smtp_user,
            'smtp_password_length_in_service': len(email_service.smtp_password) if email_service.smtp_password else 0
        }
        
        return jsonify(diagnostic), 200
        
    except Exception as e:
        return jsonify({'error': f'Email diagnostic failed: {str(e)}'}), 500

@app.route('/api/admin/test-email', methods=['POST'])
def test_email():
    """Admin-only endpoint to test email sending"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        data = request.get_json()
        test_email_address = data.get('email', user.email)
        
        # Try to send a test email
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        
        try:
            # Get sanitized credentials
            gmail_user = os.getenv('GMAIL_USER', '').strip()
            gmail_password = os.getenv('GMAIL_APP_PASSWORD', '').strip()
            
            if not gmail_user or not gmail_password:
                return jsonify({
                    'success': False,
                    'error': 'Gmail credentials not configured'
                }), 500
            
            # Create test message
            message = MIMEMultipart('alternative')
            message['From'] = f"BAK UP Test <{gmail_user}>"
            message['To'] = test_email_address
            message['Subject'] = "🧪 BAK UP Email Test"
            
            html_content = """
            <html>
            <body>
                <h2>✅ Email Service is Working!</h2>
                <p>This is a test email from the BAK UP E-Voucher System.</p>
                <p>If you received this, email sending is configured correctly.</p>
            </body>
            </html>
            """
            
            html_part = MIMEText(html_content, 'html')
            message.attach(html_part)
            
            # Try to send
            with smtplib.SMTP('smtp.gmail.com', 587, timeout=10) as server:
                server.starttls()
                server.login(gmail_user, gmail_password)
                server.send_message(message)
            
            return jsonify({
                'success': True,
                'message': f'Test email sent successfully to {test_email_address}',
                'smtp_user': gmail_user
            }), 200
            
        except smtplib.SMTPAuthenticationError as e:
            return jsonify({
                'success': False,
                'error': 'SMTP Authentication Failed',
                'details': str(e),
                'suggestion': 'Check if Gmail App Password is correct and not expired'
            }), 500
            
        except Exception as e:
            return jsonify({
                'success': False,
                'error': 'Email sending failed',
                'details': str(e)
            }), 500
        
    except Exception as e:
        return jsonify({'error': f'Test email failed: {str(e)}'}), 500

@app.route('/api/payment/create-intent', methods=['POST'])
def create_payment_intent():
    """Create a Stripe Payment Intent for VCFSE fund loading"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vcse':
            return jsonify({'error': 'Only VCFSE organizations can load funds'}), 403
        
        data = request.get_json()
        amount = data.get('amount')
        description = data.get('description', 'Fund loading for voucher distribution')
        
        # Validate amount
        if not amount:
            return jsonify({'error': 'Amount is required'}), 400
        
        try:
            amount = float(amount)
        except ValueError:
            return jsonify({'error': 'Invalid amount format'}), 400
        
        # Minimum £10, maximum £10,000
        if amount < 10:
            return jsonify({'error': 'Minimum payment amount is £10'}), 400
        if amount > 10000:
            return jsonify({'error': 'Maximum payment amount is £10,000'}), 400
        
        # Create Stripe Payment Intent
        try:
            payment_result = stripe_payment.create_payment_intent(
                amount=amount,
                vcse_id=user.id,
                vcse_email=user.email,
                description=description
            )
        except Exception as stripe_error:
            return jsonify({'error': f'Payment setup failed: {str(stripe_error)}'}), 500
        
        # Create payment transaction record
        transaction = PaymentTransaction(
            vcse_id=user.id,
            amount=amount,
            currency='GBP',
            stripe_payment_intent_id=payment_result['payment_intent_id'],
            status='pending',
            description=description
        )
        
        db.session.add(transaction)
        db.session.commit()
        
        return jsonify({
            'client_secret': payment_result['client_secret'],
            'payment_intent_id': payment_result['payment_intent_id'],
            'amount': amount,
            'transaction_id': transaction.id
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to create payment intent: {str(e)}'}), 500

@app.route('/api/payment/verify', methods=['POST'])
def verify_payment():
    """Verify payment completion and update user balance"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vcse':
            return jsonify({'error': 'Only VCFSE organizations can load funds'}), 403
        
        data = request.get_json()
        payment_intent_id = data.get('payment_intent_id')
        
        if not payment_intent_id:
            return jsonify({'error': 'Payment intent ID is required'}), 400
        
        # Verify payment with Stripe
        try:
            verification = stripe_payment.verify_payment(payment_intent_id)
        except Exception as stripe_error:
            return jsonify({'error': f'Payment verification failed: {str(stripe_error)}'}), 500
        
        # Find transaction record
        transaction = PaymentTransaction.query.filter_by(
            stripe_payment_intent_id=payment_intent_id
        ).first()
        
        if not transaction:
            return jsonify({'error': 'Transaction not found'}), 404
        
        # Verify transaction belongs to this user
        if transaction.vcse_id != user.id:
            return jsonify({'error': 'Unauthorized'}), 403
        
        # Check if already processed
        if transaction.status == 'succeeded':
            return jsonify({
                'success': True,
                'message': 'Payment already processed',
                'amount': transaction.amount,
                'new_balance': user.balance,
                'transaction_id': transaction.id
            }), 200
        
        # Update transaction based on verification
        if verification['verified']:
            transaction.status = 'succeeded'
            transaction.completed_at = datetime.utcnow()
            transaction.payment_method_id = verification.get('payment_method')
            
            # Add funds to user balance
            user.balance = (user.balance or 0) + transaction.amount
            
            # Create success notification
            create_notification(
                user.id,
                '💳 Payment Successful',
                f'Your payment of £{transaction.amount:.2f} has been processed successfully. Your new balance is £{user.balance:.2f}.',
                'success'
            )
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'amount': transaction.amount,
                'new_balance': user.balance,
                'transaction_id': transaction.id
            }), 200
        else:
            transaction.status = 'failed'
            transaction.failure_reason = f"Payment status: {verification['status']}"
            db.session.commit()
            
            return jsonify({
                'success': False,
                'error': 'Payment not completed',
                'status': verification['status']
            }), 400
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Payment verification failed: {str(e)}'}), 500

@app.route('/api/payment/history', methods=['GET'])
def get_payment_history():
    """Get payment transaction history for VCFSE user"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vcse':
            return jsonify({'error': 'Only VCFSE organizations can view payment history'}), 403
        
        # Get query parameters
        limit = request.args.get('limit', 50, type=int)
        offset = request.args.get('offset', 0, type=int)
        status_filter = request.args.get('status')
        
        # Build query
        query = PaymentTransaction.query.filter_by(vcse_id=user.id)
        
        if status_filter:
            query = query.filter_by(status=status_filter)
        
        # Get total count
        total = query.count()
        
        # Get transactions with pagination
        transactions = query.order_by(
            PaymentTransaction.created_at.desc()
        ).limit(limit).offset(offset).all()
        
        # Format response
        transaction_list = []
        for t in transactions:
            transaction_list.append({
                'id': t.id,
                'amount': t.amount,
                'currency': t.currency,
                'status': t.status,
                'description': t.description,
                'created_at': t.created_at.isoformat() if t.created_at else None,
                'completed_at': t.completed_at.isoformat() if t.completed_at else None,
                'failure_reason': t.failure_reason
            })
        
        return jsonify({
            'transactions': transaction_list,
            'total': total,
            'current_balance': user.balance,
            'allocated_balance': user.allocated_balance
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get payment history: {str(e)}'}), 500

@app.route('/api/payment/webhook', methods=['POST'])
def stripe_webhook():
    """Handle Stripe webhook events"""
    try:
        payload = request.data
        sig_header = request.headers.get('Stripe-Signature')
        webhook_secret = os.environ.get('STRIPE_WEBHOOK_SECRET')
        
        if not webhook_secret:
            print('Warning: STRIPE_WEBHOOK_SECRET not configured')
            return jsonify({'error': 'Webhook not configured'}), 500
        
        try:
            import stripe
            event = stripe.Webhook.construct_event(
                payload, sig_header, webhook_secret
            )
        except ValueError as e:
            # Invalid payload
            return jsonify({'error': 'Invalid payload'}), 400
        except stripe.error.SignatureVerificationError as e:
            # Invalid signature
            return jsonify({'error': 'Invalid signature'}), 400
        
        # Handle the event
        if event['type'] == 'payment_intent.succeeded':
            payment_intent = event['data']['object']
            payment_intent_id = payment_intent['id']
            
            # Find transaction
            transaction = PaymentTransaction.query.filter_by(
                stripe_payment_intent_id=payment_intent_id
            ).first()
            
            if transaction and transaction.status != 'succeeded':
                # Update transaction
                transaction.status = 'succeeded'
                transaction.completed_at = datetime.utcnow()
                
                # Update user balance
                user = User.query.get(transaction.vcse_id)
                if user:
                    user.balance = (user.balance or 0) + transaction.amount
                    
                    # Create notification
                    create_notification(
                        user.id,
                        '💳 Payment Confirmed',
                        f'Your payment of £{transaction.amount:.2f} has been confirmed. Your new balance is £{user.balance:.2f}.',
                        'success'
                    )
                
                db.session.commit()
                print(f'Payment succeeded webhook processed: {payment_intent_id}')
        
        elif event['type'] == 'payment_intent.payment_failed':
            payment_intent = event['data']['object']
            payment_intent_id = payment_intent['id']
            
            # Find transaction
            transaction = PaymentTransaction.query.filter_by(
                stripe_payment_intent_id=payment_intent_id
            ).first()
            
            if transaction:
                transaction.status = 'failed'
                transaction.failure_reason = payment_intent.get('last_payment_error', {}).get('message', 'Payment failed')
                
                # Create notification
                user = User.query.get(transaction.vcse_id)
                if user:
                    create_notification(
                        user.id,
                        '❌ Payment Failed',
                        f'Your payment of £{transaction.amount:.2f} failed. Please try again or use a different payment method.',
                        'error'
                    )
                
                db.session.commit()
                print(f'Payment failed webhook processed: {payment_intent_id}')
        
        elif event['type'] == 'payment_intent.canceled':
            payment_intent = event['data']['object']
            payment_intent_id = payment_intent['id']
            
            # Find transaction
            transaction = PaymentTransaction.query.filter_by(
                stripe_payment_intent_id=payment_intent_id
            ).first()
            
            if transaction:
                transaction.status = 'cancelled'
                db.session.commit()
                print(f'Payment cancelled webhook processed: {payment_intent_id}')
        
        return jsonify({'success': True}), 200
        
    except Exception as e:
        print(f'Webhook error: {str(e)}')
        return jsonify({'error': str(e)}), 500

# ============================================
# Vendor Multi-Shop Management Routes
# ============================================

@app.route('/api/vendor/shops', methods=['GET'])
def vendor_get_shops():
    """Get all shops for a vendor"""
    try:
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vendor':
            return jsonify({'error': 'Only vendors can view shops'}), 403
        
        shops = VendorShop.query.filter_by(vendor_id=user_id, is_active=True).all()
        
        # Calculate total sales across all vendor's shops
        # Note: Vouchers are redeemed by vendor (user_id), not by specific shop
        redemptions = Voucher.query.filter_by(
            redeemed_by_vendor=user_id,
            status='redeemed'
        ).all()
        total_sales = sum(float(v.value) for v in redemptions)
        
        return jsonify({
            'shops': [{
                'id': shop.id,
                'shop_name': shop.shop_name,
                'address': shop.address,
                'postcode': shop.postcode,
                'city': shop.city,
                'town': shop.town,
                'phone': shop.phone,
                'created_at': shop.created_at.isoformat()
            } for shop in shops],
            'total_sales': float(total_sales)
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get shops: {str(e)}'}), 500

@app.route('/api/vendor/shops', methods=['POST'])
def vendor_add_shop():
    """Add a new shop location for a vendor"""
    try:
        data = request.get_json()
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vendor':
            return jsonify({'error': 'Only vendors can add shops'}), 403
        
        # Validate required fields
        required_fields = ['shop_name', 'address']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'{field} is required'}), 400
        
        # Create new shop
        shop = VendorShop(
            vendor_id=user_id,
            shop_name=data['shop_name'],
            address=data['address'],
            postcode=data.get('postcode', ''),
            city=data.get('city', ''),
            town=data.get('town', ''),
            phone=data.get('phone', '')
        )
        
        db.session.add(shop)
        db.session.commit()
        
        create_notification(
            user_id,
            'Shop Added',
            f'New shop location "{shop.shop_name}" has been added successfully',
            'success'
        )
        
        return jsonify({
            'message': 'Shop added successfully',
            'shop_id': shop.id
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to add shop: {str(e)}'}), 500

@app.route('/api/vendor/shops/<int:shop_id>', methods=['PUT'])
def vendor_update_shop(shop_id):
    """Update shop details"""
    try:
        data = request.get_json()
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vendor':
            return jsonify({'error': 'Only vendors can update shops'}), 403
        
        shop = VendorShop.query.filter_by(id=shop_id, vendor_id=user_id).first()
        if not shop:
            return jsonify({'error': 'Shop not found'}), 404
        
        # Update shop fields
        if 'shop_name' in data:
            shop.shop_name = data['shop_name']
        if 'address' in data:
            shop.address = data['address']
        if 'postcode' in data:
            shop.postcode = data['postcode']
        if 'city' in data:
            shop.city = data['city']
        if 'town' in data:
            shop.town = data['town']
        if 'phone' in data:
            shop.phone = data['phone']
        
        db.session.commit()
        
        create_notification(
            user_id,
            'Shop Updated',
            f'Shop "{shop.shop_name}" has been updated successfully',
            'success'
        )
        
        return jsonify({
            'message': 'Shop updated successfully',
            'shop': {
                'id': shop.id,
                'shop_name': shop.shop_name,
                'address': shop.address,
                'postcode': shop.postcode,
                'city': shop.city,
                'town': shop.town,
                'phone': shop.phone
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to update shop: {str(e)}'}), 500

@app.route('/api/vendor/shops/<int:shop_id>', methods=['DELETE'])
def vendor_delete_shop(shop_id):
    """Delete (deactivate) a shop location"""
    try:
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vendor':
            return jsonify({'error': 'Only vendors can delete shops'}), 403
        
        shop = VendorShop.query.filter_by(id=shop_id, vendor_id=user_id).first()
        if not shop:
            return jsonify({'error': 'Shop not found'}), 404
        
        shop.is_active = False
        db.session.commit()
        
        return jsonify({'message': 'Shop deactivated successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to delete shop: {str(e)}'}), 500

# ============================================
# Voucher Shop Selection Routes
# ============================================

@app.route('/api/vouchers/<string:voucher_code>/available-shops', methods=['GET'])
def get_available_shops_for_voucher(voucher_code):
    """Get all shop locations where a voucher can be redeemed"""
    try:
        voucher = Voucher.query.filter_by(code=voucher_code).first()
        
        if not voucher:
            return jsonify({'error': 'Voucher not found'}), 404
        
        if voucher.status != 'active':
            return jsonify({'error': 'Voucher is not active'}), 400
        
        # Get all vendors (for now, vouchers can be redeemed at any vendor)
        # In future, can implement vendor restrictions
        vendors = User.query.filter_by(user_type='vendor', is_active=True).all()
        
        shops_list = []
        for vendor in vendors:
            # Get all shops for this vendor
            shops = VendorShop.query.filter_by(vendor_id=vendor.id, is_active=True).all()
            for shop in shops:
                shops_list.append({
                    'id': shop.id,
                    'vendor_id': vendor.id,
                    'vendor_name': vendor.shop_name or vendor.organization_name,
                    'shop_name': shop.shop_name,
                    'address': shop.address,
                    'postcode': shop.postcode,
                    'city': shop.city,
                    'phone': shop.phone
                })
        
        return jsonify({
            'voucher_code': voucher_code,
            'voucher_value': voucher.value,
            'expiry_date': voucher.expiry_date.isoformat(),
            'available_shops': shops_list
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get shops: {str(e)}'}), 500

# ============================================
# Admin Login Tracking Routes
# ============================================

@app.route('/api/admin/login-stats', methods=['GET'])
def admin_get_login_stats():
    """Get login frequency statistics for all users"""
    try:
        from datetime import datetime, timedelta
        
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Only admins can view login statistics'}), 403
        
        # Get all users (excluding admins)
        all_users = User.query.filter(User.user_type != 'admin', User.is_active == True).all()
        
        users_data = []
        total_logins = 0
        active_users_count = 0
        thirty_days_ago = datetime.utcnow() - timedelta(days=30)
        
        for u in all_users:
            login_count = u.login_count or 0
            total_logins += login_count
            
            # Calculate days since last login
            days_since_login = None
            if u.last_login:
                days_since_login = (datetime.utcnow() - u.last_login).days
                if u.last_login >= thirty_days_ago:
                    active_users_count += 1
            
            # Determine display name based on user type
            if u.user_type == 'vendor':
                display_name = u.shop_name or f'{u.first_name} {u.last_name}'
            elif u.user_type == 'vcse':
                display_name = u.organization_name or f'{u.first_name} {u.last_name}'
            elif u.user_type == 'school':
                # School users use organization_name field
                display_name = u.organization_name or f'{u.first_name} {u.last_name}'
            else:
                display_name = f'{u.first_name} {u.last_name}'
            
            users_data.append({
                'id': u.id,
                'first_name': u.first_name,
                'last_name': u.last_name,
                'email': u.email,
                'role': u.user_type,
                'login_count': login_count,
                'last_login': u.last_login.isoformat() if u.last_login else None,
                'days_since_login': days_since_login,
                'display_name': display_name
            })
        
        # Sort by login count (descending)
        users_data.sort(key=lambda x: x['login_count'], reverse=True)
        
        return jsonify({
            'users': users_data,
            'total_users': len(users_data),
            'active_users': active_users_count,
            'total_logins': total_logins
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get login stats: {str(e)}'}), 500

@app.route('/api/admin/prepopulate-login-stats', methods=['POST'])
def prepopulate_login_stats():
    """Prepopulate login statistics with test data for demonstration"""
    try:
        from datetime import datetime, timedelta
        import random
        
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Only admins can prepopulate data'}), 403
        
        # Get all non-admin users
        all_users = User.query.filter(User.user_type != 'admin').all()
        
        updated_count = 0
        for u in all_users:
            # Generate realistic login data
            if u.user_type == 'vcse':
                login_count = random.randint(15, 50)
                days_ago = random.randint(1, 7)
            elif u.user_type == 'vendor':
                login_count = random.randint(20, 60)
                days_ago = random.randint(1, 5)
            elif u.user_type == 'school':
                login_count = random.randint(10, 40)
                days_ago = random.randint(2, 10)
            else:  # recipient
                login_count = random.randint(5, 30)
                days_ago = random.randint(1, 14)
            
            # Update user
            u.login_count = login_count
            u.last_login = datetime.utcnow() - timedelta(days=days_ago)
            updated_count += 1
        
        db.session.commit()
        
        return jsonify({
            'message': 'Login statistics prepopulated successfully',
            'updated_users': updated_count
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to prepopulate login stats: {str(e)}'}), 500


@app.route('/api/admin/prepopulate-school-data', methods=['POST'])
def prepopulate_school_data():
    """Prepopulate school portal with test vouchers and To Go items"""
    try:
        from datetime import datetime, timedelta
        import random
        
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Only admins can prepopulate data'}), 403
        
        # Get a school user to issue vouchers from
        school_user = User.query.filter_by(user_type='school').first()
        if not school_user:
            return jsonify({'error': 'No school users found'}), 404
        
        # Create test vouchers with different statuses
        voucher_count = 0
        statuses = ['active', 'active', 'active', 'redeemed', 'redeemed', 'expired', 'reassigned']
        
        for i, status in enumerate(statuses):
            # Create or get recipient
            recipient_email = f"testrecipient{i+1}@example.com"
            recipient = User.query.filter_by(email=recipient_email).first()
            
            if not recipient:
                recipient = User(
                    email=recipient_email,
                    first_name=f"Test{i+1}",
                    last_name="Recipient",
                    phone=f"07700{900000+i}",
                    address=f"{i+1} Test Street, London",
                    user_type='recipient',
                    password_hash='dummy'
                )
                db.session.add(recipient)
                db.session.flush()
            
            # Create voucher
            voucher_code = f"TEST{random.randint(10000, 99999)}"
            expiry_date = datetime.utcnow() + timedelta(days=30 if status != 'expired' else -5)
            
            voucher = Voucher(
                code=voucher_code,
                value=random.choice([5.0, 10.0, 15.0, 20.0]),
                status=status,
                recipient_id=recipient.id,
                issued_by=school_user.id,
                created_at=datetime.utcnow() - timedelta(days=random.randint(1, 30)),
                expiry_date=expiry_date
            )
            db.session.add(voucher)
            voucher_count += 1
        
        # Create test To Go items
        # First, ensure we have a vendor with a shop
        vendor = User.query.filter_by(user_type='vendor').first()
        if vendor:
            shop = VendorShop.query.filter_by(vendor_id=vendor.id).first()
            if shop:
                # Create surplus items
                items_data = [
                    {'name': 'Fresh Bread Loaves', 'qty': 10, 'unit': 'loaves', 'category': 'Bakery', 'price': 1.50},
                    {'name': 'Milk Bottles', 'qty': 8, 'unit': 'bottles', 'category': 'Dairy', 'price': 1.20},
                    {'name': 'Fresh Vegetables Mix', 'qty': 15, 'unit': 'bags', 'category': 'Produce', 'price': 2.00},
                    {'name': 'Canned Soup', 'qty': 20, 'unit': 'cans', 'category': 'Canned Goods', 'price': 0.80},
                    {'name': 'Rice Bags', 'qty': 12, 'unit': 'bags', 'category': 'Grains', 'price': 3.00}
                ]
                
                items_count = 0
                for item_data in items_data:
                    # Check if item already exists
                    existing = SurplusItem.query.filter_by(
                        shop_id=shop.id,
                        item_name=item_data['name']
                    ).first()
                    
                    if not existing:
                        surplus_item = SurplusItem(
                            shop_id=shop.id,
                            item_name=item_data['name'],
                            quantity=item_data['qty'],
                            unit=item_data['unit'],
                            category=item_data['category'],
                            price=item_data['price'],
                            description=f"Fresh {item_data['name']} available for collection",
                            status='available',
                            posted_at=datetime.utcnow()
                        )
                        db.session.add(surplus_item)
                        items_count += 1
        
        db.session.commit()
        
        return jsonify({
            'message': 'School test data prepopulated successfully',
            'vouchers_created': voucher_count,
            'items_created': items_count if vendor and shop else 0
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to prepopulate school data: {str(e)}'}), 500


# ============================================# Admin Voucher Reassignment Routes
# ============================================

@app.route('/api/admin/vouchers/unredeemed', methods=['GET'])
def admin_get_unredeemed_vouchers():
    """Get all unredeemed vouchers for reassignment"""
    try:
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Only admins can view unredeemed vouchers'}), 403
        
        vouchers = Voucher.query.filter_by(status='active').all()
        unredeemed = []
        
        for voucher in vouchers:
            recipient = User.query.get(voucher.recipient_id) if voucher.recipient_id else None
            issuer = User.query.get(voucher.issued_by)
            
            unredeemed.append({
                'id': voucher.id,
                'code': voucher.code,
                'value': voucher.value,
                'recipient_name': f'{recipient.first_name} {recipient.last_name}' if recipient else 'Unassigned',
                'recipient_email': recipient.email if recipient else None,
                'recipient_phone': recipient.phone if recipient else None,
                'recipient_address': recipient.address if recipient else None,
                'recipient_city': recipient.city if recipient else None,
                'recipient_postcode': recipient.postcode if recipient else None,
                'issued_by': issuer.organization_name if issuer.user_type == 'vcse' else 'Admin',
                'expiry_date': voucher.expiry_date.isoformat(),
                'created_at': voucher.created_at.isoformat()
            })
        
        return jsonify({'unredeemed_vouchers': unredeemed}), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get unredeemed vouchers: {str(e)}'}), 500

@app.route('/api/admin/vouchers/<int:voucher_id>/reassign', methods=['POST'])
def admin_reassign_voucher(voucher_id):
    """Reassign a voucher to a new recipient"""
    try:
        data = request.get_json()
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Only admins can reassign vouchers'}), 403
        
        voucher = Voucher.query.get(voucher_id)
        if not voucher:
            return jsonify({'error': 'Voucher not found'}), 404
        
        if voucher.status != 'active':
            return jsonify({'error': 'Can only reassign active vouchers'}), 400
        
        new_recipient_email = data.get('new_recipient_email')
        if not new_recipient_email:
            return jsonify({'error': 'New recipient email is required'}), 400
        
        new_recipient = User.query.filter_by(email=new_recipient_email, user_type='recipient').first()
        if not new_recipient:
            return jsonify({'error': 'Recipient not found'}), 404
        
        # Store old recipient for notification
        old_recipient_id = voucher.recipient_id
        
        # Reassign voucher
        voucher.recipient_id = new_recipient.id
        voucher.status = 'reassigned'
        db.session.commit()
        
        # Create new active voucher for new recipient
        new_voucher = Voucher(
            code=voucher.code,
            value=voucher.value,
            recipient_id=new_recipient.id,
            issued_by=voucher.issued_by,
            expiry_date=voucher.expiry_date,
            status='active'
        )
        db.session.add(new_voucher)
        db.session.commit()
        
        # Notify new recipient
        create_notification(
            new_recipient.id,
            'Voucher Assigned',
            f'You have been assigned voucher {voucher.code} for £{voucher.value:.2f}',
            'success'
        )
        
        # Notify old recipient if exists
        if old_recipient_id:
            create_notification(
                old_recipient_id,
                'Voucher Reassigned',
                f'Voucher {voucher.code} has been reassigned',
                'info'
            )
        
        return jsonify({
            'message': 'Voucher reassigned successfully',
            'new_voucher_id': new_voucher.id
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to reassign voucher: {str(e)}'}), 500

# ============================================
# Surplus Food Enhanced Routes
# ============================================

@app.route('/api/items/claim', methods=['POST'])
def claim_surplus_item():
    """VCFSE organization claims a surplus food item"""
    try:
        data = request.get_json()
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vcse':
            return jsonify({'error': 'Only VCFSE organizations can claim items'}), 403
        
        item_id = data.get('item_id')
        if not item_id:
            return jsonify({'error': 'Item ID is required'}), 400
        
        item = Item.query.get(item_id)
        if not item:
            return jsonify({'error': 'Item not found'}), 404
        
        if item.status != 'available':
            return jsonify({'error': 'Item is not available'}), 400
        
        # Claim the item
        item.status = 'claimed'
        item.claimed_by = user_id
        item.claimed_at = datetime.utcnow()
        
        # Set collection time limit (e.g., 2 hours from now)
        item.collection_time_limit = datetime.utcnow() + timedelta(hours=2)
        
        db.session.commit()
        
        # Notify vendor
        create_notification(
            item.vendor_id,
            'Surplus Item Claimed',
            f'{user.organization_name} has claimed your surplus item: {item.title}. Collection by: {item.collection_time_limit.strftime("%H:%M")}',
            'success'
        )
        
        # Send SMS to vendor about collection
        vendor = User.query.get(item.vendor_id)
        if vendor and vendor.phone:
            # Get shop info if available
            shop = VendorShop.query.get(item.shop_id) if hasattr(item, 'shop_id') else None
            shop_name = shop.shop_name if shop else vendor.shop_name or f"{vendor.first_name} {vendor.last_name}"
            
            sms_result = sms_service.send_collection_notification(
                vendor.phone,
                shop_name,
                user.organization_name or f"{user.first_name} {user.last_name}",
                item.title,
                item.quantity if hasattr(item, 'quantity') else 1
            )
            if not sms_result.get('success'):
                print(f"Failed to send SMS to vendor: {sms_result.get('error')}")
        
        # Send email to vendor about collection
        if vendor and vendor.email:
            vcse_contact = user.phone or user.email
            email_result = email_service.send_collection_confirmation_email(
                vendor.email,
                shop_name,
                item.title,
                user.organization_name or f"{user.first_name} {user.last_name}",
                vcse_contact
            )
            if not email_result:
                print(f"Failed to send email to vendor: {vendor.email}")
        
        # Notify VCFSE
        create_notification(
            user_id,
            'Item Claimed Successfully',
            f'You have claimed {item.title}. Please collect by {item.collection_time_limit.strftime("%H:%M")}',
            'success'
        )
        
        return jsonify({
            'message': 'Item claimed successfully',
            'collection_deadline': item.collection_time_limit.isoformat()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to claim item: {str(e)}'}), 500

@app.route('/api/items/<int:item_id>/collect', methods=['POST'])
def mark_item_collected(item_id):
    """Mark a surplus item as collected"""
    try:
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        item = Item.query.get(item_id)
        if not item:
            return jsonify({'error': 'Item not found'}), 404
        
        user = User.query.get(user_id)
        
        # Either vendor or VCFSE can mark as collected
        if user.user_type not in ['vendor', 'vcse']:
            return jsonify({'error': 'Unauthorized'}), 403
        
        if user.user_type == 'vendor' and item.vendor_id != user_id:
            return jsonify({'error': 'Not your item'}), 403
        
        if user.user_type == 'vcse' and item.claimed_by != user_id:
            return jsonify({'error': 'Not your claimed item'}), 403
        
        if item.status != 'claimed':
            return jsonify({'error': 'Item is not in claimed status'}), 400
        
        # Mark as collected
        item.status = 'collected'
        item.collected_at = datetime.utcnow()
        db.session.commit()
        
        # Notify both parties
        create_notification(
            item.vendor_id,
            'Item Collected',
            f'Surplus item "{item.title}" has been collected',
            'success'
        )
        
        if item.claimed_by:
            create_notification(
                item.claimed_by,
                'Collection Confirmed',
                f'Collection of "{item.title}" has been confirmed',
                'success'
            )
        
        return jsonify({'message': 'Item marked as collected'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to mark item as collected: {str(e)}'}), 500

@app.route('/api/items/available', methods=['GET'])
def get_available_surplus_items():
    """Get all available surplus food items"""
    try:
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        items = Item.query.filter_by(status='available').order_by(Item.created_at.desc()).all()
        
        items_list = []
        for item in items:
            vendor = User.query.get(item.vendor_id)
            category = Category.query.get(item.category_id)
            
            items_list.append({
                'id': item.id,
                'title': item.title,
                'description': item.description,
                'category': category.name if category else 'Unknown',
                'category_icon': category.icon if category else '📦',
                'quantity': item.quantity,
                'unit': item.unit,
                'expiry_date': item.expiry_date.isoformat() if item.expiry_date else None,
                'pickup_location': item.pickup_location,
                'pickup_instructions': item.pickup_instructions,
                'vendor_name': vendor.shop_name or vendor.organization_name,
                'vendor_phone': vendor.phone,
                'created_at': item.created_at.isoformat()
            })
        
        return jsonify({'items': items_list}), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get items: {str(e)}'}), 500

# ============================================
# Reporting Routes
# ============================================

@app.route('/api/admin/reports/generate', methods=['POST'])
def admin_generate_report():
    """Generate comprehensive admin report"""
    try:
        data = request.get_json()
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Only admins can generate reports'}), 403
        
        # Get date range
        date_from = datetime.strptime(data.get('date_from', '2024-01-01'), '%Y-%m-%d').date()
        date_to = datetime.strptime(data.get('date_to', datetime.utcnow().strftime('%Y-%m-%d')), '%Y-%m-%d').date()
        
        # Calculate statistics
        total_vouchers = Voucher.query.filter(
            Voucher.created_at >= datetime.combine(date_from, datetime.min.time()),
            Voucher.created_at <= datetime.combine(date_to, datetime.max.time())
        ).count()
        
        redeemed_vouchers = Voucher.query.filter(
            Voucher.status == 'redeemed',
            Voucher.redeemed_at >= datetime.combine(date_from, datetime.min.time()),
            Voucher.redeemed_at <= datetime.combine(date_to, datetime.max.time())
        ).count()
        
        total_value = db.session.query(db.func.sum(Voucher.value)).filter(
            Voucher.created_at >= datetime.combine(date_from, datetime.min.time()),
            Voucher.created_at <= datetime.combine(date_to, datetime.max.time())
        ).scalar() or 0
        
        redeemed_value = db.session.query(db.func.sum(Voucher.value)).filter(
            Voucher.status == 'redeemed',
            Voucher.redeemed_at >= datetime.combine(date_from, datetime.min.time()),
            Voucher.redeemed_at <= datetime.combine(date_to, datetime.max.time())
        ).scalar() or 0
        
        total_vendors = User.query.filter_by(user_type='vendor', is_active=True).count()
        total_vcse = User.query.filter_by(user_type='vcse', is_active=True).count()
        total_recipients = User.query.filter_by(user_type='recipient', is_active=True).count()
        
        surplus_items_collected = Item.query.filter(
            Item.status == 'collected',
            Item.collected_at >= datetime.combine(date_from, datetime.min.time()),
            Item.collected_at <= datetime.combine(date_to, datetime.max.time())
        ).count()
        
        report_data = {
            'date_from': date_from.isoformat(),
            'date_to': date_to.isoformat(),
            'voucher_statistics': {
                'total_issued': total_vouchers,
                'total_redeemed': redeemed_vouchers,
                'redemption_rate': (redeemed_vouchers / total_vouchers * 100) if total_vouchers > 0 else 0,
                'total_value': total_value,
                'redeemed_value': redeemed_value
            },
            'user_statistics': {
                'total_vendors': total_vendors,
                'total_vcse': total_vcse,
                'total_recipients': total_recipients
            },
            'surplus_food': {
                'items_collected': surplus_items_collected
            }
        }
        
        return jsonify({
            'report': report_data,
            'generated_at': datetime.utcnow().isoformat()
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to generate report: {str(e)}'}), 500

@app.route('/api/vcse/reports/generate', methods=['POST'])
def vcse_generate_report():
    """Generate VCFSE-specific report for funders"""
    try:
        data = request.get_json()
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vcse':
            return jsonify({'error': 'Only VCFSE organizations can generate reports'}), 403
        
        # Get date range
        date_from = datetime.strptime(data.get('date_from', '2024-01-01'), '%Y-%m-%d').date()
        date_to = datetime.strptime(data.get('date_to', datetime.utcnow().strftime('%Y-%m-%d')), '%Y-%m-%d').date()
        
        # Calculate VCFSE-specific statistics
        vouchers_issued = Voucher.query.filter(
            Voucher.issued_by == user_id,
            Voucher.created_at >= datetime.combine(date_from, datetime.min.time()),
            Voucher.created_at <= datetime.combine(date_to, datetime.max.time())
        ).all()
        
        total_issued = len(vouchers_issued)
        total_value_issued = sum(v.value for v in vouchers_issued)
        
        redeemed_vouchers = [v for v in vouchers_issued if v.status == 'redeemed']
        total_redeemed = len(redeemed_vouchers)
        total_value_redeemed = sum(v.value for v in redeemed_vouchers)
        
        # Get unique recipients served
        recipients_served = len(set(v.recipient_id for v in vouchers_issued if v.recipient_id))
        
        # Get surplus items claimed
        items_claimed = Item.query.filter(
            Item.claimed_by == user_id,
            Item.claimed_at >= datetime.combine(date_from, datetime.min.time()),
            Item.claimed_at <= datetime.combine(date_to, datetime.max.time())
        ).count()
        
        items_collected = Item.query.filter(
            Item.claimed_by == user_id,
            Item.status == 'collected',
            Item.collected_at >= datetime.combine(date_from, datetime.min.time()),
            Item.collected_at <= datetime.combine(date_to, datetime.max.time())
        ).count()
        
        report_data = {
            'organization_name': user.organization_name,
            'date_from': date_from.isoformat(),
            'date_to': date_to.isoformat(),
            'voucher_activity': {
                'vouchers_issued': total_issued,
                'vouchers_redeemed': total_redeemed,
                'redemption_rate': (total_redeemed / total_issued * 100) if total_issued > 0 else 0,
                'total_value_issued': total_value_issued,
                'total_value_redeemed': total_value_redeemed,
                'recipients_served': recipients_served
            },
            'surplus_food_activity': {
                'items_claimed': items_claimed,
                'items_collected': items_collected
            },
            'financial_summary': {
                'current_balance': user.balance,
                'total_spent': total_value_issued
            }
        }
        
        return jsonify({
            'report': report_data,
            'generated_at': datetime.utcnow().isoformat()
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to generate report: {str(e)}'}), 500



# General Routes (accessible to all authenticated users)

@app.route('/api/shops', methods=['GET'])
def get_all_shops():
    """Get all shops for recipients to view where they can redeem vouchers"""
    try:
        shops = VendorShop.query.filter_by(is_active=True).all()
        
        shops_data = []
        for shop in shops:
            vendor = User.query.get(shop.vendor_id)
            shops_data.append({
                'id': shop.id,
                'name': shop.shop_name,
                'address': shop.address,
                'city': shop.city,
                'postcode': shop.postcode,
                'phone': shop.phone,
                'vendor_name': f"{vendor.first_name} {vendor.last_name}" if vendor else "Unknown",
                'accepts_vouchers': True
            })
        
        return jsonify({'shops': shops_data}), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to load shops: {str(e)}'}), 500

@app.route('/api/user/vouchers', methods=['GET'])
def get_user_vouchers():
    """Get vouchers for the current user"""
    try:
        # In a real app, you'd get user_id from session/token
        # For now, we'll accept it as a query parameter
        user_id = request.args.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'User ID required'}), 400
        
        vouchers = Voucher.query.filter_by(recipient_id=user_id).all()
        
        vouchers_data = []
        for voucher in vouchers:
            vouchers_data.append({
                'id': voucher.id,
                'code': voucher.code,
                'value': float(voucher.value),
                'status': voucher.status,
                'issued_date': voucher.issued_date.isoformat() if voucher.issued_date else None,
                'expiry_date': voucher.expiry_date.isoformat() if voucher.expiry_date else None,
                'redeemed_date': voucher.redeemed_date.isoformat() if voucher.redeemed_date else None
            })
        
        return jsonify({'vouchers': vouchers_data}), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to load vouchers: {str(e)}'}), 500

@app.route('/api/user/profile', methods=['GET'])
def get_user_profile():
    """Get user profile information"""
    try:
        user_id = request.args.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'User ID required'}), 400
        
        user = User.query.get(user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        profile_data = {
            'id': user.id,
            'email': user.email,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'phone': user.phone,
            'user_type': user.user_type,
            'organization_name': user.organization_name,
            'shop_name': user.shop_name,
            'address': user.address,
            'postcode': user.postcode,
            'city': user.city,
            'balance': float(user.balance) if user.user_type == 'vcse' else None,
            'created_at': user.created_at.isoformat() if user.created_at else None
        }
        
        return jsonify({'profile': profile_data}), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to load profile: {str(e)}'}), 500

@app.route('/api/vouchers/redeem', methods=['POST'])
def redeem_voucher():
    """Redeem a voucher at a shop"""
    try:
        data = request.get_json()
        
        voucher_code = data.get('code') or data.get('voucher_code')  # Accept both 'code' and 'voucher_code'
        shop_id = data.get('shop_id')
        vendor_id = data.get('vendor_id')
        
        if not voucher_code or not shop_id:
            return jsonify({'error': 'Voucher code and shop ID required'}), 400
        
        voucher = Voucher.query.filter_by(code=voucher_code).first()
        
        if not voucher:
            return jsonify({'error': 'Voucher not found'}), 404
        
        if voucher.status != 'active':
            return jsonify({'error': f'Voucher is {voucher.status}'}), 400
        
        if voucher.expiry_date and voucher.expiry_date < datetime.utcnow():
            voucher.status = 'expired'
            db.session.commit()
            return jsonify({'error': 'Voucher has expired'}), 400
        
        # Redeem the voucher
        voucher.status = 'redeemed'
        voucher.redeemed_date = datetime.utcnow()
        voucher.redeemed_at_shop_id = shop_id
        
        db.session.commit()
        
        # Create notification for recipient
        create_notification(
            voucher.recipient_id,
            'Voucher Redeemed',
            f'Your voucher {voucher_code} worth £{voucher.value} has been redeemed.',
            'success'
        )
        
        return jsonify({
            'message': 'Voucher redeemed successfully',
            'value': float(voucher.value),
            'code': voucher.code,
            'redeemed_date': voucher.redeemed_date.isoformat()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to redeem voucher: {str(e)}'}), 500

@app.route('/api/notifications', methods=['GET'])
def get_notifications():
    """Get notifications for a user"""
    try:
        user_id = request.args.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'User ID required'}), 400
        
        notifications = Notification.query.filter_by(user_id=user_id).order_by(Notification.created_at.desc()).limit(50).all()
        
        notifications_data = []
        for notif in notifications:
            notifications_data.append({
                'id': notif.id,
                'title': notif.title,
                'message': notif.message,
                'type': notif.type,
                'is_read': notif.is_read,
                'created_at': notif.created_at.isoformat() if notif.created_at else None
            })
        
        return jsonify({'notifications': notifications_data}), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to load notifications: {str(e)}'}), 500

@app.route('/api/notifications/<int:notification_id>/read', methods=['POST'])
def mark_notification_read(notification_id):
    """Mark a notification as read"""
    try:
        notification = Notification.query.get(notification_id)
        
        if not notification:
            return jsonify({'error': 'Notification not found'}), 404
        
        notification.is_read = True
        db.session.commit()
        
        return jsonify({'message': 'Notification marked as read'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to mark notification as read: {str(e)}'}), 500

# Vendor Routes - Post Surplus Food
@app.route('/api/items/post', methods=['POST'])
def post_surplus_item():
    """Vendor posts a surplus food item"""
    try:
        data = request.get_json()
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        # Validate required fields
        required_fields = ['item_name', 'quantity', 'category', 'shop_name', 'shop_address']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'error': f'Please complete your shop profile before posting items. Missing: {field}'}), 400
        
        # Validate shop_name is not just whitespace
        if not data['shop_name'].strip() or not data['shop_address'].strip():
            return jsonify({'error': 'Please set up your shop profile with valid shop name and address before posting items.'}), 400
        
        # Find or create shop for this vendor
        shop = VendorShop.query.filter_by(
            vendor_id=user_id,
            shop_name=data['shop_name']
        ).first()
        
        if not shop:
            # Create new shop
            try:
                shop = VendorShop(
                    vendor_id=user_id,
                    shop_name=data['shop_name'].strip(),
                    address=data['shop_address'].strip(),
                    postcode='',
                    city='',
                    phone='',
                    is_active=True
                )
                db.session.add(shop)
                db.session.flush()  # Get the shop ID
            except Exception as shop_error:
                db.session.rollback()
                return jsonify({'error': f'Failed to create shop profile: {str(shop_error)}'}), 500
        
        # Parse expiry date if provided
        expiry_date = None
        if data.get('expiry_date'):
            try:
                from datetime import datetime as dt
                expiry_date = dt.strptime(data['expiry_date'], '%Y-%m-%d').date()
            except ValueError:
                pass
        
        # Create surplus item
        new_item = SurplusItem(
            vendor_id=user_id,
            shop_id=shop.id,
            item_name=data['item_name'],
            quantity=data['quantity'],
            category=data['category'],
            description=data.get('description', ''),
            expiry_date=expiry_date,
            item_type=data.get('item_type', 'free'),  # 'free' or 'discount'
            price=data.get('price') if data.get('item_type') == 'discount' else None,
            original_price=data.get('original_price') if data.get('item_type') == 'discount' else None,
            status='available',
            posted_at=datetime.now()
        )
        
        db.session.add(new_item)
        db.session.commit()
        
        # Broadcast real-time notification via WebSocket
        from notifications_system import broadcast_new_item_notification
        broadcast_new_item_notification(
            socketio,
            item_type=data.get('item_type', 'free'),
            shop_id=shop.id,
            item_id=new_item.id,
            item_name=data['item_name'],
            shop_name=shop.shop_name,
            quantity=data['quantity']
        )
        
        print(f"Surplus item posted: {data['item_name']} at {shop.shop_name}")
        
        return jsonify({
            'message': 'To Go item posted successfully',
            'item_id': new_item.id,
            'item_name': new_item.item_name,
            'shop': shop.shop_name
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to post surplus item: {str(e)}'}), 500

# ============================================
# VENDOR SURPLUS ITEMS ENDPOINTS
# ============================================

@app.route('/api/vendor/surplus-items', methods=['GET'])
def get_vendor_surplus_items():
    """Get all surplus items posted by the logged-in vendor"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vendor':
            return jsonify({'error': 'Vendor access required'}), 403
        
        # Get all surplus items posted by this vendor
        surplus_items = SurplusItem.query.filter_by(vendor_id=user_id).order_by(SurplusItem.posted_at.desc()).all()
        
        items_list = []
        for item in surplus_items:
            shop = VendorShop.query.get(item.shop_id)
            items_list.append({
                'id': item.id,
                'item_name': item.item_name,
                'quantity': item.quantity,
                'category': item.category,
                'description': item.description,
                'status': item.status,
                'shop_name': shop.shop_name if shop else 'Unknown',
                'posted_at': item.posted_at.isoformat() if item.posted_at else None
            })
        
        return jsonify({
            'surplus_items': items_list,
            'total_count': len(items_list)
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get surplus items: {str(e)}'}), 500

@app.route('/api/vendor/to-go-items', methods=['GET'])
def get_vendor_to_go_items():
    """Alias for get_vendor_surplus_items - Get all to-go items posted by the logged-in vendor"""
    return get_vendor_surplus_items()

@app.route('/api/vendor/surplus-items/<int:item_id>', methods=['PUT'])
def update_surplus_item(item_id):
    """Update a surplus item"""
    try:
        data = request.get_json()
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vendor':
            return jsonify({'error': 'Vendor access required'}), 403
        
        item = SurplusItem.query.filter_by(id=item_id, vendor_id=user_id).first()
        if not item:
            return jsonify({'error': 'Item not found'}), 404
        
        # Update item fields
        if 'item_name' in data:
            item.item_name = data['item_name']
        if 'quantity' in data:
            item.quantity = data['quantity']
        if 'category' in data:
            item.category = data['category']
        if 'description' in data:
            item.description = data['description']
        if 'expiry_date' in data:
            try:
                from datetime import datetime as dt
                item.expiry_date = dt.strptime(data['expiry_date'], '%Y-%m-%d').date() if data['expiry_date'] else None
            except ValueError:
                pass
        
        db.session.commit()
        
        return jsonify({
            'message': 'Item updated successfully',
            'item': {
                'id': item.id,
                'item_name': item.item_name,
                'quantity': item.quantity,
                'category': item.category,
                'description': item.description,
                'expiry_date': item.expiry_date.isoformat() if item.expiry_date else None
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to update item: {str(e)}'}), 500

@app.route('/api/vendor/surplus-items/<int:item_id>', methods=['DELETE'])
def delete_surplus_item(item_id):
    """Delete a surplus item"""
    try:
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vendor':
            return jsonify({'error': 'Vendor access required'}), 403
        
        item = SurplusItem.query.filter_by(id=item_id, vendor_id=user_id).first()
        if not item:
            return jsonify({'error': 'Item not found'}), 404
        
        # Mark as unavailable instead of deleting
        item.status = 'removed'
        db.session.commit()
        
        return jsonify({'message': 'Item deleted successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to delete item: {str(e)}'}), 500

# Run the application
# ============================================
# FUND ALLOCATION ENDPOINTS
# ============================================

@app.route('/api/admin/vcse-organizations', methods=['GET'])
def get_vcse_organizations():
    """Get all VCFSE organizations with their balances"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Get all VCFSE organizations
        vcse_orgs = User.query.filter_by(user_type='vcse').all()
        
        result = []
        for org in vcse_orgs:
            result.append({
                'id': org.id,
                'name': f"{org.first_name} {org.last_name}",
                'organization_name': org.organization_name,
                'email': org.email,
                'charity_commission_number': org.charity_commission_number if hasattr(org, 'charity_commission_number') else '',
                'allocated_balance': float(org.allocated_balance) if org.allocated_balance else 0.0,
                'created_at': org.created_at.isoformat() if org.created_at else None
            })
        
        return jsonify(result), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get VCFSE organizations: {str(e)}'}), 500


@app.route('/api/admin/schools', methods=['GET'])
def get_schools():
    """Get all School/Care Organizations with their balances"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Get all School/Care Organizations
        schools = User.query.filter_by(user_type='school').all()
        
        result = []
        for school in schools:
            result.append({
                'id': school.id,
                'name': f"{school.first_name} {school.last_name}",
                'organization_name': school.organization_name,
                'email': school.email,
                'allocated_balance': float(school.allocated_balance) if school.allocated_balance else 0.0,
                'created_at': school.created_at.isoformat() if school.created_at else None
            })
        
        return jsonify(result), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get schools: {str(e)}'}), 500


@app.route('/api/admin/recipients', methods=['GET'])
def get_recipients():
    """Get all recipients with their voucher information"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Get all recipients
        recipients = User.query.filter_by(user_type='recipient').all()
        
        result = []
        for recipient in recipients:
            # Get voucher statistics for this recipient
            vouchers = Voucher.query.filter_by(recipient_id=recipient.id).all()
            total_vouchers = len(vouchers)
            active_vouchers = len([v for v in vouchers if v.status == 'active'])
            redeemed_vouchers = len([v for v in vouchers if v.status == 'redeemed'])
            total_value = sum([float(v.value) for v in vouchers if v.status == 'active'])
            
            result.append({
                'id': recipient.id,
                'first_name': recipient.first_name,
                'last_name': recipient.last_name,
                'name': f"{recipient.first_name} {recipient.last_name}",
                'email': recipient.email,
                'phone': recipient.phone if hasattr(recipient, 'phone') else '',
                'address': recipient.address if hasattr(recipient, 'address') else '',
                'city': recipient.city if hasattr(recipient, 'city') else '',
                'postcode': recipient.postcode if hasattr(recipient, 'postcode') else '',
                'date_of_birth': recipient.date_of_birth.isoformat() if hasattr(recipient, 'date_of_birth') and recipient.date_of_birth else None,
                'created_at': recipient.created_at.isoformat() if recipient.created_at else None,
                'total_vouchers': total_vouchers,
                'active_vouchers': active_vouchers,
                'redeemed_vouchers': redeemed_vouchers,
                'total_active_value': float(total_value)
            })
        
        return jsonify(result), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get recipients: {str(e)}'}), 500


@app.route('/api/admin/allocate-funds', methods=['POST'])
def allocate_funds():
    """Admin allocates funds to a VCFSE organization or School"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        data = request.get_json()
        recipient_id = data.get('vcse_id') or data.get('school_id')  # Support both
        amount = data.get('amount')
        notes = data.get('notes', '')
        
        if not recipient_id or not amount:
            return jsonify({'error': 'Recipient ID and amount are required'}), 400
        
        try:
            amount = float(amount)
            if amount <= 0:
                return jsonify({'error': 'Amount must be positive'}), 400
        except ValueError:
            return jsonify({'error': 'Invalid amount'}), 400
        
        # Get recipient organization (VCFSE or School)
        recipient = User.query.get(recipient_id)
        if not recipient or recipient.user_type not in ['vcse', 'school']:
            return jsonify({'error': 'Organization not found'}), 404
        
        # Update allocated balance (funds from admin)
        current_allocated = float(recipient.allocated_balance) if recipient.allocated_balance else 0.0
        recipient.allocated_balance = current_allocated + amount
        
        # Create allocation record (TODO: Create FundAllocation model)
        # allocation = FundAllocation(
        #     admin_id=user_id,
        #     vcse_id=recipient_id if recipient.user_type == 'vcse' else None,
        #     amount=amount,
        #     notes=notes
        # )
        # db.session.add(allocation)
        db.session.commit()
        
        # Send email notification (TODO: Implement send_fund_allocation_email function)
        # send_fund_allocation_email(recipient.email, recipient.first_name, amount, current_allocated + amount)
        
        org_name = recipient.organization_name if recipient.organization_name else f"{recipient.first_name} {recipient.last_name}"
        
        return jsonify({
            'message': 'Funds allocated successfully',
            'organization_name': org_name,
            'organization_type': recipient.user_type,
            'amount': amount,
            'new_allocated_balance': float(recipient.allocated_balance)
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to allocate funds: {str(e)}'}), 500


@app.route('/api/admin/fund-allocations', methods=['GET'])
def get_fund_allocations():
    """Get all fund allocation history"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        allocations = FundAllocation.query.order_by(FundAllocation.allocated_at.desc()).all()
        
        result = []
        for allocation in allocations:
            admin = User.query.get(allocation.admin_id)
            vcse = User.query.get(allocation.vcse_id)
            
            result.append({
                'id': allocation.id,
                'admin_name': f"{admin.first_name} {admin.last_name}" if admin else "Unknown",
                'vcse_name': f"{vcse.first_name} {vcse.last_name}" if vcse else "Unknown",
                'vcse_email': vcse.email if vcse else "",
                'amount': float(allocation.amount),
                'allocated_at': allocation.allocated_at.isoformat() if allocation.allocated_at else None,
                'notes': allocation.notes
            })
        
        return jsonify(result), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get fund allocations: {str(e)}'}), 500


@app.route('/api/vcse/balance', methods=['GET'])
def get_vcse_balance():
    """Get current VCFSE organization balance"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vcse':
            return jsonify({'error': 'VCFSE access required'}), 403
        
        balance = float(user.balance) if user.balance else 0.0
        allocated_balance = float(user.allocated_balance) if user.allocated_balance else 0.0
        
        return jsonify({
            'balance': balance,
            'allocated_balance': allocated_balance,
            'organization_name': f"{user.first_name} {user.last_name}"
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get balance: {str(e)}'}), 500


@app.route('/api/admin/vouchers', methods=['GET'])
def get_all_vouchers():
    """Get all vouchers for admin dashboard"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        vouchers = Voucher.query.order_by(Voucher.created_at.desc()).all()
        
        vouchers_data = []
        for v in vouchers:
            recipient = User.query.get(v.recipient_id) if v.recipient_id else None
            issuer = User.query.get(v.issued_by) if v.issued_by else None
            
            vouchers_data.append({
                'id': v.id,
                'code': v.code,
                'value': float(v.value),
                'status': v.status,
                'created_at': v.created_at.isoformat() if v.created_at else None,
                'expiry_date': v.expiry_date.isoformat() if v.expiry_date else None,
                'recipient': {
                    'name': f"{recipient.first_name} {recipient.last_name}" if recipient else 'N/A',
                    'email': recipient.email if recipient else 'N/A',
                    'phone': recipient.phone if recipient else 'N/A',
                    'address': recipient.address if recipient else 'N/A'
                } if recipient else None,
                'issued_by': {
                    'name': f"{issuer.first_name} {issuer.last_name}" if issuer else 'N/A',
                    'organization': issuer.organization_name if issuer else 'N/A'
                } if issuer else None
            })
        
        return jsonify({
            'vouchers': vouchers_data,
            'total_count': len(vouchers_data)
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get vouchers: {str(e)}'}), 500


@app.route('/api/admin/shops', methods=['GET'])
def admin_get_all_shops():
    """Admin endpoint to view all shops with surplus foods"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Get all vendor shops
        shops = VendorShop.query.all()
        
        shops_data = []
        for shop in shops:
            vendor = User.query.get(shop.vendor_id)
            
            # Count surplus items for this shop
            surplus_count = SurplusItem.query.filter_by(shop_id=shop.id).count()
            
            shops_data.append({
                'id': shop.id,
                'shop_name': shop.shop_name,
                'address': shop.address,
                'city': shop.city,
                'postcode': shop.postcode,
                'phone': shop.phone,
                'surplus_items_count': surplus_count,
                'vendor_name': f"{vendor.first_name} {vendor.last_name}" if vendor else 'Unknown',
                'vendor_email': vendor.email if vendor else 'N/A',
                'created_at': shop.created_at.isoformat() if shop.created_at else None
            })
        
        return jsonify({
            'shops': shops_data,
            'total_count': len(shops_data)
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get shops: {str(e)}'}), 500


@app.route('/api/admin/surplus-items', methods=['GET'])
def admin_get_all_surplus_items():
    """Admin endpoint to view all surplus items across all vendors"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Get all surplus items
        items = SurplusItem.query.order_by(SurplusItem.posted_at.desc()).all()
        
        items_data = []
        for item in items:
            shop = VendorShop.query.get(item.shop_id)
            vendor = User.query.get(shop.vendor_id) if shop else None
            
            items_data.append({
                'id': item.id,
                'item_name': item.item_name,
                'quantity': item.quantity,
                'unit': item.unit,
                'category': item.category,
                'price': float(item.price) if item.price else 0.0,
                'description': item.description,
                'status': item.status,
                'shop_name': shop.shop_name if shop else 'Unknown',
                'shop_address': shop.address if shop else 'N/A',
                'vendor_name': f"{vendor.first_name} {vendor.last_name}" if vendor else 'Unknown',
                'created_at': item.posted_at.isoformat() if item.posted_at else None
            })
        
        return jsonify({
            'items': items_data,
            'total_count': len(items_data)
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get surplus items: {str(e)}'}), 500


@app.route('/api/admin/to-go-items', methods=['GET'])
def admin_get_to_go_items():
    """Alias for admin_get_all_surplus_items - Get all to-go items across all vendors"""
    return admin_get_all_surplus_items()


@app.route('/api/vcse/to-go-items', methods=['GET'])
def vcse_get_to_go_items():
    """VCFSE endpoint to view free surplus items available for collection"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vcse':
            return jsonify({'error': 'VCFSE access required'}), 403
        
        # Get only FREE surplus items that are available
        items = SurplusItem.query.filter_by(
            item_type='free',
            status='available'
        ).order_by(SurplusItem.posted_at.desc()).all()
        
        items_data = []
        for item in items:
            shop = VendorShop.query.get(item.shop_id)
            vendor = User.query.get(shop.vendor_id) if shop else None
            
            items_data.append({
                'id': item.id,
                'item_name': item.item_name,
                'quantity': item.quantity,
                'unit': item.unit,
                'category': item.category,
                'description': item.description,
                'status': item.status,
                'expiry_date': item.expiry_date.isoformat() if item.expiry_date else None,
                'shop_name': shop.shop_name if shop else 'Unknown',
                'shop_address': shop.address if shop else 'N/A',
                'shop_phone': shop.phone if shop else 'N/A',
                'vendor_name': f"{vendor.first_name} {vendor.last_name}" if vendor else 'Unknown',
                'created_at': item.posted_at.isoformat() if item.posted_at else None
            })
        
        return jsonify({
            'items': items_data,
            'total_count': len(items_data)
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get to-go items: {str(e)}'}), 500


@app.route('/api/recipient/shops', methods=['GET'])
def get_shops_for_recipient():
    """Recipient endpoint to view all participating shops with optional town filtering"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'recipient':
            return jsonify({'error': 'Recipient access required'}), 403
        
        # Get optional town filter from query parameters
        town_filter = request.args.get('town', None)
        
        # Get vendor shops with optional town filtering
        query = VendorShop.query
        if town_filter and town_filter != 'all':
            query = query.filter_by(town=town_filter)
        
        shops = query.all()
        
        shops_data = []
        for shop in shops:
            # Count available surplus items for this shop
            surplus_count = SurplusItem.query.filter_by(
                shop_id=shop.id,
                status='available'
            ).count()
            
            shops_data.append({
                'id': shop.id,
                'shop_name': shop.shop_name,
                'address': shop.address,
                'city': shop.city,
                'town': shop.town,
                'postcode': shop.postcode,
                'phone': shop.phone,
                'surplus_items_count': surplus_count
            })
        
        return jsonify({
            'shops': shops_data,
            'total_count': len(shops_data)
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get shops: {str(e)}'}), 500


@app.route('/api/recipient/surplus-items', methods=['GET'])
def get_surplus_items_for_recipient():
    """Recipient endpoint to view all available surplus items"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'recipient':
            return jsonify({'error': 'Recipient access required'}), 403
        
        # Get all available surplus items
        items = SurplusItem.query.filter_by(status='available').order_by(
            SurplusItem.posted_at.desc()
        ).all()
        
        items_data = []
        for item in items:
            shop = VendorShop.query.get(item.shop_id)
            
            items_data.append({
                'id': item.id,
                'item_name': item.item_name,
                'quantity': item.quantity,
                'unit': item.unit,
                'category': item.category,
                'price': float(item.price) if item.price else 0.0,
                'description': item.description,
                'shop_name': shop.shop_name if shop else 'Unknown',
                'shop_address': shop.address if shop else 'N/A',
                'shop_phone': shop.phone if shop else 'N/A',
                'created_at': item.created_at.isoformat() if item.created_at else None
            })
        
        return jsonify({
            'items': items_data,
            'total_count': len(items_data)
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get surplus items: {str(e)}'}), 500


# School/Care Organization Routes
@app.route('/api/school/balance', methods=['GET'])
def get_school_balance():
    """Get current school/care organization balance"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'school':
            return jsonify({'error': 'School/Care Organization access required'}), 403
        
        allocated_balance = float(user.allocated_balance) if user.allocated_balance else 0.0
        
        return jsonify({
            'allocated_balance': allocated_balance,
            'organization_name': user.organization_name
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get balance: {str(e)}'}), 500

@app.route('/api/school/issue-voucher', methods=['POST'])
def school_issue_voucher():
    """School/Care Organization issues a voucher to a family"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'school':
            return jsonify({'error': 'School/Care Organization access required'}), 403
        
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['recipient_email', 'amount', 'recipient_first_name', 'recipient_last_name', 'recipient_phone', 'recipient_address']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        amount = float(data['amount'])
        assign_shop_method = data.get('assign_shop_method', 'specific_shop')  # 'specific_shop' or 'recipient_to_choose'
        selected_shops = data.get('selected_shops')  # List of shop IDs or 'all'
        
        # Enforce £50 maximum per voucher - split into multiple vouchers if needed
        MAX_VOUCHER_VALUE = 50.0
        num_vouchers = int(amount / MAX_VOUCHER_VALUE)
        remainder = amount % MAX_VOUCHER_VALUE
        
        voucher_amounts = [MAX_VOUCHER_VALUE] * num_vouchers
        if remainder > 0:
            voucher_amounts.append(remainder)
        
        # Check if school has sufficient wallet balance
        if user.balance < amount:
            return jsonify({'error': f'Insufficient wallet balance. Current balance: £{user.balance:.2f}, Required: £{amount:.2f}'}), 400
        
        # Find or create recipient
        recipient = User.query.filter_by(email=data['recipient_email']).first()
        
        if not recipient:
            # Create new recipient account
            from werkzeug.security import generate_password_hash
            import secrets
            
            temp_password = secrets.token_urlsafe(12)
            recipient = User(
                email=data['recipient_email'],
                password_hash=generate_password_hash(temp_password),
                first_name=data['recipient_first_name'],
                last_name=data['recipient_last_name'],
                phone=data['recipient_phone'],
                address=data['recipient_address'],
                user_type='recipient',
                is_verified=True,
                is_active=True
            )
            db.session.add(recipient)
            db.session.flush()
        
        # Generate unique voucher codes and create multiple vouchers
        import random
        import string
        from datetime import datetime, timedelta
        import json
        
        vendor_restrictions = None
        if selected_shops and selected_shops != 'all':
            vendor_restrictions = json.dumps(selected_shops)  # Store as JSON array
        
        # Deduct total amount from school wallet balance once
        balance_before = user.balance
        user.balance -= amount
        
        # Create wallet transaction record for total amount
        wallet_transaction = WalletTransaction(
            user_id=user_id,
            transaction_type='debit',
            amount=amount,
            balance_before=balance_before,
            balance_after=user.balance,
            description=f'{len(voucher_amounts)} voucher(s) issued to {recipient.first_name} {recipient.last_name} (Total: £{amount:.2f})',
            reference=f'BATCH_{datetime.utcnow().strftime("%Y%m%d%H%M%S")}',
            status='completed'
        )
        db.session.add(wallet_transaction)
        db.session.flush()
        
        # Create multiple vouchers based on split amounts
        voucher_codes = []
        for voucher_value in voucher_amounts:
            voucher_code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=10))
            voucher_codes.append(voucher_code)
            
            voucher = Voucher(
                code=voucher_code,
                value=voucher_value,
                issued_by=user_id,
                recipient_id=recipient.id,
                status='active',
                expiry_date=datetime.utcnow() + timedelta(days=90),
                vendor_restrictions=vendor_restrictions,
                assign_shop_method=assign_shop_method,
                original_recipient_id=recipient.id,
                reassignment_count=0,
                issued_by_user_id=user_id,
                deducted_from_wallet=True,
                wallet_transaction_id=wallet_transaction.id
            )
            db.session.add(voucher)
        
        db.session.commit()
        
        # Create notification for recipient
        voucher_list = ', '.join(voucher_codes)
        create_notification(
            recipient.id,
            'New Voucher(s) Received',
            f'You have received {len(voucher_codes)} voucher(s) worth £{amount:.2f} from {user.organization_name}. Check your vouchers page for codes.',
            'success'
        )
        
        # Send SMS notification to recipient with voucher codes
        if recipient.phone:
            if len(voucher_codes) == 1:
                sms_result = sms_service.send_voucher_code(
                    recipient.phone,
                    voucher_codes[0],
                    f"{recipient.first_name} {recipient.last_name}",
                    amount
                )
            else:
                # For multiple vouchers, send summary SMS
                sms_result = sms_service.send_sms(
                    recipient.phone,
                    f"You've received {len(voucher_codes)} vouchers worth £{amount:.2f} from {user.organization_name}. Check your email or login to view codes."
                )
            if not sms_result.get('success'):
                print(f"Failed to send SMS: {sms_result.get('error')}")
        
        # Send email notification to recipient with voucher details
        if recipient.email:
            if len(voucher_codes) == 1:
                email_result = email_service.send_voucher_issued_email(
                    recipient.email,
                    f"{recipient.first_name} {recipient.last_name}",
                    voucher_codes[0],
                    amount,
                    user.organization_name
                )
            else:
                # For multiple vouchers, send batch email
                email_result = email_service.send_batch_vouchers_email(
                    recipient.email,
                    f"{recipient.first_name} {recipient.last_name}",
                    voucher_codes,
                    voucher_amounts,
                    amount,
                    user.organization_name
                )
            if not email_result:
                print(f"Failed to send email to {recipient.email}")
        
        return jsonify({
            'message': f'{len(voucher_codes)} voucher(s) issued successfully',
            'voucher_codes': voucher_codes,
            'voucher_amounts': voucher_amounts,
            'total_amount': amount,
            'num_vouchers': len(voucher_codes),
            'recipient_email': recipient.email,
            'remaining_balance': float(user.balance)
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to issue voucher: {str(e)}'}), 500

@app.route('/api/school/vouchers', methods=['GET'])
def get_school_vouchers():
    """Get all vouchers issued by this school"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'school':
            return jsonify({'error': 'School/Care Organization access required'}), 403
        
        vouchers = Voucher.query.filter_by(issued_by=user_id).order_by(Voucher.created_at.desc()).all()
        
        vouchers_data = []
        for voucher in vouchers:
            recipient = User.query.get(voucher.recipient_id)
            vouchers_data.append({
                'id': voucher.id,
                'code': voucher.code,
                'value': float(voucher.value),
                'status': voucher.status,
                'recipient_name': f"{recipient.first_name} {recipient.last_name}" if recipient else 'Unknown',
                'recipient_email': recipient.email if recipient else 'Unknown',
                'recipient_phone': recipient.phone if recipient else 'N/A',
                'recipient_address': recipient.address if recipient else 'N/A',
                'created_at': voucher.created_at.isoformat() if voucher.created_at else None,
                'expiry_date': voucher.expiry_date.isoformat() if voucher.expiry_date else None
            })
        
        return jsonify({'vouchers': vouchers_data}), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get vouchers: {str(e)}'}), 500


@app.route('/api/school/to-go-items', methods=['GET'])
def get_school_to_go_items():
    """Get all available To Go items for schools to view and order"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'school':
            return jsonify({'error': 'School/Care Organization access required'}), 403
        
        # Schools should only see discounted items, not free items
        items = SurplusItem.query.filter_by(status='available', item_type='discount').order_by(SurplusItem.posted_at.desc()).all()
        
        items_data = []
        for item in items:
            shop = VendorShop.query.get(item.shop_id)
            vendor = User.query.get(shop.vendor_id) if shop else None
            
            items_data.append({
                'id': item.id,
                'item_name': item.item_name,
                'quantity': item.quantity,
                'unit': item.unit,
                'category': item.category,
                'price': float(item.price) if item.price else 0.0,
                'description': item.description,
                'status': item.status,
                'shop_name': shop.shop_name if shop else 'Unknown',
                'shop_address': shop.address if shop else 'N/A',
                'vendor_name': f"{vendor.first_name} {vendor.last_name}" if vendor else 'Unknown',
                'created_at': item.posted_at.isoformat() if item.posted_at else None
            })
        
        return jsonify({
            'items': items_data,
            'total_count': len(items_data)
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get to-go items: {str(e)}'}), 500


# ============================================
# Vendor Voucher Redemption Routes
# ============================================

@app.route('/api/vendor/redeem-voucher', methods=['POST'])
def vendor_redeem_voucher():
    """Vendor endpoint to redeem a voucher by code"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vendor':
            return jsonify({'error': 'Vendor access required'}), 403
        
        data = request.get_json()
        voucher_code = data.get('code', '').strip().upper()
        
        if not voucher_code:
            return jsonify({'error': 'Voucher code is required'}), 400
        
        # Find voucher by code
        voucher = Voucher.query.filter_by(code=voucher_code).first()
        
        if not voucher:
            return jsonify({'error': 'Invalid voucher code'}), 404
        
        # Check if voucher is already redeemed
        if voucher.status == 'redeemed':
            return jsonify({'error': 'Voucher has already been redeemed'}), 400
        
        # Check if voucher is expired
        if voucher.status == 'expired':
            return jsonify({'error': 'Voucher has expired'}), 400
        
        # Check expiry date
        from datetime import datetime
        if voucher.expiry_date and datetime.now().date() > voucher.expiry_date:
            voucher.status = 'expired'
            db.session.commit()
            return jsonify({'error': 'Voucher has expired'}), 400
        
        # Get recipient details
        recipient = User.query.get(voucher.recipient_id) if voucher.recipient_id else None
        
        # Redeem voucher
        voucher.status = 'redeemed'
        voucher.redeemed_at = datetime.now()
        voucher.redeemed_by_vendor = user_id
        
        # Update vendor balance
        current_balance = float(user.balance) if user.balance else 0.0
        user.balance = current_balance + float(voucher.value)
        
        db.session.commit()
        
        # Optional: Send SMS to recipient confirming redemption
        if recipient and recipient.phone:
            redemption_message = f"""BAK UP Voucher Update

Your voucher {voucher_code} (£{voucher.value:.2f}) has been redeemed at {user.shop_name or 'a vendor shop'}.

Thank you for using BAK UP!

BAK UP Team"""
            sms_result = sms_service.send_sms(recipient.phone, redemption_message)
            if not sms_result.get('success'):
                print(f"Failed to send redemption SMS to recipient: {sms_result.get('error')}")
        
        # Send email receipt to recipient
        if recipient and recipient.email:
            email_result = email_service.send_redemption_receipt_email(
                recipient.email,
                f"{recipient.first_name} {recipient.last_name}",
                voucher_code,
                float(voucher.value),
                0.0,  # Remaining balance (0 for full redemption)
                user.shop_name or 'Local Food Shop'
            )
            if not email_result:
                print(f"Failed to send redemption email to {recipient.email}")
        
        return jsonify({
            'message': 'Voucher redeemed successfully',
            'voucher': {
                'code': voucher.code,
                'value': float(voucher.value),
                'recipient': {
                    'name': f"{recipient.first_name} {recipient.last_name}" if recipient else 'N/A',
                    'phone': recipient.phone if recipient else 'N/A'
                } if recipient else None
            },
            'new_balance': float(user.balance)
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to redeem voucher: {str(e)}'}), 500


@app.route('/api/vendor/validate-voucher', methods=['POST'])
def validate_voucher():
    """Vendor endpoint to validate a voucher code without redeeming it"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vendor':
            return jsonify({'error': 'Vendor access required'}), 403
        
        data = request.get_json()
        voucher_code = data.get('code', '').strip().upper()
        
        if not voucher_code:
            return jsonify({'error': 'Voucher code is required'}), 400
        
        # Find voucher by code
        voucher = Voucher.query.filter_by(code=voucher_code).first()
        
        if not voucher:
            return jsonify({'valid': False, 'error': 'Invalid voucher code'}), 200
        
        # Check status
        if voucher.status != 'active':
            return jsonify({'valid': False, 'error': f'Voucher is {voucher.status}'}), 200
        
        # Check expiry
        from datetime import datetime
        if voucher.expiry_date and datetime.now().date() > voucher.expiry_date:
            return jsonify({'valid': False, 'error': 'Voucher has expired'}), 200
        
        # Check vendor restrictions (support multi-shop redemption)
        if voucher.vendor_restrictions:
            import json
            allowed_shop_ids = json.loads(voucher.vendor_restrictions)
            
            # Get all shops owned by this vendor
            vendor_shops = VendorShop.query.filter_by(vendor_id=user_id, is_active=True).all()
            vendor_shop_ids = [shop.id for shop in vendor_shops]
            
            # Check if any of vendor's shops are in the allowed list
            if not any(shop_id in allowed_shop_ids for shop_id in vendor_shop_ids):
                return jsonify({'valid': False, 'error': 'This voucher cannot be redeemed at your shops'}), 200
        
        # Get recipient details
        recipient = User.query.get(voucher.recipient_id) if voucher.recipient_id else None
        
        return jsonify({
            'valid': True,
            'voucher': {
                'code': voucher.code,
                'value': float(voucher.value),
                'expiry_date': voucher.expiry_date.isoformat() if voucher.expiry_date else None,
                'recipient': {
                    'name': f"{recipient.first_name} {recipient.last_name}" if recipient else 'N/A',
                    'phone': recipient.phone if recipient else 'N/A'
                } if recipient else None
            }
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to validate voucher: {str(e)}'}), 500


# ============================================
# Recipient Voucher Management Routes
# ============================================

@app.route('/api/recipient/vouchers', methods=['GET'])
def get_recipient_vouchers():
    """Get all vouchers assigned to the logged-in recipient"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'recipient':
            return jsonify({'error': 'Recipient access required'}), 403
        
        # Get all vouchers for this recipient
        vouchers = Voucher.query.filter_by(recipient_id=user_id).order_by(
            Voucher.created_at.desc()
        ).all()
        
        vouchers_data = []
        for voucher in vouchers:
            issued_by_user = User.query.get(voucher.issued_by)
            redeemed_vendor = User.query.get(voucher.redeemed_by_vendor) if voucher.redeemed_by_vendor else None
            
            vouchers_data.append({
                'id': voucher.id,
                'code': voucher.code,
                'value': float(voucher.value),
                'status': voucher.status,
                'expiry_date': voucher.expiry_date.isoformat() if voucher.expiry_date else None,
                'created_at': voucher.created_at.isoformat() if voucher.created_at else None,
                'redeemed_at': voucher.redeemed_at.isoformat() if voucher.redeemed_at else None,
                'issued_by': {
                    'name': issued_by_user.organization_name or f"{issued_by_user.first_name} {issued_by_user.last_name}",
                    'type': issued_by_user.user_type
                } if issued_by_user else None,
                'redeemed_by': {
                    'name': redeemed_vendor.shop_name or f"{redeemed_vendor.first_name} {redeemed_vendor.last_name}"
                } if redeemed_vendor else None,
                'vendor_restrictions': voucher.vendor_restrictions
            })
        
        # Calculate summary statistics
        active_vouchers = [v for v in vouchers if v.status == 'active']
        redeemed_vouchers = [v for v in vouchers if v.status == 'redeemed']
        expired_vouchers = [v for v in vouchers if v.status == 'expired']
        
        total_value = sum(float(v.value) for v in active_vouchers)
        total_redeemed = sum(float(v.value) for v in redeemed_vouchers)
        
        return jsonify({
            'vouchers': vouchers_data,
            'summary': {
                'total_vouchers': len(vouchers),
                'active_count': len(active_vouchers),
                'redeemed_count': len(redeemed_vouchers),
                'expired_count': len(expired_vouchers),
                'total_active_value': total_value,
                'total_redeemed_value': total_redeemed
            }
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get vouchers: {str(e)}'}), 500


@app.route('/api/recipient/shops', methods=['GET'])
def get_recipient_shops():
    """Get all participating shops with their to-go items count"""
    try:
        # Only show active shops
        shops = VendorShop.query.filter_by(is_active=True).all()
        shops_data = []
        
        for shop in shops:
            # Count available to-go items for this shop
            to_go_count = SurplusItem.query.filter_by(
                shop_id=shop.id,
                status='available'
            ).count()
            
            shops_data.append({
                'id': shop.id,
                'shop_name': shop.shop_name,
                'address': shop.address or '',
                'city': shop.city or '',
                'postcode': shop.postcode or '',
                'phone': shop.phone or '',
                'to_go_items_count': to_go_count
            })
        
        return jsonify({'shops': shops_data}), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get shops: {str(e)}'}), 500


@app.route('/api/recipient/to-go-items', methods=['GET'])
def get_recipient_to_go_items():
    """Get all available discounted to-go items from all shops (no free items)"""
    try:
        # Recipients should only see discounted items, not free items
        items = SurplusItem.query.filter_by(status='available', item_type='discount').all()
        items_data = []
        
        for item in items:
            # Get shop information
            shop = VendorShop.query.get(item.shop_id)
            
            items_data.append({
                'id': item.id,
                'item_name': item.item_name,
                'quantity': item.quantity,
                'category': item.category,
                'description': item.description or '',
                'item_type': item.item_type or 'free',
                'price': float(item.price) if item.price else 0.0,
                'original_price': float(item.original_price) if item.original_price else 0.0,
                'quantity_available': item.quantity or '0',
                'shop_name': shop.shop_name if shop else 'Unknown Shop',
                'shop_address': shop.address if shop else '',
                'shop_city': shop.city if shop else '',
                'shop_postcode': shop.postcode if shop else '',
                'posted_at': item.posted_at.isoformat() if item.posted_at else None
            })
        
        return jsonify({'items': items_data}), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get to-go items: {str(e)}'}), 500


@app.route('/api/recipient/vouchers/<int:voucher_id>/qr', methods=['GET'])
def get_voucher_qr_code(voucher_id):
    """Generate QR code for a voucher"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        voucher = Voucher.query.get(voucher_id)
        if not voucher or voucher.recipient_id != user_id:
            return jsonify({'error': 'Voucher not found'}), 404
        
        # Generate QR code data (voucher code + validation token)
        import hashlib
        import time
        timestamp = int(time.time())
        validation_token = hashlib.sha256(f"{voucher.code}{timestamp}{voucher.id}".encode()).hexdigest()[:16]
        
        qr_data = {
            'voucher_code': voucher.code,
            'voucher_id': voucher.id,
            'value': float(voucher.value),
            'token': validation_token,
            'timestamp': timestamp
        }
        
        return jsonify({
            'qr_data': qr_data,
            'qr_string': f"BAKUP-{voucher.code}-{validation_token}"
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to generate QR code: {str(e)}'}), 500


@app.route('/api/recipient/vouchers/<int:voucher_id>/pdf', methods=['GET'])
def recipient_voucher_pdf(voucher_id):
    """Generate printable PDF voucher with QR code for recipient"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.units import inch
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from io import BytesIO
        import qrcode
        from PIL import Image
        from flask import send_file
        
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        # Get voucher
        voucher = Voucher.query.get(voucher_id)
        if not voucher or voucher.recipient_id != user_id:
            return jsonify({'error': 'Voucher not found or access denied'}), 404
        
        recipient = User.query.get(voucher.recipient_id)
        
        # Create PDF
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        
        # Header - BAK UP branding
        c.setFillColor(colors.HexColor('#4CAF50'))
        c.rect(0, height - 2*inch, width, 2*inch, fill=True, stroke=False)
        c.setFillColor(colors.white)
        c.setFont('Helvetica-Bold', 32)
        c.drawCentredString(width/2, height - 1.2*inch, 'BAK UP E-Voucher')
        c.setFont('Helvetica', 14)
        c.drawCentredString(width/2, height - 1.6*inch, 'Supporting Families Through Education & Care')
        
        # Voucher Code - Large and prominent
        c.setFillColor(colors.black)
        c.setFont('Helvetica-Bold', 48)
        c.drawCentredString(width/2, height - 3*inch, voucher.code)
        c.setFont('Helvetica', 12)
        c.setFillColor(colors.grey)
        c.drawCentredString(width/2, height - 3.3*inch, 'Voucher Code')
        
        # Generate QR Code
        qr = qrcode.QRCode(version=1, box_size=10, border=2)
        qr.add_data(voucher.code)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")
        
        # Save QR code to BytesIO
        qr_buffer = BytesIO()
        qr_img.save(qr_buffer, format='PNG')
        qr_buffer.seek(0)
        
        # Draw QR code on PDF
        c.drawImage(qr_buffer, width/2 - 1.5*inch, height - 6*inch, width=3*inch, height=3*inch)
        
        # Instruction text above QR code
        c.setFont('Helvetica-Bold', 12)
        c.setFillColor(colors.HexColor('#4CAF50'))
        c.drawCentredString(width/2, height - 3.8*inch, 'Show this QR code at participating shops')
        
        # Voucher Details Box
        y_position = height - 7*inch
        c.setFillColor(colors.HexColor('#f5f5f5'))
        c.rect(1*inch, y_position - 2.5*inch, width - 2*inch, 2.5*inch, fill=True, stroke=True)
        
        c.setFillColor(colors.black)
        c.setFont('Helvetica-Bold', 14)
        c.drawString(1.2*inch, y_position - 0.4*inch, 'Voucher Details')
        
        c.setFont('Helvetica', 11)
        c.drawString(1.2*inch, y_position - 0.8*inch, f'Recipient: {recipient.first_name} {recipient.last_name}' if recipient else 'Recipient: Unknown')
        c.drawString(1.2*inch, y_position - 1.1*inch, f'Email: {recipient.email}' if recipient else '')
        c.drawString(1.2*inch, y_position - 1.4*inch, f'Phone: {recipient.phone}' if recipient else '')
        
        c.setFont('Helvetica-Bold', 16)
        c.setFillColor(colors.HexColor('#4CAF50'))
        c.drawString(1.2*inch, y_position - 1.9*inch, f'Value: £{float(voucher.value):.2f}')
        
        c.setFillColor(colors.black)
        c.setFont('Helvetica', 11)
        c.drawString(1.2*inch, y_position - 2.2*inch, f'Issue Date: {voucher.created_at.strftime("%d %B %Y") if voucher.created_at else ""}')
        c.drawString(1.2*inch, y_position - 2.5*inch, f'Expiry Date: {voucher.expiry_date.strftime("%d %B %Y") if voucher.expiry_date else ""}')
        
        # Status indicator
        status_color = colors.HexColor('#4CAF50') if voucher.status == 'active' else colors.HexColor('#f44336')
        c.setFillColor(status_color)
        c.setFont('Helvetica-Bold', 12)
        status_text = voucher.status.upper()
        c.drawString(width - 2.5*inch, y_position - 0.4*inch, f'Status: {status_text}')
        
        # How to Use Section
        y_position = height - 10*inch
        c.setFillColor(colors.black)
        c.setFont('Helvetica-Bold', 12)
        c.drawString(1*inch, y_position, 'How to Use Your Voucher:')
        c.setFont('Helvetica', 10)
        instructions = [
            '1. Print this voucher or show it on your mobile device',
            '2. Visit any participating local shop',
            '3. Show the QR code or voucher code to the shop staff',
            '4. Shop staff will scan the QR code or enter the voucher code',
            '5. Purchase food and essential items up to the voucher value'
        ]
        y_pos = y_position - 0.25*inch
        for instruction in instructions:
            c.drawString(1*inch, y_pos, instruction)
            y_pos -= 0.2*inch
        
        # Terms and Conditions
        y_pos -= 0.3*inch
        c.setFont('Helvetica-Bold', 10)
        c.drawString(1*inch, y_pos, 'Terms & Conditions:')
        c.setFont('Helvetica', 8)
        terms = [
            '• This voucher can be redeemed at participating local shops for food and essential items.',
            '• The voucher must be presented at the time of purchase.',
            '• The voucher cannot be exchanged for cash.',
            '• The voucher is valid until the expiry date shown above.',
            '• Any unused balance will be forfeited after expiry.',
            '• For assistance, contact your VCFSE organization or visit backup-voucher-system.onrender.com'
        ]
        y_pos -= 0.2*inch
        for term in terms:
            c.drawString(1*inch, y_pos, term)
            y_pos -= 0.15*inch
        
        # Footer
        c.setFont('Helvetica-Oblique', 8)
        c.setFillColor(colors.grey)
        c.drawCentredString(width/2, 0.5*inch, f'Generated on {datetime.utcnow().strftime("%d %B %Y at %H:%M UTC")} | BAK UP E-Voucher System')
        
        c.save()
        buffer.seek(0)
        
        # Send PDF
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'voucher_{voucher.code}.pdf'
        )
        
    except Exception as e:
        return jsonify({'error': f'Failed to generate PDF: {str(e)}'}), 500


@app.route('/api/recipient/vouchers/<int:voucher_id>/resend-sms', methods=['POST'])
def resend_voucher_sms(voucher_id):
    """Resend voucher code via SMS"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        voucher = Voucher.query.get(voucher_id)
        if not voucher or voucher.recipient_id != user_id:
            return jsonify({'error': 'Voucher not found'}), 404
        
        recipient = User.query.get(user_id)
        if not recipient.phone:
            return jsonify({'error': 'No phone number on file'}), 400
        
        # Send SMS with voucher code
        sms_result = sms_service.send_voucher_code(
            recipient.phone,
            voucher.code,
            recipient.first_name or recipient.name,
            float(voucher.value)
        )
        
        if not sms_result.get('success'):
            return jsonify({'error': sms_result.get('error', 'Failed to send SMS')}), 500
        
        return jsonify({
            'success': True,
            'message': 'SMS sent successfully'
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to send SMS: {str(e)}'}), 500


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        
        # Run automatic migration for date_of_birth field
        try:
            # Check if date_of_birth column exists
            result = db.session.execute(text("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name='user' AND column_name='date_of_birth'
            """))
            
            if not result.fetchone():
                print("Adding date_of_birth column to user table...")
                db.session.execute(text('ALTER TABLE "user" ADD COLUMN date_of_birth VARCHAR(50)'))
                db.session.commit()
                print("✅ date_of_birth column added successfully")
            else:
                print("✅ date_of_birth column already exists")
        except Exception as e:
            print(f"Migration check: {e}")
            db.session.rollback()
        
        # Create default categories if they don't exist
        if not Category.query.first():
            categories = [
                Category(name='Fresh Produce', type='edible', description='Fruits, vegetables, and fresh items'),
                Category(name='Bakery', type='edible', description='Bread, pastries, and baked goods'),
                Category(name='Dairy', type='edible', description='Milk, cheese, yogurt, and dairy products'),
                Category(name='Meat & Fish', type='edible', description='Fresh and frozen meat and fish'),
                Category(name='Packaged Foods', type='edible', description='Canned, boxed, and packaged items'),
                Category(name='Non-Food Items', type='non-edible', description='Toiletries, household items, and other essentials')
            ]
            db.session.add_all(categories)
            db.session.commit()
            print("Database initialized with default categories")

# ============================================
# Shopping Cart Routes
# ============================================

@app.route('/api/cart/add', methods=['POST'])
def add_to_cart():
    """Add item to shopping cart"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'recipient':
            return jsonify({'error': 'Recipient access required'}), 403
        
        data = request.get_json()
        item_id = data.get('item_id')
        quantity = data.get('quantity', 1)
        
        if not item_id:
            return jsonify({'error': 'Item ID is required'}), 400
        
        # Check if item exists and is available
        item = SurplusItem.query.get(item_id)
        if not item or item.status != 'available':
            return jsonify({'error': 'Item not available'}), 400
        
        # Add to cart or update quantity
        result = db.session.execute(
            text("SELECT id, quantity FROM shopping_cart WHERE recipient_id = :user_id AND surplus_item_id = :item_id"),
            {"user_id": user_id, "item_id": item_id}
        ).fetchone()
        
        if result:
            # Update existing cart item
            cart_id, current_qty = result
            new_qty = current_qty + quantity
            db.session.execute(
                text("UPDATE shopping_cart SET quantity = :qty WHERE id = :cart_id"),
                {"qty": new_qty, "cart_id": cart_id}
            )
        else:
            # Add new cart item
            db.session.execute(
                text("INSERT INTO shopping_cart (recipient_id, surplus_item_id, quantity) VALUES (:user_id, :item_id, :qty)"),
                {"user_id": user_id, "item_id": item_id, "qty": quantity}
            )
        
        # Notify vendor that someone added item to cart
        shop = VendorShop.query.get(item.shop_id)
        if shop:
            db.session.execute(
                text("INSERT INTO cart_notification (user_id, message, type, surplus_item_id) VALUES (:vendor_id, :message, :type, :item_id)"),
                {"vendor_id": shop.vendor_id, "message": f"{user.first_name} added {item.item_name} to their cart", "type": 'item_added_to_cart', "item_id": item_id}
            )
        
        db.session.commit()
        
        return jsonify({'message': 'Item added to cart successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to add to cart: {str(e)}'}), 500

@app.route('/api/cart', methods=['GET'])
def get_cart():
    """Get user's shopping cart"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'recipient':
            return jsonify({'error': 'Recipient access required'}), 403
        
        # Get cart items
        cart_items = db.session.execute(text("""
            SELECT c.id, c.quantity, c.added_at, s.id as item_id, s.item_name, s.quantity as item_quantity,
                   s.unit, s.category, s.price, s.description, s.status,
                   v.shop_name, v.address
            FROM shopping_cart c
            JOIN surplus_item s ON c.surplus_item_id = s.id
            JOIN vendor_shop v ON s.shop_id = v.id
            WHERE c.recipient_id = :user_id
            ORDER BY c.added_at DESC
        """), {"user_id": user_id}).fetchall()
        
        cart_data = []
        for item in cart_items:
            cart_data.append({
                'cart_id': item[0],
                'quantity': item[1],
                'added_at': item[2],
                'item': {
                    'id': item[3],
                    'name': item[4],
                    'available_quantity': item[5],
                    'unit': item[6],
                    'category': item[7],
                    'price': float(item[8]) if item[8] else 0.0,
                    'description': item[9],
                    'status': item[10]
                },
                'shop': {
                    'name': item[11],
                    'address': item[12]
                }
            })
        
        return jsonify({'cart': cart_data, 'count': len(cart_data)}), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get cart: {str(e)}'}), 500

@app.route('/api/cart/remove/<int:cart_id>', methods=['DELETE'])
def remove_from_cart(cart_id):
    """Remove item from cart"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        # Verify cart item belongs to user
        result = db.session.execute(
            text("SELECT recipient_id FROM shopping_cart WHERE id = :cart_id"),
            {"cart_id": cart_id}
        ).fetchone()
        
        if not result or result[0] != user_id:
            return jsonify({'error': 'Cart item not found'}), 404
        
        db.session.execute(text("DELETE FROM shopping_cart WHERE id = :cart_id"), {"cart_id": cart_id})
        db.session.commit()
        
        return jsonify({'message': 'Item removed from cart'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to remove from cart: {str(e)}'}), 500

@app.route('/api/cart/notifications', methods=['GET'])
def get_cart_notifications():
    """Get cart-related notifications"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        notifications = db.session.execute(text("""
            SELECT id, message, type, is_read, created_at
            FROM cart_notification
            WHERE user_id = :user_id
            ORDER BY created_at DESC
            LIMIT 50
        """), {"user_id": user_id}).fetchall()
        
        notif_data = []
        for n in notifications:
            notif_data.append({
                'id': n[0],
                'message': n[1],
                'type': n[2],
                'is_read': bool(n[3]),
                'created_at': n[4]
            })
        
        return jsonify({'notifications': notif_data}), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to get notifications: {str(e)}'}), 500


# ============================================
# Admin Edit and Delete Routes for Schools and VCFSE
# ============================================

@app.route('/api/admin/schools/<int:school_id>', methods=['PUT'])
def edit_school(school_id):
    """Admin edits a school/care organization"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Get the school
        school = User.query.get(school_id)
        if not school or school.user_type != 'school':
            return jsonify({'error': 'School not found'}), 404
        
        # Get update data
        data = request.get_json()
        
        # Update fields if provided
        if 'organization_name' in data:
            school.organization_name = data['organization_name']
        if 'first_name' in data:
            school.first_name = data['first_name']
        if 'last_name' in data:
            school.last_name = data['last_name']
        if 'email' in data:
            # Check if email is already taken by another user
            existing_user = User.query.filter_by(email=data['email']).first()
            if existing_user and existing_user.id != school_id:
                return jsonify({'error': 'Email already in use'}), 400
            school.email = data['email']
        if 'phone' in data:
            school.phone = data['phone']
        if 'address' in data:
            school.address = data['address']
        if 'city' in data:
            school.city = data['city']
        if 'postcode' in data:
            school.postcode = data['postcode']
        if 'allocated_balance' in data:
            school.allocated_balance = float(data['allocated_balance'])
        
        # Handle password reset if provided
        if 'new_password' in data and data['new_password']:
            from werkzeug.security import generate_password_hash
            school.password_hash = generate_password_hash(data['new_password'])
        
        db.session.commit()
        
        return jsonify({
            'message': 'School updated successfully',
            'school': {
                'id': school.id,
                'organization_name': school.organization_name,
                'first_name': school.first_name,
                'last_name': school.last_name,
                'email': school.email,
                'phone': school.phone,
                'address': school.address,
                'city': school.city,
                'postcode': school.postcode,
                'allocated_balance': float(school.allocated_balance) if school.allocated_balance else 0.0
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to update school: {str(e)}'}), 500


@app.route('/api/admin/schools/<int:school_id>', methods=['DELETE'])
def delete_school(school_id):
    """Admin deletes a school/care organization"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Get the school
        school = User.query.get(school_id)
        if not school or school.user_type != 'school':
            return jsonify({'error': 'School not found'}), 404
        
        # Check if school has issued any vouchers
        vouchers_count = Voucher.query.filter_by(issued_by=school_id).count()
        if vouchers_count > 0:
            return jsonify({
                'error': f'Cannot delete school. It has issued {vouchers_count} voucher(s). Please reassign or delete vouchers first.'
            }), 400
        
        # Delete associated records first to avoid foreign key constraints
        # Delete login sessions
        LoginSession.query.filter_by(user_id=school_id).delete()
        
        # Delete notifications
        Notification.query.filter_by(user_id=school_id).delete()
        
        # Delete the school
        db.session.delete(school)
        db.session.commit()
        
        return jsonify({
            'message': 'School deleted successfully',
            'school_name': school.organization_name
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to delete school: {str(e)}'}), 500


@app.route('/api/admin/vcse/<int:vcse_id>', methods=['PUT'])
def edit_vcse(vcse_id):
    """Admin edits a VCFSE organization"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Get the VCFSE organization
        vcse = User.query.get(vcse_id)
        if not vcse or vcse.user_type != 'vcse':
            return jsonify({'error': 'VCFSE organization not found'}), 404
        
        # Get update data
        data = request.get_json()
        
        # Update fields if provided
        if 'name' in data:
            vcse.organization_name = data['name']
        if 'email' in data:
            # Check if email is already taken by another user
            existing_user = User.query.filter_by(email=data['email']).first()
            if existing_user and existing_user.id != vcse_id:
                return jsonify({'error': 'Email already in use'}), 400
            vcse.email = data['email']
        if 'charity_commission_number' in data:
            vcse.charity_commission_number = data['charity_commission_number']
        if 'allocated_balance' in data:
            vcse.allocated_balance = float(data['allocated_balance'])
        if 'new_password' in data and data['new_password']:
            # Reset password if provided
            vcse.password_hash = generate_password_hash(data['new_password'])
        
        db.session.commit()
        
        return jsonify({
            'message': 'VCFSE organization updated successfully',
            'vcse': {
                'id': vcse.id,
                'name': vcse.organization_name,
                'email': vcse.email,
                'charity_commission_number': vcse.charity_commission_number,
                'allocated_balance': float(vcse.allocated_balance) if vcse.allocated_balance else 0.0
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to update VCFSE organization: {str(e)}'}), 500


@app.route('/api/admin/vcse/<int:vcse_id>', methods=['DELETE'])
def delete_vcse(vcse_id):
    """Admin deletes a VCFSE organization"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Get the VCFSE organization
        vcse = User.query.get(vcse_id)
        if not vcse or vcse.user_type != 'vcse':
            return jsonify({'error': 'VCFSE organization not found'}), 404
        
        # Check if VCFSE has issued any vouchers
        vouchers_count = Voucher.query.filter_by(issued_by=vcse_id).count()
        if vouchers_count > 0:
            return jsonify({
                'error': f'Cannot delete VCFSE organization. It has issued {vouchers_count} voucher(s). Please reassign or delete vouchers first.'
            }), 400
        
        # Delete associated records first to avoid foreign key constraints
        # Delete login sessions
        LoginSession.query.filter_by(user_id=vcse_id).delete()
        
        # Delete notifications
        Notification.query.filter_by(user_id=vcse_id).delete()
        
        # Delete the VCFSE organization
        db.session.delete(vcse)
        db.session.commit()
        
        return jsonify({
            'message': 'VCFSE organization deleted successfully',
            'organization_name': vcse.organization_name or vcse.first_name
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to delete VCFSE organization: {str(e)}'}), 500


@app.route('/api/admin/shops/<int:shop_id>', methods=['PUT'])
def admin_edit_shop(shop_id):
    """Admin edits a local shop"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Get the shop
        shop = VendorShop.query.get(shop_id)
        if not shop:
            return jsonify({'error': 'Shop not found'}), 404
        
        # Get update data
        data = request.get_json()
        
        # Update fields if provided
        if 'shop_name' in data:
            shop.shop_name = data['shop_name']
        if 'address' in data:
            shop.address = data['address']
        if 'city' in data:
            shop.city = data['city']
        if 'postcode' in data:
            shop.postcode = data['postcode']
        if 'phone' in data:
            shop.phone = data['phone']
        
        db.session.commit()
        
        return jsonify({
            'message': 'Shop updated successfully',
            'shop': {
                'id': shop.id,
                'shop_name': shop.shop_name,
                'address': shop.address,
                'city': shop.city,
                'postcode': shop.postcode,
                'phone': shop.phone
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to update shop: {str(e)}'}), 500


@app.route('/api/admin/shops/<int:shop_id>', methods=['DELETE'])
def admin_delete_shop(shop_id):
    """Admin deletes a local shop"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Unauthorized'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Get the shop
        shop = VendorShop.query.get(shop_id)
        if not shop:
            return jsonify({'error': 'Shop not found'}), 404
        
        # Check if shop has to-go items
        togo_count = SurplusItem.query.filter_by(shop_id=shop_id).count()
        if togo_count > 0:
            return jsonify({
                'error': f'Cannot delete shop. It has {togo_count} to-go item(s). Please remove items first.'
            }), 400
        
        shop_name = shop.shop_name
        
        # Delete the shop
        db.session.delete(shop)
        db.session.commit()
        
        return jsonify({
            'message': 'Shop deleted successfully',
            'shop_name': shop_name
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to delete shop: {str(e)}'}), 500


@app.route('/api/initialize-admin', methods=['POST'])
def initialize_admin():
    """One-time endpoint to create the admin account"""
    try:
        # Check if admin already exists
        existing_admin = User.query.filter_by(email='prince.caesar@bakup.org').first()
        if existing_admin:
            return jsonify({'message': 'Admin account already exists'}), 200
        
        # Create admin user
        admin = User(
            email='prince.caesar@bakup.org',
            password_hash=generate_password_hash('Prince@2024'),
            first_name='Prince',
            last_name='Caesar',
            user_type='admin',
            is_verified=True,
            is_active=True
        )
        
        db.session.add(admin)
        db.session.commit()
        
        return jsonify({
            'message': 'Admin account created successfully',
            'email': 'prince.caesar@bakup.org'
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to create admin: {str(e)}'}), 500


@app.route('/api/change-password', methods=['POST'])
def change_password():
    """Allow any authenticated user to change their password"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        data = request.get_json()
        current_password = data.get('current_password')
        new_password = data.get('new_password')
        
        if not current_password or not new_password:
            return jsonify({'error': 'Current password and new password are required'}), 400
        
        if len(new_password) < 8:
            return jsonify({'error': 'New password must be at least 8 characters long'}), 400
        
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Verify current password
        if not check_password_hash(user.password_hash, current_password):
            return jsonify({'error': 'Current password is incorrect'}), 401
        
        # Update password
        user.password_hash = generate_password_hash(new_password)
        db.session.commit()
        
        return jsonify({'message': 'Password changed successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to change password: {str(e)}'}), 500


@app.route('/api/admin/admins', methods=['GET'])
def get_all_admins():
    """Get list of all admin accounts (admin only)"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        admins = User.query.filter_by(user_type='admin', is_active=True).all()
        
        admin_list = []
        for admin in admins:
            admin_list.append({
                'id': admin.id,
                'email': admin.email,
                'first_name': admin.first_name,
                'last_name': admin.last_name,
                'created_at': admin.created_at.strftime('%Y-%m-%d %H:%M:%S') if admin.created_at else None,
                'last_login': admin.last_login.strftime('%Y-%m-%d %H:%M:%S') if admin.last_login else None,
                'login_count': admin.login_count
            })
        
        return jsonify({'admins': admin_list}), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to fetch admins: {str(e)}'}), 500


@app.route('/api/admin/admins', methods=['POST'])
def create_admin_account():
    """Create a new admin account (admin only)"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        data = request.get_json()
        email = data.get('email')
        password = data.get('password')
        first_name = data.get('first_name')
        last_name = data.get('last_name')
        
        if not all([email, password, first_name, last_name]):
            return jsonify({'error': 'All fields are required'}), 400
        
        if len(password) < 8:
            return jsonify({'error': 'Password must be at least 8 characters long'}), 400
        
        # Check if email already exists
        existing_user = User.query.filter_by(email=email).first()
        if existing_user:
            return jsonify({'error': 'Email already exists'}), 400
        
        # Create new admin
        new_admin = User(
            email=email,
            password_hash=generate_password_hash(password),
            first_name=first_name,
            last_name=last_name,
            user_type='admin',
            is_verified=True,
            is_active=True
        )
        
        db.session.add(new_admin)
        db.session.commit()
        
        return jsonify({
            'message': 'Admin account created successfully',
            'admin': {
                'id': new_admin.id,
                'email': new_admin.email,
                'first_name': new_admin.first_name,
                'last_name': new_admin.last_name
            }
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to create admin: {str(e)}'}), 500


@app.route('/api/admin/admins/<int:admin_id>', methods=['DELETE'])
def delete_admin_account(admin_id):
    """Delete an admin account (admin only, cannot delete self)"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        
        # Prevent self-deletion
        if user_id == admin_id:
            return jsonify({'error': 'Cannot delete your own admin account'}), 400
        
        admin_to_delete = User.query.get(admin_id)
        if not admin_to_delete or admin_to_delete.user_type != 'admin':
            return jsonify({'error': 'Admin account not found'}), 404
        
        # Check if this is the last admin
        admin_count = User.query.filter_by(user_type='admin', is_active=True).count()
        if admin_count <= 1:
            return jsonify({'error': 'Cannot delete the last admin account'}), 400
        
        # Delete associated records
        LoginSession.query.filter_by(user_id=admin_id).delete()
        Notification.query.filter_by(user_id=admin_id).delete()
        
        # Delete admin
        db.session.delete(admin_to_delete)
        db.session.commit()
        
        return jsonify({
            'message': 'Admin account deleted successfully',
            'email': admin_to_delete.email
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to delete admin: {str(e)}'}), 500
    
@app.route('/api/voucher/reassign', methods=['POST'])
def reassign_voucher():
    """Reassign a voucher to a different recipient (max 3 times)"""
    try:
        data = request.json
        voucher_id = data.get('voucher_id')
        new_recipient_email = data.get('new_recipient_email')
        reason = data.get('reason', '')
        
        # Get voucher
        voucher = Voucher.query.get(voucher_id)
        if not voucher:
            return jsonify({'error': 'Voucher not found'}), 404
        
        # Check if voucher can be reassigned
        if voucher.status != 'active':
            return jsonify({'error': f'Cannot reassign {voucher.status} voucher'}), 400
        
        if voucher.reassignment_count >= 3:
            return jsonify({'error': 'Voucher has reached maximum reassignment limit (3)'}), 400
        
        # Get new recipient
        new_recipient = User.query.filter_by(email=new_recipient_email, user_type='recipient').first()
        if not new_recipient:
            return jsonify({'error': 'Recipient not found'}), 404
        
        # Store original recipient on first reassignment
        if voucher.reassignment_count == 0:
            voucher.original_recipient_id = voucher.recipient_id
        
        # Update reassignment history
        import json
        history = json.loads(voucher.reassignment_history) if voucher.reassignment_history else []
        history.append({
            'from_recipient_id': voucher.recipient_id,
            'from_recipient_name': voucher.recipient.first_name + ' ' + voucher.recipient.last_name if voucher.recipient else 'N/A',
            'to_recipient_id': new_recipient.id,
            'to_recipient_name': new_recipient.first_name + ' ' + new_recipient.last_name,
            'reason': reason,
            'reassigned_at': datetime.utcnow().isoformat()
        })
        voucher.reassignment_history = json.dumps(history)
        
        # Update voucher
        voucher.recipient_id = new_recipient.id
        voucher.reassignment_count += 1
        
        db.session.commit()
        
        return jsonify({
            'message': 'Voucher reassigned successfully',
            'voucher_code': voucher.code,
            'new_recipient': new_recipient.first_name + ' ' + new_recipient.last_name,
            'reassignment_count': voucher.reassignment_count
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/vendor/shops/all', methods=['GET'])
def get_all_vendor_shops():
    """Get all vendor shops for dropdown selection"""
    try:
        shops = VendorShop.query.filter_by(is_active=True).all()
        return jsonify([
            {
                'id': shop.id,
                'shop_name': shop.shop_name,
                'vendor_name': shop.vendor.first_name + ' ' + shop.vendor.last_name,
                'vendor_id': shop.vendor_id,
                'address': shop.address,
                'city': shop.city
            }
            for shop in shops
        ]), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/login-stats', methods=['GET'])
def get_login_stats():
    """Get login frequency statistics for all users"""
    try:
        users = User.query.filter(User.user_type != 'admin').all()
        
        stats = []
        for user in users:
            stats.append({
                'id': user.id,
                'name': user.first_name + ' ' + user.last_name,
                'email': user.email,
                'user_type': user.user_type,
                'organization': user.organization_name or user.shop_name or 'N/A',
                'last_login': user.last_login.isoformat() if user.last_login else 'Never',
                'login_count': user.login_count,
                'is_active': user.is_active,
                'created_at': user.created_at.isoformat()
            })
        
        # Calculate summary statistics
        total_users = len(stats)
        active_users = sum(1 for s in stats if s['is_active'])
        logged_in_users = sum(1 for s in stats if s['last_login'] != 'Never')
        
        return jsonify({
            'users': stats,
            'summary': {
                'total_users': total_users,
                'active_users': active_users,
                'logged_in_users': logged_in_users,
                'never_logged_in': total_users - logged_in_users
            }
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== PAYOUT REQUEST ROUTES ====================

@app.route('/api/vendor/payout/request', methods=['POST'])
def request_payout():
    """Vendor requests a payout for redeemed vouchers"""
    try:
        data = request.json
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vendor':
            return jsonify({'error': 'Unauthorized'}), 403
        
        # Validate required fields
        required_fields = ['shop_id', 'amount', 'bank_name', 'account_number', 'sort_code', 'account_holder_name']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Verify shop belongs to vendor
        shop = VendorShop.query.get(data['shop_id'])
        if not shop or shop.vendor_id != user_id:
            return jsonify({'error': 'Invalid shop'}), 400
        
        # Create payout request
        payout = PayoutRequest(
            vendor_id=user_id,
            shop_id=data['shop_id'],
            amount=float(data['amount']),
            bank_name=data['bank_name'],
            account_number=data['account_number'],
            sort_code=data['sort_code'],
            account_holder_name=data['account_holder_name'],
            notes=data.get('notes', '')
        )
        
        db.session.add(payout)
        db.session.commit()
        
        # Send email notification to admin
        try:
            email_service.send_payout_request_notification(
                vendor_name=f"{user.first_name} {user.last_name}",
                shop_name=shop.shop_name,
                amount=payout.amount
            )
        except Exception as e:
            print(f"Failed to send payout request email: {e}")
        
        return jsonify({
            'message': 'Payout request submitted successfully',
            'payout_id': payout.id
        }), 201
        
    except Exception as e:
        import traceback
        db.session.rollback()
        error_trace = traceback.format_exc()
        print(f"\n{'='*80}")
        print(f"ERROR in payout request endpoint:")
        print(error_trace)
        print(f"{'='*80}\n")
        return jsonify({'error': str(e)}), 500

@app.route('/api/vendor/payout/history', methods=['GET'])
def get_payout_history():
    """Get payout request history for vendor"""
    try:
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'vendor':
            return jsonify({'error': 'Unauthorized'}), 403
        
        payouts = PayoutRequest.query.filter_by(vendor_id=user_id).order_by(PayoutRequest.requested_at.desc()).all()
        
        payout_list = []
        for payout in payouts:
            payout_list.append({
                'id': payout.id,
                'shop_name': payout.shop.shop_name,
                'amount': payout.amount,
                'status': payout.status,
                'requested_at': payout.requested_at.isoformat(),
                'reviewed_at': payout.reviewed_at.isoformat() if payout.reviewed_at else None,
                'paid_at': payout.paid_at.isoformat() if payout.paid_at else None,
                'notes': payout.notes,
                'admin_notes': payout.admin_notes
            })
        
        return jsonify({'payouts': payout_list}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/payout/requests', methods=['GET'])
def get_all_payout_requests():
    """Admin: Get all payout requests"""
    try:
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Unauthorized'}), 403
        
        # Get filter parameters
        status_filter = request.args.get('status', 'all')
        
        query = PayoutRequest.query
        
        if status_filter != 'all':
            query = query.filter_by(status=status_filter)
        
        payouts = query.order_by(PayoutRequest.requested_at.desc()).all()
        
        payout_list = []
        for payout in payouts:
            payout_list.append({
                'id': payout.id,
                'vendor_name': f"{payout.vendor.first_name} {payout.vendor.last_name}",
                'vendor_email': payout.vendor.email,
                'vendor_phone': payout.vendor.phone,
                'shop_name': payout.shop.shop_name,
                'amount': payout.amount,
                'status': payout.status,
                'bank_name': payout.bank_name,
                'account_number': payout.account_number,
                'sort_code': payout.sort_code,
                'account_holder_name': payout.account_holder_name,
                'requested_at': payout.requested_at.isoformat(),
                'reviewed_at': payout.reviewed_at.isoformat() if payout.reviewed_at else None,
                'paid_at': payout.paid_at.isoformat() if payout.paid_at else None,
                'notes': payout.notes,
                'admin_notes': payout.admin_notes
            })
        
        # Calculate summary
        summary = {
            'total_requests': len(payouts),
            'pending': sum(1 for p in payouts if p.status == 'pending'),
            'approved': sum(1 for p in payouts if p.status == 'approved'),
            'rejected': sum(1 for p in payouts if p.status == 'rejected'),
            'paid': sum(1 for p in payouts if p.status == 'paid'),
            'total_amount_pending': sum(p.amount for p in payouts if p.status == 'pending'),
            'total_amount_approved': sum(p.amount for p in payouts if p.status == 'approved'),
            'total_amount_paid': sum(p.amount for p in payouts if p.status == 'paid')
        }
        
        return jsonify({
            'payouts': payout_list,
            'summary': summary
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/payout/<int:payout_id>/review', methods=['POST'])
def review_payout(payout_id):
    """Admin: Approve or reject payout request"""
    try:
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Unauthorized'}), 403
        
        data = request.json
        action = data.get('action')  # 'approve' or 'reject'
        admin_notes = data.get('admin_notes', '')
        
        if action not in ['approve', 'reject']:
            return jsonify({'error': 'Invalid action'}), 400
        
        payout = PayoutRequest.query.get(payout_id)
        if not payout:
            return jsonify({'error': 'Payout request not found'}), 404
        
        if payout.status != 'pending':
            return jsonify({'error': 'Payout request already reviewed'}), 400
        
        # Update payout status
        payout.status = 'approved' if action == 'approve' else 'rejected'
        payout.reviewed_by = user_id
        payout.reviewed_at = datetime.utcnow()
        payout.admin_notes = admin_notes
        
        db.session.commit()
        
        # Send email notification to vendor
        try:
            email_service.send_payout_status_notification(
                vendor_email=payout.vendor.email,
                vendor_name=f"{payout.vendor.first_name} {payout.vendor.last_name}",
                shop_name=payout.shop.shop_name,
                amount=payout.amount,
                status=payout.status,
                admin_notes=admin_notes
            )
        except Exception as e:
            print(f"Failed to send payout status email: {e}")
        
        return jsonify({
            'message': f'Payout request {action}d successfully',
            'payout': {
                'id': payout.id,
                'status': payout.status,
                'reviewed_at': payout.reviewed_at.isoformat()
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/payout/<int:payout_id>/mark-paid', methods=['POST'])
def mark_payout_paid(payout_id):
    """Admin: Mark approved payout as paid"""
    try:
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'admin':
            return jsonify({'error': 'Unauthorized'}), 403
        
        payout = PayoutRequest.query.get(payout_id)
        if not payout:
            return jsonify({'error': 'Payout request not found'}), 404
        
        if payout.status != 'approved':
            return jsonify({'error': 'Can only mark approved payouts as paid'}), 400
        
        payout.status = 'paid'
        payout.paid_at = datetime.utcnow()
        
        db.session.commit()
        
        # Send payment confirmation email
        try:
            email_service.send_payout_paid_notification(
                vendor_email=payout.vendor.email,
                vendor_name=f"{payout.vendor.first_name} {payout.vendor.last_name}",
                shop_name=payout.shop.shop_name,
                amount=payout.amount
            )
        except Exception as e:
            print(f"Failed to send payment confirmation email: {e}")
        
        return jsonify({
            'message': 'Payout marked as paid successfully',
            'payout': {
                'id': payout.id,
                'status': payout.status,
                'paid_at': payout.paid_at.isoformat()
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/admin/run-migration', methods=['POST'])
def run_migration():
    """Run database migrations (admin only)"""
    user_id = session.get('user_id')
    if not user_id:
        return jsonify({'error': 'Unauthorized'}), 401
    
    user = User.query.get(user_id)
    if not user or user.user_type != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        # Check if payout_request table exists
        from sqlalchemy import inspect, text
        inspector = inspect(db.engine)
        tables = inspector.get_table_names()
        
        if 'payout_request' not in tables:
            # Create the table using raw SQL
            create_table_sql = text("""
            CREATE TABLE IF NOT EXISTS payout_request (
                id SERIAL PRIMARY KEY,
                vendor_id INTEGER NOT NULL REFERENCES "user"(id),
                shop_id INTEGER NOT NULL REFERENCES vendor_shop(id),
                amount FLOAT NOT NULL,
                status VARCHAR(20) DEFAULT 'pending',
                bank_name VARCHAR(100),
                account_number VARCHAR(50),
                sort_code VARCHAR(20),
                account_holder_name VARCHAR(100),
                notes TEXT,
                admin_notes TEXT,
                requested_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                reviewed_at TIMESTAMP,
                reviewed_by INTEGER REFERENCES "user"(id),
                paid_at TIMESTAMP
            );
            """)
            db.session.execute(create_table_sql)
            
            # Create indexes
            db.session.execute(text("CREATE INDEX IF NOT EXISTS idx_payout_vendor ON payout_request(vendor_id);"))
            db.session.execute(text("CREATE INDEX IF NOT EXISTS idx_payout_shop ON payout_request(shop_id);"))
            db.session.execute(text("CREATE INDEX IF NOT EXISTS idx_payout_status ON payout_request(status);"))
            
            db.session.commit()
            return jsonify({'message': 'Migration completed: payout_request table created'}), 200
        else:
            return jsonify({'message': 'Migration not needed: payout_request table already exists'}), 200
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Migration failed: {str(e)}'}), 500

@app.route('/api/admin/update-user-type/<int:user_id>', methods=['POST'])
def update_user_type(user_id):
    """Update a user's type (for fixing registration issues)"""
    admin_id = session.get('user_id')
    if not admin_id:
        return jsonify({'error': 'Not authenticated'}), 401
    
    admin_user = User.query.get(admin_id)
    if not admin_user or admin_user.user_type != 'admin':
        return jsonify({'error': 'Admin access required'}), 403
    
    data = request.get_json()
    new_user_type = data.get('user_type')
    
    if new_user_type not in ['recipient', 'vendor', 'vcse', 'school', 'admin']:
        return jsonify({'error': 'Invalid user type'}), 400
    
    user = User.query.get(user_id)
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    old_type = user.user_type
    user.user_type = new_user_type
    db.session.commit()
    
    return jsonify({
        'message': f'User type updated from {old_type} to {new_user_type}',
        'user_id': user_id,
        'email': user.email,
        'old_type': old_type,
        'new_type': new_user_type
    })

# ============================================
# Food To Go - Shop Selection Endpoints
# ============================================

@app.route('/api/recipient/voucher-shop-status/<code>', methods=['GET'])
def get_voucher_shop_status(code):
    """Check if voucher requires shop selection and return shop assignment status"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        voucher = Voucher.query.filter_by(code=code).first()
        if not voucher:
            return jsonify({'error': 'Voucher not found'}), 404
        
        # Check if this voucher belongs to the logged-in user
        if voucher.recipient_id != user_id:
            return jsonify({'error': 'Unauthorized'}), 403
        
        response_data = {
            'assign_shop_method': voucher.assign_shop_method or 'specific_shop',
            'requires_selection': voucher.assign_shop_method == 'recipient_to_choose' and not voucher.recipient_selected_shop_id,
            'assigned_shop': None,
            'recipient_selected_shop': None
        }
        
        # If shop was preselected by VCFSE/School (specific_shop method)
        if voucher.assign_shop_method == 'specific_shop':
            # Get shop from vendor_restrictions (legacy) or assigned shop
            if voucher.vendor_restrictions:
                import json
                try:
                    vendor_ids = json.loads(voucher.vendor_restrictions)
                    if vendor_ids and len(vendor_ids) > 0:
                        # Get the first shop from the first vendor
                        shop = VendorShop.query.filter_by(vendor_id=vendor_ids[0], is_active=True).first()
                        if shop:
                            response_data['assigned_shop'] = {
                                'id': shop.id,
                                'shop_name': shop.shop_name,
                                'address': shop.address,
                                'town': shop.town,
                                'phone': shop.phone
                            }
                except:
                    pass
        
        # If recipient has selected a shop
        if voucher.recipient_selected_shop_id:
            shop = VendorShop.query.get(voucher.recipient_selected_shop_id)
            if shop:
                response_data['recipient_selected_shop'] = {
                    'id': shop.id,
                    'shop_name': shop.shop_name,
                    'address': shop.address,
                    'town': shop.town,
                    'phone': shop.phone
                }
        
        return jsonify(response_data), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/recipient/select-shop', methods=['POST'])
def select_shop_for_voucher():
    """Save recipient's shop selection and link it to their voucher"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        data = request.get_json()
        voucher_code = data.get('voucher_code')
        shop_id = data.get('shop_id')
        
        if not voucher_code or not shop_id:
            return jsonify({'error': 'Voucher code and shop ID are required'}), 400
        
        # Get the voucher
        voucher = Voucher.query.filter_by(code=voucher_code).first()
        if not voucher:
            return jsonify({'error': 'Voucher not found'}), 404
        
        # Check if this voucher belongs to the logged-in user
        if voucher.recipient_id != user_id:
            return jsonify({'error': 'Unauthorized'}), 403
        
        # Check if voucher allows recipient to choose
        if voucher.assign_shop_method != 'recipient_to_choose':
            return jsonify({'error': 'This voucher does not allow shop selection'}), 400
        
        # Check if shop has already been selected
        if voucher.recipient_selected_shop_id:
            return jsonify({'error': 'Shop has already been selected for this voucher'}), 400
        
        # Verify shop exists and is active
        shop = VendorShop.query.get(shop_id)
        if not shop or not shop.is_active:
            return jsonify({'error': 'Shop not found or inactive'}), 404
        
        # Save shop selection to voucher
        voucher.recipient_selected_shop_id = shop_id
        
        # Also save to recipient's preferred shop
        user = User.query.get(user_id)
        user.preferred_shop_id = shop_id
        
        db.session.commit()
        
        return jsonify({
            'message': 'Shop selected successfully',
            'shop': {
                'id': shop.id,
                'shop_name': shop.shop_name,
                'address': shop.address,
                'town': shop.town,
                'phone': shop.phone
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/recipient/discounted-items', methods=['GET'])
def get_discounted_items_for_recipient():
    """Get discounted items from recipient's assigned/selected shop"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.user_type != 'recipient':
            return jsonify({'error': 'Recipient access required'}), 403
        
        # Get recipient's shop (from preferred_shop_id or from their voucher)
        shop_id = user.preferred_shop_id
        
        # If no preferred shop, try to get from their active voucher
        if not shop_id:
            voucher = Voucher.query.filter_by(recipient_id=user_id, status='active').first()
            if voucher and voucher.recipient_selected_shop_id:
                shop_id = voucher.recipient_selected_shop_id
        
        if not shop_id:
            return jsonify({
                'items': [],
                'message': 'No shop assigned yet. Please select a shop first.'
            }), 200
        
        # Get all discounted items from the shop
        items = SurplusItem.query.filter_by(
            shop_id=shop_id,
            status='available',
            item_type='discount'
        ).all()
        
        items_data = []
        for item in items:
            shop = VendorShop.query.get(item.shop_id)
            items_data.append({
                'id': item.id,
                'item_name': item.item_name,
                'description': item.description,
                'quantity': item.quantity,
                'unit': item.unit,
                'price': float(item.price) if item.price else 0,
                'original_price': float(item.original_price) if item.original_price else 0,
                'savings': float(item.original_price - item.price) if (item.original_price and item.price) else 0,
                'shop_name': shop.shop_name if shop else 'Unknown',
                'shop_address': shop.address if shop else '',
                'shop_town': shop.town if shop else '',
                'available_until': item.available_until.isoformat() if item.available_until else None
            })
        
        return jsonify({
            'items': items_data,
            'total_count': len(items_data)
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Add Food To Go migration endpoint
from add_food_to_go_endpoint import create_food_to_go_migration_endpoint
create_food_to_go_migration_endpoint(app, db)

# Add Admin Password Reset endpoint
from admin_password_reset import create_admin_password_reset_endpoint
create_admin_password_reset_endpoint(app, db, User, generate_password_hash, session, request, jsonify)

# Add date_of_birth field migration
from add_date_of_birth_field import create_date_of_birth_migration_endpoint
create_date_of_birth_migration_endpoint(app, db)

@app.route('/api/admin/run-wallet-migration', methods=['POST'])
def run_wallet_migration():
    """Run wallet system database migration"""
    # Check for secret key in request
    secret_key = request.json.get('secret_key') if request.json else request.args.get('secret_key')
    
    MIGRATION_SECRET_KEY = os.environ.get('MIGRATION_SECRET_KEY', 'Food_To_Go_2024_Migration_Key')
    
    if secret_key != MIGRATION_SECRET_KEY:
        return jsonify({'error': 'Invalid secret key'}), 403
    
    try:
        from sqlalchemy import inspect, text
        inspector = inspect(db.engine)
        tables = inspector.get_table_names()
        
        results = []
        
        # 1. Create wallet_transaction table if it doesn't exist
        if 'wallet_transaction' not in tables:
            results.append("Creating wallet_transaction table...")
            create_table_sql = text("""
                CREATE TABLE IF NOT EXISTS wallet_transaction (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    user_id INT NOT NULL,
                    transaction_type VARCHAR(20) NOT NULL,
                    amount DECIMAL(10, 2) NOT NULL,
                    balance_before DECIMAL(10, 2) NOT NULL,
                    balance_after DECIMAL(10, 2) NOT NULL,
                    description TEXT,
                    reference VARCHAR(100),
                    payment_method VARCHAR(50),
                    payment_reference VARCHAR(100),
                    status VARCHAR(20) DEFAULT 'completed',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    created_by INT,
                    
                    FOREIGN KEY (user_id) REFERENCES `user`(id) ON DELETE CASCADE,
                    FOREIGN KEY (created_by) REFERENCES `user`(id) ON DELETE SET NULL,
                    
                    INDEX idx_user_id (user_id),
                    INDEX idx_transaction_type (transaction_type),
                    INDEX idx_created_at (created_at),
                    INDEX idx_status (status)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            """)
            db.session.execute(create_table_sql)
            results.append("✓ wallet_transaction table created")
        else:
            results.append("✓ wallet_transaction table already exists")
        
        # 2. Add columns to voucher table
        voucher_columns = inspector.get_columns('voucher')
        existing_columns = [col['name'] for col in voucher_columns]
        
        if 'issued_by_user_id' not in existing_columns:
            results.append("Adding issued_by_user_id column to voucher table...")
            db.session.execute(text("ALTER TABLE voucher ADD COLUMN issued_by_user_id INT"))
            results.append("✓ issued_by_user_id column added")
        else:
            results.append("✓ issued_by_user_id column already exists")
        
        if 'deducted_from_wallet' not in existing_columns:
            results.append("Adding deducted_from_wallet column to voucher table...")
            db.session.execute(text("ALTER TABLE voucher ADD COLUMN deducted_from_wallet BOOLEAN DEFAULT FALSE"))
            results.append("✓ deducted_from_wallet column added")
        else:
            results.append("✓ deducted_from_wallet column already exists")
        
        if 'wallet_transaction_id' not in existing_columns:
            results.append("Adding wallet_transaction_id column to voucher table...")
            db.session.execute(text("ALTER TABLE voucher ADD COLUMN wallet_transaction_id INT"))
            results.append("✓ wallet_transaction_id column added")
        else:
            results.append("✓ wallet_transaction_id column already exists")
        
        # 3. Add foreign key constraints (if they don't exist)
        try:
            results.append("Adding foreign key constraints...")
            db.session.execute(text("""
                ALTER TABLE voucher 
                ADD CONSTRAINT fk_voucher_issued_by_user 
                FOREIGN KEY (issued_by_user_id) REFERENCES "user"(id) ON DELETE SET NULL
            """))
            results.append("✓ Foreign key fk_voucher_issued_by_user added")
        except Exception as e:
            if "Duplicate foreign key" in str(e) or "already exists" in str(e):
                results.append("✓ Foreign key fk_voucher_issued_by_user already exists")
            else:
                results.append(f"⚠ Foreign key fk_voucher_issued_by_user: {str(e)}")
        
        try:
            db.session.execute(text("""
                ALTER TABLE voucher 
                ADD CONSTRAINT fk_voucher_wallet_transaction 
                FOREIGN KEY (wallet_transaction_id) REFERENCES wallet_transaction(id) ON DELETE SET NULL
            """))
            results.append("✓ Foreign key fk_voucher_wallet_transaction added")
        except Exception as e:
            if "Duplicate foreign key" in str(e) or "already exists" in str(e):
                results.append("✓ Foreign key fk_voucher_wallet_transaction already exists")
            else:
                results.append(f"⚠ Foreign key fk_voucher_wallet_transaction: {str(e)}")
        
        # 4. Add indexes
        try:
            db.session.execute(text("CREATE INDEX IF NOT EXISTS idx_issued_by_user_id ON voucher (issued_by_user_id)"))
            results.append("✓ Index idx_issued_by_user_id added")
        except Exception as e:
            if "Duplicate key name" in str(e):
                results.append("✓ Index idx_issued_by_user_id already exists")
            else:
                results.append(f"⚠ Index idx_issued_by_user_id: {str(e)}")
        
        try:
            db.session.execute(text("CREATE INDEX IF NOT EXISTS idx_wallet_transaction_id ON voucher (wallet_transaction_id)"))
            results.append("✓ Index idx_wallet_transaction_id added")
        except Exception as e:
            if "Duplicate key name" in str(e):
                results.append("✓ Index idx_wallet_transaction_id already exists")
            else:
                results.append(f"⚠ Index idx_wallet_transaction_id: {str(e)}")
        
        try:
            db.session.execute(text("CREATE INDEX IF NOT EXISTS idx_deducted_from_wallet ON voucher (deducted_from_wallet)"))
            results.append("✓ Index idx_deducted_from_wallet added")
        except Exception as e:
            if "Duplicate key name" in str(e):
                results.append("✓ Index idx_deducted_from_wallet already exists")
            else:
                results.append(f"⚠ Index idx_deducted_from_wallet: {str(e)}")
        
        db.session.commit()
        results.append("\n✅ Wallet system migration completed successfully!")
        
        return jsonify({
            'message': 'Wallet system migration completed',
            'details': results
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'error': f'Migration failed: {str(e)}',
            'details': results
        }), 500

# Auto-migration on startup
def check_and_migrate_database():
    """Check and add missing database columns on startup"""
    with app.app_context():
        try:
            from sqlalchemy import inspect
            inspector = inspect(db.engine)
            
            # Check if redeemed_at_shop_id column exists in voucher table
            voucher_columns = [col['name'] for col in inspector.get_columns('voucher')]
            
            if 'redeemed_at_shop_id' not in voucher_columns:
                print("⚠ Missing column 'redeemed_at_shop_id' - adding now...")
                db.session.execute(text(
                    "ALTER TABLE voucher ADD COLUMN redeemed_at_shop_id INTEGER"
                ))
                db.session.commit()
                print("✓ Successfully added 'redeemed_at_shop_id' column")
            else:
                print("✓ Database schema is up to date")
                
        except Exception as e:
            print(f"⚠ Migration check failed: {str(e)}")
            # Don't crash the app if migration fails
            pass

# Run migration automatically when module is imported (for Gunicorn/production)
check_and_migrate_database()

if __name__ == '__main__':
    socketio.run(app, host='0.0.0.0', port=5000, debug=True)

