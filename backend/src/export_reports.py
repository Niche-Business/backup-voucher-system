"""
Report Export Functionality
Provides PDF and Excel export capabilities for system reports
"""

from flask import Blueprint, jsonify, session, request, send_file
from datetime import datetime
import io
import logging
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

export_bp = Blueprint('export', __name__)
logger = logging.getLogger(__name__)

# Global references
db = None
User = None
Voucher = None
ToGoItem = None
Transaction = None

def init_export_system(app_db, user_model, voucher_model, togo_model, transaction_model):
    """Initialize the export system with database models"""
    global db, User, Voucher, ToGoItem, Transaction
    db = app_db
    User = user_model
    Voucher = voucher_model
    ToGoItem = togo_model
    Transaction = transaction_model
    logger.info("Export system initialized")


def generate_pdf_report(report_type, data, filters=None):
    """
    Generate a PDF report
    
    Args:
        report_type: Type of report ('vouchers', 'users', 'transactions', 'togo', 'financial')
        data: List of data dictionaries to include in report
        filters: Dictionary of filters applied (for report header)
    
    Returns:
        BytesIO object containing the PDF
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#2E7D32'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    # Header
    elements.append(Paragraph("BAK UP E-Voucher System", title_style))
    
    # Report title
    report_titles = {
        'vouchers': 'Voucher Management Report',
        'users': 'User Management Report',
        'transactions': 'Transaction History Report',
        'togo': 'Food to Go Items Report',
        'financial': 'Financial Summary Report'
    }
    elements.append(Paragraph(report_titles.get(report_type, 'System Report'), styles['Heading2']))
    
    # Generation info
    generation_time = datetime.now().strftime('%d %B %Y at %H:%M')
    elements.append(Paragraph(f"Generated on {generation_time}", subtitle_style))
    
    # Filters info
    if filters:
        filter_text = "Filters: " + ", ".join([f"{k}: {v}" for k, v in filters.items() if v])
        elements.append(Paragraph(filter_text, subtitle_style))
    
    elements.append(Spacer(1, 0.3*inch))
    
    # Data table
    if report_type == 'vouchers':
        table_data = [['Code', 'Value', 'Status', 'Recipient', 'Issued By', 'Expiry Date']]
        for item in data:
            table_data.append([
                item.get('code', 'N/A'),
                f"£{item.get('value', 0):.2f}",
                item.get('status', 'N/A'),
                item.get('recipient_name', 'N/A'),
                item.get('issued_by_name', 'N/A'),
                item.get('expiry_date', 'N/A')
            ])
    
    elif report_type == 'users':
        table_data = [['Name', 'Email', 'User Type', 'Status', 'Registered']]
        for item in data:
            table_data.append([
                item.get('name', 'N/A'),
                item.get('email', 'N/A'),
                item.get('user_type', 'N/A'),
                item.get('status', 'Active'),
                item.get('created_at', 'N/A')
            ])
    
    elif report_type == 'transactions':
        table_data = [['Date', 'Type', 'Amount', 'From', 'To', 'Status']]
        for item in data:
            table_data.append([
                item.get('date', 'N/A'),
                item.get('type', 'N/A'),
                f"£{item.get('amount', 0):.2f}",
                item.get('from_user', 'N/A'),
                item.get('to_user', 'N/A'),
                item.get('status', 'N/A')
            ])
    
    elif report_type == 'togo':
        table_data = [['Item Name', 'Shop', 'Price', 'Discount', 'Status', 'Posted Date']]
        for item in data:
            table_data.append([
                item.get('name', 'N/A'),
                item.get('shop_name', 'N/A'),
                f"£{item.get('original_price', 0):.2f}",
                f"{item.get('discount_percentage', 0)}%",
                item.get('status', 'N/A'),
                item.get('posted_date', 'N/A')
            ])
    
    elif report_type == 'financial':
        table_data = [['Metric', 'Value']]
        for item in data:
            table_data.append([
                item.get('metric', 'N/A'),
                item.get('value', 'N/A')
            ])
    
    # Create table
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    elements.append(table)
    
    # Footer
    elements.append(Spacer(1, 0.5*inch))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    elements.append(Paragraph("BAK UP CIC | Enterprise Centre, Warth Park, Raunds NN9 6GR", footer_style))
    elements.append(Paragraph("Contact: prince@bakupcic.co.uk | 01933698347", footer_style))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_excel_report(report_type, data, filters=None):
    """
    Generate an Excel report
    
    Args:
        report_type: Type of report ('vouchers', 'users', 'transactions', 'togo', 'financial')
        data: List of data dictionaries to include in report
        filters: Dictionary of filters applied (for report header)
    
    Returns:
        BytesIO object containing the Excel file
    """
    buffer = io.BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    # Report titles
    report_titles = {
        'vouchers': 'Voucher Management Report',
        'users': 'User Management Report',
        'transactions': 'Transaction History Report',
        'togo': 'Food to Go Items Report',
        'financial': 'Financial Summary Report'
    }
    sheet.title = report_titles.get(report_type, 'Report')[:31]  # Excel sheet name limit
    
    # Styling
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16, color="2E7D32")
    subtitle_font = Font(size=10, color="666666")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    sheet.merge_cells('A1:F1')
    sheet['A1'] = "BAK UP E-Voucher System"
    sheet['A1'].font = title_font
    sheet['A1'].alignment = Alignment(horizontal='center')
    
    # Report type
    sheet.merge_cells('A2:F2')
    sheet['A2'] = report_titles.get(report_type, 'System Report')
    sheet['A2'].font = Font(bold=True, size=14)
    sheet['A2'].alignment = Alignment(horizontal='center')
    
    # Generation info
    sheet.merge_cells('A3:F3')
    generation_time = datetime.now().strftime('%d %B %Y at %H:%M')
    sheet['A3'] = f"Generated on {generation_time}"
    sheet['A3'].font = subtitle_font
    sheet['A3'].alignment = Alignment(horizontal='center')
    
    # Filters
    if filters:
        sheet.merge_cells('A4:F4')
        filter_text = "Filters: " + ", ".join([f"{k}: {v}" for k, v in filters.items() if v])
        sheet['A4'] = filter_text
        sheet['A4'].font = subtitle_font
        sheet['A4'].alignment = Alignment(horizontal='center')
        start_row = 6
    else:
        start_row = 5
    
    # Headers and data
    if report_type == 'vouchers':
        headers = ['Code', 'Value', 'Status', 'Recipient', 'Issued By', 'Expiry Date']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=start_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, item in enumerate(data, start_row + 1):
            sheet.cell(row=row_idx, column=1, value=item.get('code', 'N/A')).border = border
            sheet.cell(row=row_idx, column=2, value=f"£{item.get('value', 0):.2f}").border = border
            sheet.cell(row=row_idx, column=3, value=item.get('status', 'N/A')).border = border
            sheet.cell(row=row_idx, column=4, value=item.get('recipient_name', 'N/A')).border = border
            sheet.cell(row=row_idx, column=5, value=item.get('issued_by_name', 'N/A')).border = border
            sheet.cell(row=row_idx, column=6, value=item.get('expiry_date', 'N/A')).border = border
    
    elif report_type == 'users':
        headers = ['Name', 'Email', 'User Type', 'Status', 'Registered']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=start_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, item in enumerate(data, start_row + 1):
            sheet.cell(row=row_idx, column=1, value=item.get('name', 'N/A')).border = border
            sheet.cell(row=row_idx, column=2, value=item.get('email', 'N/A')).border = border
            sheet.cell(row=row_idx, column=3, value=item.get('user_type', 'N/A')).border = border
            sheet.cell(row=row_idx, column=4, value=item.get('status', 'Active')).border = border
            sheet.cell(row=row_idx, column=5, value=item.get('created_at', 'N/A')).border = border
    
    elif report_type == 'transactions':
        headers = ['Date', 'Type', 'Amount', 'From', 'To', 'Status']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=start_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, item in enumerate(data, start_row + 1):
            sheet.cell(row=row_idx, column=1, value=item.get('date', 'N/A')).border = border
            sheet.cell(row=row_idx, column=2, value=item.get('type', 'N/A')).border = border
            sheet.cell(row=row_idx, column=3, value=f"£{item.get('amount', 0):.2f}").border = border
            sheet.cell(row=row_idx, column=4, value=item.get('from_user', 'N/A')).border = border
            sheet.cell(row=row_idx, column=5, value=item.get('to_user', 'N/A')).border = border
            sheet.cell(row=row_idx, column=6, value=item.get('status', 'N/A')).border = border
    
    elif report_type == 'togo':
        headers = ['Item Name', 'Shop', 'Price', 'Discount', 'Status', 'Posted Date']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=start_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, item in enumerate(data, start_row + 1):
            sheet.cell(row=row_idx, column=1, value=item.get('name', 'N/A')).border = border
            sheet.cell(row=row_idx, column=2, value=item.get('shop_name', 'N/A')).border = border
            sheet.cell(row=row_idx, column=3, value=f"£{item.get('original_price', 0):.2f}").border = border
            sheet.cell(row=row_idx, column=4, value=f"{item.get('discount_percentage', 0)}%").border = border
            sheet.cell(row=row_idx, column=5, value=item.get('status', 'N/A')).border = border
            sheet.cell(row=row_idx, column=6, value=item.get('posted_date', 'N/A')).border = border
    
    elif report_type == 'financial':
        headers = ['Metric', 'Value']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=start_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, item in enumerate(data, start_row + 1):
            sheet.cell(row=row_idx, column=1, value=item.get('metric', 'N/A')).border = border
            sheet.cell(row=row_idx, column=2, value=item.get('value', 'N/A')).border = border
    
    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


@export_bp.route('/api/export/<report_type>/<format>', methods=['POST'])
def export_report(report_type, format):
    """
    Export a report in PDF or Excel format
    
    Args:
        report_type: Type of report ('vouchers', 'users', 'transactions', 'togo', 'financial')
        format: Export format ('pdf' or 'excel')
    """
    # Check authentication
    user_id = session.get('user_id')
    if not user_id:
        return jsonify({'error': 'Unauthorized'}), 401
    
    user = User.query.get(user_id)
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    # Check if user has permission (admin or relevant user type)
    if user.user_type not in ['admin', 'vcse', 'school']:
        return jsonify({'error': 'Forbidden - Insufficient permissions'}), 403
    
    try:
        # Get filters from request
        filters = request.json or {}
        
        # Fetch data based on report type
        data = []
        
        if report_type == 'vouchers':
            vouchers = Voucher.query.all()
            for v in vouchers:
                recipient = User.query.get(v.recipient_id) if v.recipient_id else None
                issued_by = User.query.get(v.issued_by) if v.issued_by else None
                data.append({
                    'code': v.code,
                    'value': float(v.value),
                    'status': v.status,
                    'recipient_name': f"{recipient.first_name} {recipient.last_name}" if recipient else 'N/A',
                    'issued_by_name': f"{issued_by.first_name} {issued_by.last_name}" if issued_by else 'N/A',
                    'expiry_date': v.expiry_date.strftime('%Y-%m-%d') if v.expiry_date else 'N/A'
                })
        
        elif report_type == 'users':
            users = User.query.all()
            for u in users:
                data.append({
                    'name': f"{u.first_name} {u.last_name}",
                    'email': u.email,
                    'user_type': u.user_type,
                    'status': 'Active',
                    'created_at': u.created_at.strftime('%Y-%m-%d') if hasattr(u, 'created_at') and u.created_at else 'N/A'
                })
        
        elif report_type == 'togo':
            items = ToGoItem.query.all()
            for item in items:
                shop = User.query.get(item.shop_id) if item.shop_id else None
                data.append({
                    'name': item.name,
                    'shop_name': shop.shop_name if shop and hasattr(shop, 'shop_name') else 'N/A',
                    'original_price': float(item.original_price),
                    'discount_percentage': float(item.discount_percentage),
                    'status': item.status,
                    'posted_date': item.posted_date.strftime('%Y-%m-%d') if item.posted_date else 'N/A'
                })
        
        # Generate report
        if format == 'pdf':
            buffer = generate_pdf_report(report_type, data, filters)
            filename = f"{report_type}_report_{datetime.now().strftime('%Y%m%d')}.pdf"
            mimetype = 'application/pdf'
        elif format == 'excel':
            buffer = generate_excel_report(report_type, data, filters)
            filename = f"{report_type}_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            return jsonify({'error': 'Invalid format. Use "pdf" or "excel"'}), 400
        
        return send_file(
            buffer,
            mimetype=mimetype,
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        logger.error(f"Error exporting {report_type} report as {format}: {str(e)}")
        return jsonify({'error': str(e)}), 500
